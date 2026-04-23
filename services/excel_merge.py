"""excel_merge — 「使用者主權」Excel 合併工具。

問題：原本各 service 用 `wb = Workbook()` → `wb.save(filepath)` 整檔覆蓋，
洗掉使用者在 Excel 上手改的欄位、註記、分類調整。

解法：shadow 機制 — DB 記住「上次 service 寫入的值」，重產時：
  - 讀現有檔，逐 cell 比對「現值 vs shadow」
  - 若 ≠ → 使用者改過 → 標為 user-owned → 跳過寫入
  - 若 == → DB 主權 → 用新值覆寫
  - 寫完更新 shadow，下次再比對

不需要使用者手動標欄位，不需要白名單，自動偵測。
"""
from __future__ import annotations

import json
import logging
import os
import sqlite3
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger("shanbot.excel_merge")

_DB_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "data", "shanbot.db")


def _get_conn():
    conn = sqlite3.connect(_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _ensure_table():
    conn = _get_conn()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS excel_cell_shadow (
            filepath TEXT NOT NULL,
            sheet_name TEXT NOT NULL,
            cell_addr TEXT NOT NULL,
            db_value TEXT,
            PRIMARY KEY (filepath, sheet_name, cell_addr)
        )
    """)
    conn.commit()
    conn.close()


def _serialize(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float, bool)):
        return json.dumps(v)
    return str(v)


def _load_shadow(filepath: str) -> dict[tuple[str, str], str]:
    _ensure_table()
    conn = _get_conn()
    rows = conn.execute(
        "SELECT sheet_name, cell_addr, db_value FROM excel_cell_shadow WHERE filepath=?",
        (filepath,),
    ).fetchall()
    conn.close()
    return {(r["sheet_name"], r["cell_addr"]): r["db_value"] for r in rows}


def _save_shadow(filepath: str, shadow: dict[tuple[str, str], str]):
    _ensure_table()
    conn = _get_conn()
    conn.execute("DELETE FROM excel_cell_shadow WHERE filepath=?", (filepath,))
    conn.executemany(
        "INSERT INTO excel_cell_shadow (filepath, sheet_name, cell_addr, db_value) "
        "VALUES (?, ?, ?, ?)",
        [(filepath, sn, ca, v) for (sn, ca), v in shadow.items()],
    )
    conn.commit()
    conn.close()


class MergeContext:
    """一次 Excel 重產期間，追蹤使用者主權 cell + 累積新 shadow。

    用法：
        ctx = MergeContext.open(filepath)
        wb = ctx.workbook
        ws = wb.active
        ctx.set_cell(ws, 1, 1, "標題")        # 自動偵測 user edit
        ctx.set_cell(ws, 2, 1, "DB 算的值")
        ctx.commit(filepath)                  # save + 更新 shadow
    """

    def __init__(self, workbook: Workbook, shadow: dict[tuple[str, str], str],
                 user_owned: set[tuple[str, str]]):
        self.workbook = workbook
        self._shadow = shadow
        self._user_owned = user_owned
        self._new_shadow: dict[tuple[str, str], str] = {}
        self._row_cursor: dict[str, int] = {}

    @classmethod
    def open(cls, filepath: str) -> "MergeContext":
        if os.path.exists(filepath):
            try:
                wb = load_workbook(filepath)
            except Exception as e:
                logger.warning(f"load_workbook fail ({filepath}), 改用空白：{e}")
                wb = Workbook()
                wb.remove(wb.active)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        shadow = _load_shadow(filepath)

        # 標出使用者主權 cell：兩種情況
        #   (a) shadow 有記錄但現值與 shadow 不同 → user 改過
        #   (b) shadow 沒記但現值非空 → user 新加的欄位/註記
        user_owned: set[tuple[str, str]] = set()
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    addr = cell.coordinate
                    key = (sheet_name, addr)
                    if key in shadow:
                        if _serialize(cell.value) != shadow[key]:
                            user_owned.add(key)
                    else:
                        user_owned.add(key)

        if user_owned:
            logger.info(f"[excel_merge] {os.path.basename(filepath)}: "
                        f"偵測 {len(user_owned)} 個使用者主權 cell（保留不覆寫）")

        return cls(wb, shadow, user_owned)

    def get_or_create_sheet(self, name: str) -> Worksheet:
        if name in self.workbook.sheetnames:
            return self.workbook[name]
        return self.workbook.create_sheet(name)

    def clear_sheet(self, ws: Worksheet, *, preserve_user_edits: bool = True):
        """清空 sheet 內容（user-owned cell 保留），並重置 append cursor=1。"""
        sheet_name = ws.title
        for row in ws.iter_rows():
            for cell in row:
                if preserve_user_edits and (sheet_name, cell.coordinate) in self._user_owned:
                    continue
                cell.value = None
        self._row_cursor[sheet_name] = 1

    def set_cell(self, ws: Worksheet, row: int, col: int, value):
        cell = ws.cell(row=row, column=col)
        addr = cell.coordinate
        sheet_name = ws.title
        if (sheet_name, addr) not in self._user_owned:
            cell.value = value
        self._new_shadow[(sheet_name, addr)] = _serialize(value)
        return cell

    def append_row(self, ws: Worksheet, values: list, *, start_row: Optional[int] = None):
        """類似 ws.append，但走 set_cell 以記錄 shadow，並用 ctx 內部 cursor。

        cursor 邏輯：第一次呼叫某 sheet 時掃實際非空列；之後維護 ctx 內 counter。
        clear_sheet 會把 cursor 重置為 1。
        """
        sheet_name = ws.title
        if start_row is not None:
            target_row = start_row
        elif sheet_name in self._row_cursor:
            target_row = self._row_cursor[sheet_name]
        else:
            last = 0
            for row in ws.iter_rows():
                if any(c.value is not None for c in row):
                    if row[0].row > last:
                        last = row[0].row
            target_row = last + 1

        for col_idx, v in enumerate(values, 1):
            self.set_cell(ws, target_row, col_idx, v)

        self._row_cursor[sheet_name] = target_row + 1
        return target_row

    def commit(self, filepath: str):
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        self.workbook.save(filepath)
        merged_shadow = dict(self._shadow)
        merged_shadow.update(self._new_shadow)
        _save_shadow(filepath, merged_shadow)
        logger.info(f"[excel_merge] saved + shadow updated: {filepath} "
                    f"({len(self._new_shadow)} cells written, "
                    f"{len(self._user_owned)} preserved)")


def reset_shadow(filepath: str):
    """清除某檔的 shadow（debug 用，全部 cell 都會被視為 DB 主權）。"""
    _ensure_table()
    conn = _get_conn()
    conn.execute("DELETE FROM excel_cell_shadow WHERE filepath=?", (filepath,))
    conn.commit()
    conn.close()


def save_with_shadow(new_wb: Workbook, filepath: str):
    """Drop-in 替代 wb.save(filepath) — 自動偵測使用者編輯並合併保留。

    流程：
      1. 若 filepath 已存在：load 舊檔 → 比對 shadow → 找出使用者主權 cell
      2. 把舊檔的 user-owned cell value 貼回 new_wb 對應位置
      3. save new_wb
      4. 把 new_wb 所有 cell 寫進新 shadow

    用法（service 改一行即可）：
        # wb.save(filepath)
        from services.excel_merge import save_with_shadow
        save_with_shadow(wb, filepath)
    """
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    user_overrides: list[tuple[str, str, object]] = []
    if os.path.exists(filepath):
        try:
            old_wb = load_workbook(filepath)
        except Exception as e:
            logger.warning(f"[save_with_shadow] load 舊檔失敗，整檔覆蓋：{filepath} ({e})")
            old_wb = None

        if old_wb is not None:
            shadow = _load_shadow(filepath)
            for sheet_name in old_wb.sheetnames:
                ws = old_wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        addr = cell.coordinate
                        key = (sheet_name, addr)
                        if key in shadow:
                            if _serialize(cell.value) != shadow[key]:
                                user_overrides.append((sheet_name, addr, cell.value))
                        else:
                            user_overrides.append((sheet_name, addr, cell.value))

    # 把使用者改的值貼回 new_wb
    preserved = 0
    for sheet_name, addr, value in user_overrides:
        if sheet_name in new_wb.sheetnames:
            new_wb[sheet_name][addr] = value
            preserved += 1
        else:
            ws = new_wb.create_sheet(sheet_name)
            ws[addr] = value
            preserved += 1

    new_wb.save(filepath)

    # 重建 shadow：以 new_wb 寫出後的最終值為基準（含已貼回的 user 值）
    new_shadow: dict[tuple[str, str], str] = {}
    saved_wb = load_workbook(filepath)
    for sheet_name in saved_wb.sheetnames:
        ws = saved_wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                # user_overrides 的 cell 不寫入 shadow → 之後仍會被視為 user-owned
                if any(sn == sheet_name and ad == cell.coordinate
                       for sn, ad, _ in user_overrides):
                    continue
                new_shadow[(sheet_name, cell.coordinate)] = _serialize(cell.value)
    _save_shadow(filepath, new_shadow)

    logger.info(f"[save_with_shadow] {filepath}: "
                f"shadow {len(new_shadow)} cells, 保留 user {preserved} cells")
