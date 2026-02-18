"""四大財務報表生成器（中小企業會計準則）"""

import logging
import os
from datetime import datetime

logger = logging.getLogger("shanbot.financial_report")

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "exports")


def _ensure_output_dir(year_month: str) -> str:
    """確保輸出目錄存在"""
    d = os.path.join(OUTPUT_DIR, year_month)
    os.makedirs(d, exist_ok=True)
    return d


def _get_income_data(year_month: str) -> dict:
    """取得收入資料"""
    import state_manager as sm
    income_rows = sm.get_income_summary(year_month)
    total = sum(r.get("amount", 0) for r in income_rows)
    return {"rows": income_rows, "total": total}


def _get_expense_data(year_month: str) -> dict:
    """取得支出資料（按分類）"""
    import state_manager as sm
    stagings = sm.get_stagings_by_month(year_month)
    cat_totals = {}
    total = 0
    for s in stagings:
        items = sm.get_purchase_items(s["id"])
        for item in items:
            cat = item.get("category", "其他")
            amt = item.get("amount", 0) or 0
            cat_totals[cat] = cat_totals.get(cat, 0) + amt
            total += amt
    return {"categories": cat_totals, "total": total}


def _get_monthly_cost_data(year_month: str) -> dict:
    """取得月度成本結構"""
    import state_manager as sm
    cost = sm.get_monthly_cost(year_month)
    return cost or {
        "ingredient_total": 0, "labor_total": 0, "overhead_total": 0,
        "taxable_purchase_total": 0, "input_tax_total": 0,
    }


def generate_balance_sheet(year_month: str, output_dir: str = None) -> str | None:
    """生成資產負債表"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        logger.error("openpyxl not installed")
        return None

    out = output_dir or _ensure_output_dir(year_month)
    os.makedirs(out, exist_ok=True)

    income = _get_income_data(year_month)
    expense = _get_expense_data(year_month)
    cost = _get_monthly_cost_data(year_month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "資產負債表"

    # 標題
    header_font = Font(name="微軟正黑體", size=14, bold=True)
    sub_font = Font(name="微軟正黑體", size=11, bold=True)
    normal_font = Font(name="微軟正黑體", size=10)
    money_fmt = '#,##0'

    ws.merge_cells("A1:D1")
    ws["A1"] = "資 產 負 債 表"
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = f"期間：{year_month}"
    ws["A2"].font = normal_font
    ws["A2"].alignment = Alignment(horizontal="center")

    # 資產
    row = 4
    ws.cell(row=row, column=1, value="資產").font = sub_font
    row += 1

    # 流動資產
    ws.cell(row=row, column=1, value="流動資產").font = sub_font
    row += 1

    total_revenue = income["total"]
    total_expense = expense["total"]
    ingredient = cost.get("ingredient_total", 0) or 0
    cash_estimate = total_revenue - total_expense  # 簡化：淨現金

    ws.cell(row=row, column=2, value="現金及約當現金")
    ws.cell(row=row, column=4, value=max(cash_estimate, 0)).number_format = money_fmt
    row += 1

    ws.cell(row=row, column=2, value="應收帳款")
    ws.cell(row=row, column=4, value=0).number_format = money_fmt
    row += 1

    ws.cell(row=row, column=2, value="存貨（食材）")
    ws.cell(row=row, column=4, value=ingredient).number_format = money_fmt
    row += 1

    current_assets = max(cash_estimate, 0) + ingredient
    ws.cell(row=row, column=2, value="流動資產合計").font = sub_font
    ws.cell(row=row, column=4, value=current_assets).number_format = money_fmt
    ws.cell(row=row, column=4).font = sub_font
    row += 2

    # 固定資產（簡化）
    ws.cell(row=row, column=1, value="非流動資產").font = sub_font
    row += 1
    ws.cell(row=row, column=2, value="不動產、廠房及設備")
    ws.cell(row=row, column=4, value=0).number_format = money_fmt
    row += 1

    total_assets = current_assets
    ws.cell(row=row, column=1, value="資產總計").font = Font(name="微軟正黑體", size=11, bold=True)
    ws.cell(row=row, column=4, value=total_assets).number_format = money_fmt
    ws.cell(row=row, column=4).font = Font(name="微軟正黑體", size=11, bold=True)
    row += 2

    # 負債
    ws.cell(row=row, column=1, value="負債").font = sub_font
    row += 1
    ws.cell(row=row, column=2, value="應付帳款")
    ap = total_expense * 0.3  # 簡化估計
    ws.cell(row=row, column=4, value=ap).number_format = money_fmt
    row += 1

    tax_payable = cost.get("input_tax_total", 0) or 0
    ws.cell(row=row, column=2, value="應付稅款")
    ws.cell(row=row, column=4, value=tax_payable).number_format = money_fmt
    row += 1

    total_liabilities = ap + tax_payable
    ws.cell(row=row, column=1, value="負債合計").font = sub_font
    ws.cell(row=row, column=4, value=total_liabilities).number_format = money_fmt
    ws.cell(row=row, column=4).font = sub_font
    row += 2

    # 權益
    ws.cell(row=row, column=1, value="權益").font = sub_font
    row += 1
    equity = total_assets - total_liabilities
    ws.cell(row=row, column=2, value="業主權益（淨值）")
    ws.cell(row=row, column=4, value=equity).number_format = money_fmt
    row += 1
    ws.cell(row=row, column=1, value="負債及權益合計").font = Font(name="微軟正黑體", size=11, bold=True)
    ws.cell(row=row, column=4, value=total_assets).number_format = money_fmt
    ws.cell(row=row, column=4).font = Font(name="微軟正黑體", size=11, bold=True)

    # 欄寬
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 18

    filename = f"{year_month}_資產負債表.xlsx"
    filepath = os.path.join(out, filename)
    wb.save(filepath)
    logger.info(f"Balance sheet generated: {filepath}")
    return filepath


def generate_income_statement(year_month: str, output_dir: str = None) -> str | None:
    """生成損益表"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        logger.error("openpyxl not installed")
        return None

    out = output_dir or _ensure_output_dir(year_month)
    os.makedirs(out, exist_ok=True)

    income = _get_income_data(year_month)
    expense = _get_expense_data(year_month)
    cost = _get_monthly_cost_data(year_month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "損益表"

    header_font = Font(name="微軟正黑體", size=14, bold=True)
    sub_font = Font(name="微軟正黑體", size=11, bold=True)
    normal_font = Font(name="微軟正黑體", size=10)
    money_fmt = '#,##0'

    ws.merge_cells("A1:D1")
    ws["A1"] = "損 益 表"
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = f"期間：{year_month}"
    ws["A2"].font = normal_font
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    # 營業收入
    ws.cell(row=row, column=1, value="營業收入").font = sub_font
    row += 1
    total_revenue = income["total"]
    for entry in income["rows"]:
        ws.cell(row=row, column=2, value=entry.get("description", "團膳收入"))
        ws.cell(row=row, column=4, value=entry.get("amount", 0)).number_format = money_fmt
        row += 1
    ws.cell(row=row, column=2, value="營業收入合計").font = sub_font
    ws.cell(row=row, column=4, value=total_revenue).number_format = money_fmt
    ws.cell(row=row, column=4).font = sub_font
    row += 2

    # 營業成本
    ws.cell(row=row, column=1, value="營業成本").font = sub_font
    row += 1
    ingredient = cost.get("ingredient_total", 0) or 0
    labor = cost.get("labor_total", 0) or 0

    ws.cell(row=row, column=2, value="食材成本")
    ws.cell(row=row, column=4, value=ingredient).number_format = money_fmt
    row += 1
    ws.cell(row=row, column=2, value="直接人工")
    ws.cell(row=row, column=4, value=labor).number_format = money_fmt
    row += 1

    cost_total = ingredient + labor
    ws.cell(row=row, column=2, value="營業成本合計").font = sub_font
    ws.cell(row=row, column=4, value=cost_total).number_format = money_fmt
    ws.cell(row=row, column=4).font = sub_font
    row += 2

    # 營業毛利
    gross_profit = total_revenue - cost_total
    ws.cell(row=row, column=1, value="營業毛利").font = sub_font
    ws.cell(row=row, column=4, value=gross_profit).number_format = money_fmt
    ws.cell(row=row, column=4).font = sub_font
    row += 2

    # 營業費用
    ws.cell(row=row, column=1, value="營業費用").font = sub_font
    row += 1
    overhead = cost.get("overhead_total", 0) or 0
    for cat, amt in sorted(expense["categories"].items(), key=lambda x: -x[1]):
        if cat not in ("蔬菜", "肉類", "水產", "蛋豆", "乾貨", "調味料", "油品", "米糧"):
            ws.cell(row=row, column=2, value=cat)
            ws.cell(row=row, column=4, value=amt).number_format = money_fmt
            row += 1
    if overhead:
        ws.cell(row=row, column=2, value="其他營業費用")
        ws.cell(row=row, column=4, value=overhead).number_format = money_fmt
        row += 1
    row += 1

    # 本期淨利
    net_income = gross_profit - overhead
    ws.cell(row=row, column=1, value="本期淨利（淨損）").font = Font(name="微軟正黑體", size=12, bold=True)
    ws.cell(row=row, column=4, value=net_income).number_format = money_fmt
    ws.cell(row=row, column=4).font = Font(name="微軟正黑體", size=12, bold=True)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 18

    filename = f"{year_month}_損益表.xlsx"
    filepath = os.path.join(out, filename)
    wb.save(filepath)
    logger.info(f"Income statement generated: {filepath}")
    return filepath


def generate_cash_flow(year_month: str, output_dir: str = None) -> str | None:
    """生成現金流量表"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        logger.error("openpyxl not installed")
        return None

    out = output_dir or _ensure_output_dir(year_month)
    os.makedirs(out, exist_ok=True)

    income = _get_income_data(year_month)
    expense = _get_expense_data(year_month)
    cost = _get_monthly_cost_data(year_month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "現金流量表"

    header_font = Font(name="微軟正黑體", size=14, bold=True)
    sub_font = Font(name="微軟正黑體", size=11, bold=True)
    normal_font = Font(name="微軟正黑體", size=10)
    money_fmt = '#,##0'

    ws.merge_cells("A1:D1")
    ws["A1"] = "現 金 流 量 表"
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = f"期間：{year_month}"
    ws["A2"].font = normal_font
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    total_revenue = income["total"]
    total_expense = expense["total"]
    ingredient = cost.get("ingredient_total", 0) or 0
    labor = cost.get("labor_total", 0) or 0
    overhead = cost.get("overhead_total", 0) or 0

    # 營業活動
    ws.cell(row=row, column=1, value="營業活動之現金流量").font = sub_font
    row += 1
    net_income = total_revenue - ingredient - labor - overhead
    ws.cell(row=row, column=2, value="本期淨利")
    ws.cell(row=row, column=4, value=net_income).number_format = money_fmt
    row += 1

    ws.cell(row=row, column=2, value="調整項目：")
    row += 1
    ws.cell(row=row, column=3, value="折舊費用")
    ws.cell(row=row, column=4, value=0).number_format = money_fmt
    row += 1
    ws.cell(row=row, column=3, value="存貨變動")
    ws.cell(row=row, column=4, value=-ingredient).number_format = money_fmt
    row += 1

    operating_cf = net_income - ingredient
    ws.cell(row=row, column=2, value="營業活動淨現金").font = sub_font
    ws.cell(row=row, column=4, value=operating_cf).number_format = money_fmt
    ws.cell(row=row, column=4).font = sub_font
    row += 2

    # 投資活動
    ws.cell(row=row, column=1, value="投資活動之現金流量").font = sub_font
    row += 1
    ws.cell(row=row, column=2, value="購置設備")
    ws.cell(row=row, column=4, value=0).number_format = money_fmt
    row += 1
    ws.cell(row=row, column=2, value="投資活動淨現金").font = sub_font
    ws.cell(row=row, column=4, value=0).number_format = money_fmt
    ws.cell(row=row, column=4).font = sub_font
    row += 2

    # 籌資活動
    ws.cell(row=row, column=1, value="籌資活動之現金流量").font = sub_font
    row += 1
    ws.cell(row=row, column=2, value="業主提款/投入")
    ws.cell(row=row, column=4, value=0).number_format = money_fmt
    row += 1
    ws.cell(row=row, column=2, value="籌資活動淨現金").font = sub_font
    ws.cell(row=row, column=4, value=0).number_format = money_fmt
    ws.cell(row=row, column=4).font = sub_font
    row += 2

    # 淨增減
    ws.cell(row=row, column=1, value="本期現金淨增減").font = Font(name="微軟正黑體", size=12, bold=True)
    ws.cell(row=row, column=4, value=operating_cf).number_format = money_fmt
    ws.cell(row=row, column=4).font = Font(name="微軟正黑體", size=12, bold=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18

    filename = f"{year_month}_現金流量表.xlsx"
    filepath = os.path.join(out, filename)
    wb.save(filepath)
    logger.info(f"Cash flow statement generated: {filepath}")
    return filepath


def generate_equity_changes(year_month: str, output_dir: str = None) -> str | None:
    """生成權益變動表"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        logger.error("openpyxl not installed")
        return None

    out = output_dir or _ensure_output_dir(year_month)
    os.makedirs(out, exist_ok=True)

    income = _get_income_data(year_month)
    expense = _get_expense_data(year_month)
    cost = _get_monthly_cost_data(year_month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "權益變動表"

    header_font = Font(name="微軟正黑體", size=14, bold=True)
    sub_font = Font(name="微軟正黑體", size=11, bold=True)
    normal_font = Font(name="微軟正黑體", size=10)
    money_fmt = '#,##0'

    ws.merge_cells("A1:D1")
    ws["A1"] = "權 益 變 動 表"
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = f"期間：{year_month}"
    ws["A2"].font = normal_font
    ws["A2"].alignment = Alignment(horizontal="center")

    # 表頭
    row = 4
    ws.cell(row=row, column=1, value="項目").font = sub_font
    ws.cell(row=row, column=2, value="資本").font = sub_font
    ws.cell(row=row, column=3, value="保留盈餘").font = sub_font
    ws.cell(row=row, column=4, value="合計").font = sub_font
    row += 1

    total_revenue = income["total"]
    ingredient = cost.get("ingredient_total", 0) or 0
    labor = cost.get("labor_total", 0) or 0
    overhead = cost.get("overhead_total", 0) or 0
    net_income = total_revenue - ingredient - labor - overhead

    # 期初餘額（簡化為 0）
    ws.cell(row=row, column=1, value="期初餘額")
    ws.cell(row=row, column=2, value=0).number_format = money_fmt
    ws.cell(row=row, column=3, value=0).number_format = money_fmt
    ws.cell(row=row, column=4, value=0).number_format = money_fmt
    row += 1

    # 本期損益
    ws.cell(row=row, column=1, value="本期淨利（淨損）")
    ws.cell(row=row, column=2, value=0).number_format = money_fmt
    ws.cell(row=row, column=3, value=net_income).number_format = money_fmt
    ws.cell(row=row, column=4, value=net_income).number_format = money_fmt
    row += 1

    # 業主增資/提款
    ws.cell(row=row, column=1, value="業主增資")
    ws.cell(row=row, column=2, value=0).number_format = money_fmt
    ws.cell(row=row, column=3, value=0).number_format = money_fmt
    ws.cell(row=row, column=4, value=0).number_format = money_fmt
    row += 1

    ws.cell(row=row, column=1, value="業主提款")
    ws.cell(row=row, column=2, value=0).number_format = money_fmt
    ws.cell(row=row, column=3, value=0).number_format = money_fmt
    ws.cell(row=row, column=4, value=0).number_format = money_fmt
    row += 2

    # 期末餘額
    ws.cell(row=row, column=1, value="期末餘額").font = Font(name="微軟正黑體", size=12, bold=True)
    ws.cell(row=row, column=2, value=0).number_format = money_fmt
    ws.cell(row=row, column=2).font = sub_font
    ws.cell(row=row, column=3, value=net_income).number_format = money_fmt
    ws.cell(row=row, column=3).font = sub_font
    ws.cell(row=row, column=4, value=net_income).number_format = money_fmt
    ws.cell(row=row, column=4).font = Font(name="微軟正黑體", size=12, bold=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    filename = f"{year_month}_權益變動表.xlsx"
    filepath = os.path.join(out, filename)
    wb.save(filepath)
    logger.info(f"Equity changes statement generated: {filepath}")
    return filepath
