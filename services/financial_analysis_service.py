"""小膳 Bot — 月度財務分析報告

每月結束時自動產出：
1. 成本結構分析（食材分類佔比、前 N 大供應商）
2. 趨勢分析（環比成長/衰退、同期比較）
3. 異常與風險偵測（價格波動、集中度、未分類）
4. 經營指標（毛利率、成本率、每餐成本）
5. 具體改進建議

產出格式：Excel 報告 + 文字摘要（可推送 LINE）
"""

import logging
import os
from datetime import datetime
from typing import Optional

import state_manager as sm

logger = logging.getLogger("shanbot.analysis")


def generate_monthly_analysis(year_month: str, company_id: int = None) -> dict:
    """產出月度財務分析報告

    Args:
        year_month: 如 "2026-03"
        company_id: 指定公司（None = 全部彙總）

    Returns:
        {
            "year_month": str,
            "cost_structure": dict,      # 成本結構
            "trend": dict,               # 趨勢分析
            "kpi": dict,                 # 經營指標
            "risks": list,               # 風險項目
            "recommendations": list,     # 改進建議
            "summary_text": str,         # LINE 推送用摘要
        }
    """
    stagings = sm.get_stagings_by_month(year_month, company_id=company_id)
    confirmed = [s for s in stagings if s.get("status") in ("confirmed", "exported")]
    incomes = sm.get_income_summary(year_month)

    total_revenue = sum(i.get("amount", 0) for i in incomes)
    total_purchase = sum(s.get("total_amount", 0) or 0 for s in confirmed)
    total_subtotal = sum(s.get("subtotal", 0) or 0 for s in confirmed)
    total_tax = sum(s.get("tax_amount", 0) or 0 for s in confirmed)

    # === 1. 成本結構分析 ===
    cost_structure = _analyze_cost_structure(confirmed)

    # === 2. 趨勢分析（環比） ===
    trend = _analyze_trend(year_month, total_purchase, total_revenue, company_id)

    # === 3. 經營指標 ===
    kpi = _calculate_kpi(total_revenue, total_subtotal, total_tax, confirmed, incomes)

    # === 4. 風險偵測 ===
    risks = _detect_risks(confirmed, cost_structure, trend, kpi)

    # === 5. 改進建議 ===
    recommendations = _generate_recommendations(risks, cost_structure, trend, kpi)

    # === 摘要文字 ===
    summary_text = _format_summary(year_month, cost_structure, trend, kpi, risks, recommendations)

    return {
        "year_month": year_month,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "cost_structure": cost_structure,
        "trend": trend,
        "kpi": kpi,
        "risks": risks,
        "recommendations": recommendations,
        "summary_text": summary_text,
    }


# =====================================================================
# 分析模組
# =====================================================================

def _analyze_cost_structure(confirmed: list) -> dict:
    """成本結構分析：分類佔比 + 供應商排名"""
    category_totals = {}
    supplier_totals = {}
    daily_totals = {}
    total = 0

    for s in confirmed:
        name = s.get("supplier_name") or "未知"
        amt = s.get("total_amount", 0) or 0
        date = s.get("purchase_date", "")
        supplier_totals[name] = supplier_totals.get(name, 0) + amt
        daily_totals[date] = daily_totals.get(date, 0) + amt
        total += amt

        items = sm.get_purchase_items(s["id"])
        for item in items:
            cat = item.get("category") or "其他"
            item_amt = item.get("amount", 0) or 0
            category_totals[cat] = category_totals.get(cat, 0) + item_amt

    # 排序
    cat_sorted = sorted(category_totals.items(), key=lambda x: -x[1])
    sup_sorted = sorted(supplier_totals.items(), key=lambda x: -x[1])

    # 佔比
    cat_pct = [(c, a, round(a / total * 100, 1) if total > 0 else 0) for c, a in cat_sorted]
    sup_pct = [(s, a, round(a / total * 100, 1) if total > 0 else 0) for s, a in sup_sorted]

    return {
        "total_cost": total,
        "categories": cat_pct,  # [(name, amount, pct), ...]
        "suppliers": sup_pct,
        "daily_totals": daily_totals,
        "transaction_count": len(confirmed),
        "avg_per_transaction": round(total / len(confirmed)) if confirmed else 0,
    }


def _analyze_trend(year_month: str, current_purchase: float,
                   current_revenue: float, company_id: int = None) -> dict:
    """趨勢分析：與上月比較"""
    parts = year_month.split("-")
    y, m = int(parts[0]), int(parts[1])
    if m == 1:
        prev_ym = f"{y - 1}-12"
    else:
        prev_ym = f"{y}-{m - 1:02d}"

    prev_stagings = sm.get_stagings_by_month(prev_ym, company_id=company_id)
    prev_confirmed = [s for s in prev_stagings if s.get("status") in ("confirmed", "exported")]
    prev_purchase = sum(s.get("total_amount", 0) or 0 for s in prev_confirmed)

    prev_incomes = sm.get_income_summary(prev_ym)
    prev_revenue = sum(i.get("amount", 0) for i in prev_incomes)

    # 環比變化
    purchase_change = _calc_change(current_purchase, prev_purchase)
    revenue_change = _calc_change(current_revenue, prev_revenue)

    return {
        "prev_month": prev_ym,
        "current_purchase": current_purchase,
        "prev_purchase": prev_purchase,
        "purchase_change_pct": purchase_change,
        "current_revenue": current_revenue,
        "prev_revenue": prev_revenue,
        "revenue_change_pct": revenue_change,
        "current_count": len([s for s in sm.get_stagings_by_month(year_month, company_id=company_id)
                              if s.get("status") in ("confirmed", "exported")]),
        "prev_count": len(prev_confirmed),
    }


def _calc_change(current: float, previous: float) -> float:
    """計算環比變化百分比"""
    if previous == 0:
        return 100.0 if current > 0 else 0.0
    return round((current - previous) / previous * 100, 1)


def _calculate_kpi(revenue: float, subtotal: float, tax: float,
                   confirmed: list, incomes: list) -> dict:
    """經營指標"""
    gross_profit = revenue - subtotal
    gross_margin = round(gross_profit / revenue * 100, 1) if revenue > 0 else 0
    cost_ratio = round(subtotal / revenue * 100, 1) if revenue > 0 else 0

    # 營業天數估算（有交易的天數）
    dates = set(s.get("purchase_date") for s in confirmed if s.get("purchase_date"))
    operating_days = len(dates)

    avg_daily_cost = round(subtotal / operating_days) if operating_days > 0 else 0
    avg_daily_revenue = round(revenue / 30)  # 月均

    # 進項稅額可扣抵比例
    deductible_count = sum(1 for s in confirmed if s.get("deduction_code") == "1")
    deductible_ratio = round(deductible_count / len(confirmed) * 100, 1) if confirmed else 0

    return {
        "revenue": revenue,
        "cost": subtotal,
        "tax": tax,
        "gross_profit": gross_profit,
        "gross_margin_pct": gross_margin,
        "cost_ratio_pct": cost_ratio,
        "operating_days": operating_days,
        "avg_daily_cost": avg_daily_cost,
        "avg_daily_revenue": avg_daily_revenue,
        "invoice_count": sum(1 for s in confirmed if s.get("supplier_tax_id")),
        "receipt_count": sum(1 for s in confirmed if not s.get("supplier_tax_id")),
        "deductible_ratio_pct": deductible_ratio,
    }


def _detect_risks(confirmed: list, cost_structure: dict,
                  trend: dict, kpi: dict) -> list:
    """風險偵測"""
    risks = []

    # 1. 毛利率過低
    if kpi["gross_margin_pct"] < 20 and kpi["revenue"] > 0:
        risks.append({
            "level": "high",
            "type": "low_margin",
            "message": f"毛利率偏低 {kpi['gross_margin_pct']}%（建議 ≥ 30%）",
            "value": kpi["gross_margin_pct"],
        })
    elif kpi["gross_margin_pct"] < 30 and kpi["revenue"] > 0:
        risks.append({
            "level": "medium",
            "type": "low_margin",
            "message": f"毛利率一般 {kpi['gross_margin_pct']}%（行業標準 30-45%）",
            "value": kpi["gross_margin_pct"],
        })

    # 2. 採購成本環比大幅上升
    if trend["purchase_change_pct"] > 20:
        risks.append({
            "level": "medium",
            "type": "cost_surge",
            "message": f"採購成本環比上升 {trend['purchase_change_pct']}%（上月 ${trend['prev_purchase']:,.0f}）",
            "value": trend["purchase_change_pct"],
        })

    # 3. 供應商集中度
    for name, amt, pct in cost_structure["suppliers"]:
        if pct >= 50:
            risks.append({
                "level": "high",
                "type": "supplier_concentration",
                "message": f"供應商 {name} 佔比達 {pct}%，依賴度過高",
                "value": pct,
            })
        elif pct >= 35:
            risks.append({
                "level": "medium",
                "type": "supplier_concentration",
                "message": f"供應商 {name} 佔比 {pct}%，建議分散採購",
                "value": pct,
            })

    # 4. 免用發票比例偏高（影響進項扣抵）
    if kpi["deductible_ratio_pct"] < 60 and len(confirmed) > 3:
        risks.append({
            "level": "medium",
            "type": "low_deductible",
            "message": f"可扣抵發票比例僅 {kpi['deductible_ratio_pct']}%，影響營業稅抵扣",
            "value": kpi["deductible_ratio_pct"],
        })

    # 5. 單一分類佔比過高（食材結構失衡）
    for cat, amt, pct in cost_structure["categories"]:
        if pct >= 60 and cat != "其他":
            risks.append({
                "level": "low",
                "type": "category_imbalance",
                "message": f"{cat}佔採購成本 {pct}%，食材結構可能不夠多元",
                "value": pct,
            })

    # 6. 收入下降
    if trend["revenue_change_pct"] < -10 and trend["prev_revenue"] > 0:
        risks.append({
            "level": "high",
            "type": "revenue_decline",
            "message": f"收入環比下降 {abs(trend['revenue_change_pct'])}%",
            "value": trend["revenue_change_pct"],
        })

    return risks


def _generate_recommendations(risks: list, cost_structure: dict,
                               trend: dict, kpi: dict) -> list:
    """根據風險產出改進建議"""
    recs = []

    risk_types = {r["type"] for r in risks}

    if "low_margin" in risk_types:
        recs.append({
            "priority": "高",
            "area": "定價策略",
            "action": "檢討團膳報價，考慮調高單價或降低食材成本。目標毛利率 30% 以上。",
        })

    if "cost_surge" in risk_types:
        surge = next((r for r in risks if r["type"] == "cost_surge"), None)
        recs.append({
            "priority": "高",
            "area": "成本控管",
            "action": f"本月採購較上月增加 {surge['value']}%，建議比對品項單價變化，找出漲幅最大的食材。",
        })

    if "supplier_concentration" in risk_types:
        recs.append({
            "priority": "中",
            "area": "供應鏈",
            "action": "主力供應商佔比過高，建議發展替代供應商以降低斷貨風險和議價能力。",
        })

    if "low_deductible" in risk_types:
        recs.append({
            "priority": "中",
            "area": "稅務優化",
            "action": "免用發票比例偏高，盡量向有統一發票的供應商採購，提高進項稅額扣抵比例。",
        })

    if "revenue_decline" in risk_types:
        recs.append({
            "priority": "高",
            "area": "營收",
            "action": "收入下滑，確認是否有合約到期、客戶流失或季節性因素，及早調整行銷策略。",
        })

    if "category_imbalance" in risk_types:
        recs.append({
            "priority": "低",
            "area": "菜單設計",
            "action": "食材結構偏重單一類別，建議豐富菜色以平衡營養和成本。",
        })

    # 通用建議
    if not risks:
        recs.append({
            "priority": "提醒",
            "area": "整體",
            "action": "本月財務健康度良好，持續維持目前的採購和成本控管模式。",
        })

    if kpi["operating_days"] < 20 and kpi["revenue"] > 0:
        recs.append({
            "priority": "低",
            "area": "記帳完整度",
            "action": f"本月僅 {kpi['operating_days']} 天有採購記錄，確認是否所有單據都已上傳。",
        })

    return recs


def _format_summary(year_month: str, cost: dict, trend: dict,
                    kpi: dict, risks: list, recs: list) -> str:
    """格式化為 LINE 推送用的文字摘要"""
    lines = [
        f"📊 {year_month} 月度財務分析報告",
        "",
        "【經營指標】",
        f"  營收：${kpi['revenue']:,.0f}",
        f"  成本：${kpi['cost']:,.0f}（成本率 {kpi['cost_ratio_pct']}%）",
        f"  毛利：${kpi['gross_profit']:,.0f}（毛利率 {kpi['gross_margin_pct']}%）",
        f"  發票：{kpi['invoice_count']} 張 / 收據：{kpi['receipt_count']} 張",
    ]

    # 趨勢
    if trend["prev_purchase"] > 0:
        arrow_p = "↑" if trend["purchase_change_pct"] > 0 else "↓"
        arrow_r = "↑" if trend["revenue_change_pct"] > 0 else "↓"
        lines.append("")
        lines.append("【環比趨勢】")
        lines.append(f"  採購：{arrow_p} {abs(trend['purchase_change_pct'])}%（vs {trend['prev_month']}）")
        lines.append(f"  營收：{arrow_r} {abs(trend['revenue_change_pct'])}%")

    # 成本結構 Top 3
    lines.append("")
    lines.append("【成本結構 Top 3】")
    for cat, amt, pct in cost["categories"][:3]:
        lines.append(f"  {cat}：${amt:,.0f}（{pct}%）")

    # 風險
    high_risks = [r for r in risks if r["level"] == "high"]
    med_risks = [r for r in risks if r["level"] == "medium"]
    if high_risks or med_risks:
        lines.append("")
        lines.append("【⚠️ 風險提醒】")
        for r in high_risks:
            lines.append(f"  🔴 {r['message']}")
        for r in med_risks:
            lines.append(f"  🟡 {r['message']}")

    # 建議
    if recs:
        lines.append("")
        lines.append("【💡 改進建議】")
        for r in recs[:3]:
            lines.append(f"  [{r['priority']}] {r['area']}：{r['action']}")

    return "\n".join(lines)


# =====================================================================
# Excel 報告
# =====================================================================

def generate_analysis_excel(year_month: str, output_dir: str = None,
                            company_id: int = None) -> Optional[str]:
    """產出月度財務分析報告 Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed")
        return None

    analysis = generate_monthly_analysis(year_month, company_id)

    if not output_dir:
        output_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "data", "exports", year_month
        )
    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    title_font = Font(name="微軟正黑體", size=14, bold=True)
    header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    normal_font = Font(name="微軟正黑體", size=10)
    money_fmt = '#,##0'
    pct_fmt = '0.0%'
    high_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    med_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    good_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

    # === Sheet 1: 總覽 ===
    ws1 = wb.active
    ws1.title = "財務分析總覽"
    ws1.merge_cells("A1:E1")
    ws1["A1"] = f"{year_month} 月度財務分析報告"
    ws1["A1"].font = title_font
    ws1["A2"] = f"產生時間：{analysis['timestamp']}"
    ws1["A2"].font = normal_font

    row = 4
    ws1.cell(row=row, column=1, value="經營指標").font = Font(name="微軟正黑體", size=12, bold=True)
    row += 1
    kpi = analysis["kpi"]
    kpi_items = [
        ("營業收入", kpi["revenue"]),
        ("採購成本（未稅）", kpi["cost"]),
        ("進項稅額", kpi["tax"]),
        ("毛利", kpi["gross_profit"]),
        ("毛利率", f"{kpi['gross_margin_pct']}%"),
        ("成本率", f"{kpi['cost_ratio_pct']}%"),
        ("採購天數", kpi["operating_days"]),
        ("日均採購成本", kpi["avg_daily_cost"]),
        ("發票張數", kpi["invoice_count"]),
        ("收據張數", kpi["receipt_count"]),
        ("可扣抵比例", f"{kpi['deductible_ratio_pct']}%"),
    ]
    for label, val in kpi_items:
        ws1.cell(row=row, column=1, value=label).font = normal_font
        cell = ws1.cell(row=row, column=2, value=val)
        cell.font = normal_font
        if isinstance(val, (int, float)):
            cell.number_format = money_fmt
        row += 1

    # 趨勢
    row += 1
    ws1.cell(row=row, column=1, value="環比趨勢").font = Font(name="微軟正黑體", size=12, bold=True)
    row += 1
    trend = analysis["trend"]
    trend_items = [
        ("上月採購", trend["prev_purchase"]),
        ("本月採購", trend["current_purchase"]),
        ("採購變動", f"{trend['purchase_change_pct']:+.1f}%"),
        ("上月營收", trend["prev_revenue"]),
        ("本月營收", trend["current_revenue"]),
        ("營收變動", f"{trend['revenue_change_pct']:+.1f}%"),
    ]
    for label, val in trend_items:
        ws1.cell(row=row, column=1, value=label).font = normal_font
        cell = ws1.cell(row=row, column=2, value=val)
        cell.font = normal_font
        if isinstance(val, (int, float)):
            cell.number_format = money_fmt
        row += 1

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 18

    # === Sheet 2: 成本結構 ===
    ws2 = wb.create_sheet("成本結構")
    ws2.merge_cells("A1:D1")
    ws2["A1"] = f"{year_month} 成本結構分析"
    ws2["A1"].font = title_font

    # 分類
    row = 3
    ws2.cell(row=row, column=1, value="食材分類").font = Font(name="微軟正黑體", size=12, bold=True)
    row += 1
    for col, h in enumerate(["分類", "金額", "佔比"], 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for cat, amt, pct in analysis["cost_structure"]["categories"]:
        row += 1
        ws2.cell(row=row, column=1, value=cat).font = normal_font
        ws2.cell(row=row, column=2, value=amt).number_format = money_fmt
        ws2.cell(row=row, column=3, value=f"{pct}%").font = normal_font

    # 圓餅圖
    if analysis["cost_structure"]["categories"]:
        pie = PieChart()
        pie.title = "食材成本佔比"
        pie.style = 10
        data_start = 5
        data_end = data_start + len(analysis["cost_structure"]["categories"]) - 1
        data = Reference(ws2, min_col=2, min_row=data_start, max_row=data_end)
        cats = Reference(ws2, min_col=1, min_row=data_start, max_row=data_end)
        pie.add_data(data)
        pie.set_categories(cats)
        pie.width = 15
        pie.height = 10
        ws2.add_chart(pie, "E3")

    # 供應商
    row += 2
    ws2.cell(row=row, column=1, value="供應商排名").font = Font(name="微軟正黑體", size=12, bold=True)
    row += 1
    for col, h in enumerate(["供應商", "金額", "佔比"], 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for name, amt, pct in analysis["cost_structure"]["suppliers"]:
        row += 1
        ws2.cell(row=row, column=1, value=name).font = normal_font
        ws2.cell(row=row, column=2, value=amt).number_format = money_fmt
        ws2.cell(row=row, column=3, value=f"{pct}%").font = normal_font

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 10

    # === Sheet 3: 風險與建議 ===
    ws3 = wb.create_sheet("風險與建議")
    ws3.merge_cells("A1:D1")
    ws3["A1"] = f"{year_month} 風險偵測與改進建議"
    ws3["A1"].font = title_font

    row = 3
    ws3.cell(row=row, column=1, value="風險項目").font = Font(name="微軟正黑體", size=12, bold=True)
    row += 1
    for col, h in enumerate(["等級", "類型", "說明", "數值"], 1):
        cell = ws3.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    level_labels = {"high": "🔴 高", "medium": "🟡 中", "low": "🟢 低"}
    level_fills = {"high": high_fill, "medium": med_fill, "low": good_fill}

    if analysis["risks"]:
        for r in analysis["risks"]:
            row += 1
            cell = ws3.cell(row=row, column=1, value=level_labels.get(r["level"], r["level"]))
            cell.font = normal_font
            cell.fill = level_fills.get(r["level"], good_fill)
            ws3.cell(row=row, column=2, value=r["type"]).font = normal_font
            ws3.cell(row=row, column=3, value=r["message"]).font = normal_font
            ws3.cell(row=row, column=4, value=r.get("value", "")).font = normal_font
    else:
        row += 1
        ws3.cell(row=row, column=1, value="✅ 未偵測到顯著風險").font = normal_font
        ws3.cell(row=row, column=1).fill = good_fill

    row += 2
    ws3.cell(row=row, column=1, value="改進建議").font = Font(name="微軟正黑體", size=12, bold=True)
    row += 1
    for col, h in enumerate(["優先度", "領域", "建議行動"], 1):
        cell = ws3.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for rec in analysis["recommendations"]:
        row += 1
        ws3.cell(row=row, column=1, value=rec["priority"]).font = normal_font
        ws3.cell(row=row, column=2, value=rec["area"]).font = normal_font
        ws3.cell(row=row, column=3, value=rec["action"]).font = normal_font

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 50
    ws3.column_dimensions["D"].width = 12

    filepath = os.path.join(output_dir, f"{year_month}_財務分析報告.xlsx")
    wb.save(filepath)
    logger.info(f"Financial analysis Excel generated: {filepath}")
    return filepath
