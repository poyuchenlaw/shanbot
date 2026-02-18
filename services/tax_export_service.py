"""小膳 Bot — 稅務匯出服務

提供三種匯出格式：
1. MOF 進項 TXT（81-byte 固定寬度，供財政部媒體申報）
2. 文中資訊 Excel（會計系統匯入用）
3. 經手人證明表 PDF（免用發票之市場採購）

以及匯出前驗證與稅期工具函數。
"""

import logging
import math
import os
import re
from datetime import date, datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm, mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

import state_manager as sm

logger = logging.getLogger("shanbot.tax_export")

# === 環境變數 ===
COMPANY_TAX_ID = os.environ.get("COMPANY_TAX_ID", "")
COMPANY_TAX_REG_NO = os.environ.get("COMPANY_TAX_REG_NO", "")
COMPANY_NAME = os.environ.get("COMPANY_NAME", "")

# === 中文字型路徑（reportlab PDF 用）===
_CJK_FONT_PATHS = [
    "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/noto-cjk/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
    "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
    "C:/Windows/Fonts/msjh.ttc",
]

_CJK_FONT_NAME = None


def _register_cjk_font() -> str:
    """註冊中文字型，回傳字型名稱。找不到則用 Helvetica。"""
    global _CJK_FONT_NAME
    if _CJK_FONT_NAME:
        return _CJK_FONT_NAME

    for font_path in _CJK_FONT_PATHS:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont("CJK", font_path))
                _CJK_FONT_NAME = "CJK"
                logger.info(f"CJK 字型已註冊: {font_path}")
                return _CJK_FONT_NAME
            except Exception as e:
                logger.warning(f"字型註冊失敗 {font_path}: {e}")
                continue

    logger.warning("找不到 CJK 字型，PDF 中文可能無法顯示")
    _CJK_FONT_NAME = "Helvetica"
    return _CJK_FONT_NAME


# =====================================================================
# 1. MOF 進項 TXT 匯出（81-byte 固定寬度）
# =====================================================================

def export_mof_txt(tax_period: str, output_dir: str) -> str:
    """匯出財政部媒體申報用進項 TXT 檔。

    每行恰好 81 bytes（ASCII），符合財政部規定格式。

    Args:
        tax_period: 稅期，如 "2026-01-02"
        output_dir: 輸出目錄路徑

    Returns:
        產出檔案的完整路徑

    Raises:
        ValueError: 環境變數未設定或無資料可匯出
    """
    # 驗證必要環境變數
    tax_id = COMPANY_TAX_ID
    tax_reg_no = COMPANY_TAX_REG_NO
    if not tax_id:
        raise ValueError("環境變數 COMPANY_TAX_ID 未設定")
    if not tax_reg_no:
        raise ValueError("環境變數 COMPANY_TAX_REG_NO 未設定")

    # 取得已確認的進項記錄
    stagings = sm.get_confirmed_stagings(tax_period)
    if not stagings:
        raise ValueError(f"稅期 {tax_period} 無已確認之進項記錄")

    # 篩選有統編發票的記錄（市場免用發票不列入 MOF TXT）
    invoice_stagings = [
        s for s in stagings
        if s.get("invoice_type") != "免用發票"
        and s.get("supplier_tax_id")
    ]

    if not invoice_stagings:
        raise ValueError(f"稅期 {tax_period} 無可匯出之發票記錄（皆為免用發票或缺統編）")

    # 解析稅期的民國年與月份
    start_month, end_month = get_tax_period_months(tax_period)
    year_part = tax_period.split("-")[0]
    roc_year = int(year_part) - 1911  # 西元轉民國

    # 確保輸出目錄存在
    os.makedirs(output_dir, exist_ok=True)

    lines = []
    total_amount = 0
    total_tax = 0

    for seq, staging in enumerate(invoice_stagings, start=1):
        line = _format_mof_line(
            staging=staging,
            seq=seq,
            roc_year=roc_year,
            end_month=end_month,
            buyer_tax_id=tax_id,
            tax_reg_no=tax_reg_no,
        )
        assert len(line) == 81, f"MOF 行長度錯誤: {len(line)} (應為 81)"
        lines.append(line)
        total_amount += int(staging.get("subtotal", 0))
        total_tax += int(staging.get("tax_amount", 0))

    # 寫入檔案（ASCII 編碼，CRLF 換行符合 DOS 格式）
    filename = f"MOF_PURCHASE_{tax_period}.txt"
    filepath = os.path.join(output_dir, filename)
    with open(filepath, "w", encoding="ascii", newline="") as f:
        for line in lines:
            f.write(line + "\r\n")

    # 記錄匯出紀錄
    sm.add_tax_export(
        tax_period=tax_period,
        export_type="MOF_TXT",
        file_path=filepath,
        record_count=len(lines),
        total_amount=total_amount,
        total_tax=total_tax,
    )

    logger.info(f"MOF TXT 匯出完成: {filepath} ({len(lines)} 筆)")
    return filepath


def _format_mof_line(
    staging: dict,
    seq: int,
    roc_year: int,
    end_month: int,
    buyer_tax_id: str,
    tax_reg_no: str,
) -> str:
    """格式化單筆 MOF 固定寬度行（81 bytes）。

    欄位定義：
    1-2    格式代號 (2) AN    21/22/25
    3-11   稅籍編號 (9) AN    左補零
    12-18  流水號   (7) AN    左補零
    19-21  所屬年度 (3) N     民國年，左補零
    22-23  所屬月份 (2) N     月份
    24-31  買受人統編 (8) AN   我方統編
    32-39  銷售人統編 (8) AN   供應商統編
    40-41  發票字軌 (2) A     英文字母
    42-49  發票號碼 (8) N     左補零
    50-61  銷售額   (12) N    左補零（未稅金額）
    62     課稅別   (1) AN    1/2/3
    63-72  稅額     (10) N    左補零
    73     扣抵代號 (1) AN    1/2/3/4
    74-78  保留     (5) AN    空白
    79     特種稅率 (1) AN    空白
    80     彙加註記 (1) A     空白
    81     通關註記 (1) AN    空白
    """
    # 格式代號：預設從 staging 取，fallback 21（三聯式）
    format_code = str(staging.get("invoice_format_code", "21")).ljust(2)[:2]

    # 稅籍編號：9 碼，左補零
    reg_no = tax_reg_no.zfill(9)[:9]

    # 流水號：7 碼，左補零
    seq_no = str(seq).zfill(7)[:7]

    # 所屬年度：3 碼民國年
    year_str = str(roc_year).zfill(3)[:3]

    # 所屬月份：2 碼（使用稅期結束月份）
    month_str = str(end_month).zfill(2)[:2]

    # 買受人統編（我方）：8 碼
    buyer_id = buyer_tax_id.ljust(8)[:8]

    # 銷售人統編（供應商）：8 碼
    seller_id = str(staging.get("supplier_tax_id", "")).ljust(8)[:8]

    # 發票字軌：2 碼英文字母
    prefix = str(staging.get("invoice_prefix", "")).ljust(2)[:2]

    # 發票號碼：8 碼數字，左補零
    inv_num = str(staging.get("invoice_number", "")).zfill(8)[:8]

    # 銷售額（未稅）：12 碼整數，左補零
    subtotal = int(staging.get("subtotal", 0))
    subtotal_str = str(abs(subtotal)).zfill(12)[:12]

    # 課稅別：1 碼
    tax_type = str(staging.get("tax_type", "1"))[:1]

    # 稅額：10 碼整數，左補零
    tax_amt = int(staging.get("tax_amount", 0))
    tax_str = str(abs(tax_amt)).zfill(10)[:10]

    # 扣抵代號：1 碼
    deduction = str(staging.get("deduction_code", "1"))[:1]

    # 保留（5 碼空白）+ 特種稅率（1 碼空白）+ 彙加註記（1 碼空白）+ 通關註記（1 碼空白）
    reserved = " " * 5
    special_rate = " "
    aggregate_mark = " "
    customs_mark = " "

    line = (
        f"{format_code}"      # 1-2    (2)
        f"{reg_no}"            # 3-11   (9)
        f"{seq_no}"            # 12-18  (7)
        f"{year_str}"          # 19-21  (3)
        f"{month_str}"         # 22-23  (2)
        f"{buyer_id}"          # 24-31  (8)
        f"{seller_id}"         # 32-39  (8)
        f"{prefix}"            # 40-41  (2)
        f"{inv_num}"           # 42-49  (8)
        f"{subtotal_str}"      # 50-61  (12)
        f"{tax_type}"          # 62     (1)
        f"{tax_str}"           # 63-72  (10)
        f"{deduction}"         # 73     (1)
        f"{reserved}"          # 74-78  (5)
        f"{special_rate}"      # 79     (1)
        f"{aggregate_mark}"    # 80     (1)
        f"{customs_mark}"      # 81     (1)
    )

    return line


# =====================================================================
# 2. 文中資訊 Excel 匯出
# =====================================================================

def export_winton_excel(tax_period: str, output_dir: str) -> str:
    """匯出文中資訊會計系統用 Excel 檔。

    欄位 A-K：日期、傳票號碼、摘要、借方科目、借方金額、
    貸方科目、貸方金額、進項稅額、發票號碼、統一編號、備註

    Args:
        tax_period: 稅期，如 "2026-01-02"
        output_dir: 輸出目錄路徑

    Returns:
        產出檔案的完整路徑
    """
    stagings = sm.get_confirmed_stagings(tax_period)
    if not stagings:
        raise ValueError(f"稅期 {tax_period} 無已確認之進項記錄")

    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"進項_{tax_period}"

    # 標題列（粗體）
    headers = [
        "日期", "傳票號碼", "摘要", "借方科目", "借方金額",
        "貸方科目", "貸方金額", "進項稅額", "發票號碼", "統一編號", "備註", "扣抵類別",
    ]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 資料列
    total_amount = 0
    total_tax = 0
    voucher_seq = 0

    for staging in stagings:
        voucher_seq += 1
        row_num = voucher_seq + 1  # 第 1 列是標題

        # A: 日期（YYYY/MM/DD）
        purchase_date = staging.get("purchase_date", "")
        date_display = purchase_date.replace("-", "/") if purchase_date else ""

        # B: 傳票號碼（自動遞增）
        voucher_no = f"V{str(voucher_seq).zfill(4)}"

        # C: 摘要 = 供應商名稱 + 品類摘要
        supplier_name = staging.get("supplier_name", "")
        category_summary = _get_category_summary(staging["id"])
        summary_text = f"{supplier_name} {category_summary}".strip()

        # D: 借方科目 — 從品項類別對應會計科目
        account_code = _get_primary_account_code(staging["id"])

        # E: 借方金額（未稅）
        subtotal = staging.get("subtotal", 0) or 0

        # F: 貸方科目（應付帳款）
        payable_code = "2101"

        # G: 貸方金額（含稅）
        total_amt = staging.get("total_amount", 0) or 0

        # H: 進項稅額
        tax_amt = staging.get("tax_amount", 0) or 0

        # I: 發票號碼
        inv_prefix = staging.get("invoice_prefix", "")
        inv_number = staging.get("invoice_number", "")
        invoice_display = f"{inv_prefix}{inv_number}" if inv_prefix else ""

        # J: 統一編號（供應商）
        supplier_tax_id = staging.get("supplier_tax_id", "")

        # K: 備註
        notes = staging.get("notes", "")

        # L: 扣抵類別
        deduction_code = str(staging.get("deduction_code", "1"))
        deduction_label = "可扣抵" if deduction_code == "1" else "不可扣抵"

        ws.cell(row=row_num, column=1, value=date_display)
        ws.cell(row=row_num, column=2, value=voucher_no)
        ws.cell(row=row_num, column=3, value=summary_text)
        ws.cell(row=row_num, column=4, value=account_code)
        ws.cell(row=row_num, column=5, value=round(subtotal))
        ws.cell(row=row_num, column=6, value=payable_code)
        ws.cell(row=row_num, column=7, value=round(total_amt))
        ws.cell(row=row_num, column=8, value=round(tax_amt))
        ws.cell(row=row_num, column=9, value=invoice_display)
        ws.cell(row=row_num, column=10, value=supplier_tax_id)
        ws.cell(row=row_num, column=11, value=notes)
        ws.cell(row=row_num, column=12, value=deduction_label)

        total_amount += subtotal
        total_tax += tax_amt

    # 欄寬自動調整
    column_widths = [12, 12, 30, 12, 14, 12, 14, 12, 16, 12, 20, 12]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 數字欄靠右
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=8):
        for cell in row:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")

    filename = f"WINTON_{tax_period}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

    # 記錄匯出紀錄
    sm.add_tax_export(
        tax_period=tax_period,
        export_type="WINTON_EXCEL",
        file_path=filepath,
        record_count=voucher_seq,
        total_amount=total_amount,
        total_tax=total_tax,
    )

    logger.info(f"文中 Excel 匯出完成: {filepath} ({voucher_seq} 筆)")
    return filepath


def _get_category_summary(staging_id: int) -> str:
    """取得該筆採購的品類摘要（如「蔬菜/肉類」）。"""
    items = sm.get_purchase_items(staging_id)
    if not items:
        return ""
    categories = list(dict.fromkeys(item.get("category", "其他") for item in items))
    return "/".join(categories[:3])  # 最多顯示 3 個類別


def _get_primary_account_code(staging_id: int) -> str:
    """取得該筆採購的主要會計科目代碼。

    以金額最大的品項類別為主。若查不到對應科目則預設 5110。
    """
    items = sm.get_purchase_items(staging_id)
    if not items:
        return "5110"

    # 以金額最大的品項決定主要科目
    best_item = max(items, key=lambda x: x.get("amount", 0))
    category = best_item.get("category", "其他")

    mapping = sm.get_account_mapping(category)
    if mapping:
        return mapping["account_code"]

    # fallback: 使用品項自身的 account_code
    return best_item.get("account_code", "5110")


# =====================================================================
# 3. 經手人證明表 PDF
# =====================================================================

def export_handler_cert(tax_period: str, output_dir: str) -> str:
    """匯出經手人證明表 PDF（免用發票之市場採購）。

    每日一頁，包含品名、數量、單位、單價、金額明細表。

    Args:
        tax_period: 稅期，如 "2026-01-02"
        output_dir: 輸出目錄路徑

    Returns:
        產出檔案的完整路徑

    Raises:
        ValueError: 無符合條件的市場採購記錄
    """
    stagings = sm.get_confirmed_stagings(tax_period)

    # 篩選免用發票的市場採購
    market_stagings = [
        s for s in stagings
        if _is_market_purchase(s)
    ]

    if not market_stagings:
        raise ValueError(f"稅期 {tax_period} 無免用發票之市場採購記錄")

    os.makedirs(output_dir, exist_ok=True)

    # 依日期分組
    by_date: dict[str, list[dict]] = {}
    for s in market_stagings:
        d = s.get("purchase_date", "unknown")
        by_date.setdefault(d, []).append(s)

    # 註冊中文字型
    font_name = _register_cjk_font()

    filename = f"HANDLER_CERT_{tax_period}.pdf"
    filepath = os.path.join(output_dir, filename)

    company_name = COMPANY_NAME or "（公司名稱未設定）"

    doc = SimpleDocTemplate(
        filepath,
        pagesize=A4,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
    )

    # 樣式定義
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CertTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=18,
        leading=24,
        alignment=1,  # 置中
    )
    subtitle_style = ParagraphStyle(
        "CertSubtitle",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=12,
        leading=16,
        alignment=1,
    )
    normal_style = ParagraphStyle(
        "CertNormal",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=10,
        leading=14,
    )
    footer_style = ParagraphStyle(
        "CertFooter",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=11,
        leading=16,
    )

    elements = []
    total_amount_all = 0
    total_records = 0

    sorted_dates = sorted(by_date.keys())
    for date_idx, purchase_date in enumerate(sorted_dates):
        day_stagings = by_date[purchase_date]

        # 標題
        elements.append(Paragraph(company_name, title_style))
        elements.append(Spacer(1, 4 * mm))
        elements.append(Paragraph("經手人證明表", title_style))
        elements.append(Spacer(1, 4 * mm))

        # 日期
        date_display = purchase_date.replace("-", "/")
        elements.append(Paragraph(f"日期：{date_display}", subtitle_style))
        elements.append(Spacer(1, 6 * mm))

        # 收集所有品項
        all_items = []
        day_total = 0
        for staging in day_stagings:
            items = sm.get_purchase_items(staging["id"])
            handler = staging.get("handler_name", "")
            supplier = staging.get("supplier_name", "")
            for item in items:
                all_items.append({
                    "item_name": item.get("item_name", ""),
                    "quantity": item.get("quantity", 0),
                    "unit": item.get("unit", ""),
                    "unit_price": item.get("unit_price", 0),
                    "amount": item.get("amount", 0),
                    "handler": handler,
                    "supplier": supplier,
                })
                day_total += item.get("amount", 0)

        total_amount_all += day_total
        total_records += len(all_items)

        # 表格
        table_data = [["品名", "數量", "單位", "單價", "金額"]]
        for item in all_items:
            table_data.append([
                item["item_name"],
                _format_number(item["quantity"]),
                item["unit"],
                _format_number(item["unit_price"]),
                _format_number(item["amount"]),
            ])

        # 合計列
        table_data.append(["合計", "", "", "", _format_number(day_total)])

        table = Table(
            table_data,
            colWidths=[6 * cm, 3 * cm, 2.5 * cm, 3 * cm, 3.5 * cm],
        )

        table_style = TableStyle([
            # 表頭
            ("FONTNAME", (0, 0), (-1, 0), font_name),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.85, 0.85, 0.85)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            # 內容
            ("FONTNAME", (0, 1), (-1, -1), font_name),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),
            ("ALIGN", (2, 1), (2, -1), "CENTER"),
            # 合計列
            ("FONTNAME", (0, -1), (-1, -1), font_name),
            ("BACKGROUND", (0, -1), (-1, -1), colors.Color(0.92, 0.92, 0.92)),
            ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black),
            # 框線
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ])
        table.setStyle(table_style)
        elements.append(table)

        elements.append(Spacer(1, 10 * mm))

        # 合計金額文字
        elements.append(Paragraph(
            f"合計金額：新臺幣 {_format_number(day_total)} 元整",
            footer_style,
        ))
        elements.append(Spacer(1, 15 * mm))

        # 簽章欄
        # 收集該日所有經手人（去重）
        handlers = list(dict.fromkeys(
            s.get("handler_name", "") for s in day_stagings if s.get("handler_name")
        ))
        handler_display = "、".join(handlers) if handlers else "_______________"

        sig_data = [
            ["經手人簽名", "主管覆核", "公司章"],
            [handler_display, "", ""],
        ]
        sig_table = Table(sig_data, colWidths=[6 * cm, 6 * cm, 6 * cm])
        sig_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("TOPPADDING", (0, 0), (-1, 0), 6),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("TOPPADDING", (0, 1), (-1, 1), 20),
            ("BOTTOMPADDING", (0, 1), (-1, 1), 20),
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.92, 0.92, 0.92)),
        ]))
        elements.append(sig_table)

        # 非最後一日則換頁
        if date_idx < len(sorted_dates) - 1:
            from reportlab.platypus import PageBreak
            elements.append(PageBreak())

    doc.build(elements)

    # 記錄匯出紀錄
    sm.add_tax_export(
        tax_period=tax_period,
        export_type="HANDLER_CERT",
        file_path=filepath,
        record_count=total_records,
        total_amount=total_amount_all,
        total_tax=0,
    )

    logger.info(f"經手人證明表匯出完成: {filepath} ({len(sorted_dates)} 日, {total_records} 品項)")
    return filepath


def _is_market_purchase(staging: dict) -> bool:
    """判斷是否為免用發票的市場採購。"""
    # 明確標記免用發票
    if staging.get("invoice_type") == "免用發票":
        return True

    # 供應商無統一發票
    supplier_id = staging.get("supplier_id")
    if supplier_id:
        supplier = sm.get_supplier(supplier_id=supplier_id)
        if supplier and not supplier.get("has_uniform_invoice", 1):
            return True

    return False


def _format_number(value) -> str:
    """格式化數字顯示（整數不帶小數，浮點數保留適當位數）。"""
    if value is None:
        return "0"
    if isinstance(value, float):
        if value == int(value):
            return f"{int(value):,}"
        return f"{value:,.2f}"
    return f"{int(value):,}"


# =====================================================================
# 4. 匯出前驗證（7 點檢核）
# =====================================================================

def validate_before_export(tax_period: str) -> tuple[bool, list[str]]:
    """匯出前 7 點驗證檢核。

    檢核項目：
    1. 所有暫存記錄皆已確認（無 pending）
    2. 每張發票有供應商統編（市場採購除外）
    3. 發票號碼格式正確（2 英文 + 8 數字）
    4. 稅額驗算：|tax_amount - subtotal * 0.05| <= 1
    5. 借貸平衡：sum(借方) == sum(貸方)
    6. 期別一致：所有記錄歸屬同一稅期
    7. 市場採購有填經手人

    Args:
        tax_period: 稅期，如 "2026-01-02"

    Returns:
        (True, []) 全部通過
        (False, [錯誤訊息列表]) 有未通過項目
    """
    errors: list[str] = []

    # 取得該稅期所有暫存記錄（含 pending 與 confirmed）
    confirmed = sm.get_confirmed_stagings(tax_period)

    # --- 檢核 1: 所有記錄皆已確認 ---
    pending = _get_pending_for_period(tax_period)
    if pending:
        errors.append(
            f"[1] 有 {len(pending)} 筆未確認記錄（staging_id: "
            f"{', '.join(str(p['id']) for p in pending)}）"
        )

    # 合併全部記錄做後續檢核
    all_stagings = confirmed  # 主要檢核已確認的

    if not all_stagings:
        errors.append(f"[0] 稅期 {tax_period} 無任何已確認記錄")
        return (False, errors)

    # --- 檢核 2: 發票需有供應商統編（市場採購除外）---
    for s in all_stagings:
        if _is_market_purchase(s):
            continue  # 市場採購豁免統編檢查
        tax_id = s.get("supplier_tax_id", "")
        if not tax_id or not tax_id.strip():
            errors.append(
                f"[2] staging_id={s['id']} 缺少供應商統編"
                f"（供應商: {s.get('supplier_name', 'N/A')}）"
            )

    # --- 檢核 3: 發票號碼格式（2 英文 + 8 數字）---
    invoice_pattern = re.compile(r"^[A-Z]{2}$")
    number_pattern = re.compile(r"^\d{8}$")
    for s in all_stagings:
        if _is_market_purchase(s):
            continue  # 市場採購無發票號碼
        prefix = s.get("invoice_prefix", "")
        number = s.get("invoice_number", "")
        if prefix or number:  # 有填寫才檢核
            if not invoice_pattern.match(prefix):
                errors.append(
                    f"[3] staging_id={s['id']} 發票字軌格式錯誤: '{prefix}'"
                    f"（應為 2 碼大寫英文）"
                )
            if not number_pattern.match(str(number).zfill(8)):
                # 嘗試補零後再驗
                raw = str(number)
                if not raw.isdigit() or len(raw) > 8:
                    errors.append(
                        f"[3] staging_id={s['id']} 發票號碼格式錯誤: '{number}'"
                        f"（應為 8 碼數字）"
                    )

    # --- 檢核 4: 稅額驗算 ---
    for s in all_stagings:
        if _is_market_purchase(s):
            continue  # 免用發票無稅額
        tax_type = str(s.get("tax_type", "1"))
        if tax_type != "1":
            continue  # 僅驗算應稅（tax_type=1）
        subtotal = s.get("subtotal", 0) or 0
        tax_amount = s.get("tax_amount", 0) or 0
        expected_tax = subtotal * 0.05
        if abs(tax_amount - expected_tax) > 1:
            errors.append(
                f"[4] staging_id={s['id']} 稅額異常: "
                f"稅額={tax_amount}, 預期={expected_tax:.0f} "
                f"(未稅={subtotal})"
            )

    # --- 檢核 5: 借貸平衡 ---
    total_debit = 0   # 借方 = 未稅金額 + 進項稅額
    total_credit = 0  # 貸方 = 含稅金額
    for s in all_stagings:
        subtotal = s.get("subtotal", 0) or 0
        tax_amount = s.get("tax_amount", 0) or 0
        total_amount = s.get("total_amount", 0) or 0
        total_debit += subtotal + tax_amount
        total_credit += total_amount

    if abs(total_debit - total_credit) > 1:
        errors.append(
            f"[5] 借貸不平衡: 借方合計={total_debit:.0f}, "
            f"貸方合計={total_credit:.0f}, "
            f"差額={abs(total_debit - total_credit):.0f}"
        )

    # --- 檢核 6: 期別一致 ---
    for s in all_stagings:
        record_period = s.get("tax_period", "")
        if record_period and record_period != tax_period:
            errors.append(
                f"[6] staging_id={s['id']} 稅期不一致: "
                f"記錄={record_period}, 目標={tax_period}"
            )

    # --- 檢核 7: 市場採購需填經手人 ---
    for s in all_stagings:
        if not _is_market_purchase(s):
            continue
        handler = s.get("handler_name", "")
        if not handler or not handler.strip():
            errors.append(
                f"[7] staging_id={s['id']} 市場採購缺經手人"
                f"（供應商: {s.get('supplier_name', 'N/A')}）"
            )

    # --- 檢核 8: 扣抵代號一致性 ---
    for s in all_stagings:
        if _is_market_purchase(s):
            continue
        tax_id = s.get("supplier_tax_id", "").strip()
        inv_type = s.get("invoice_type", "")
        deduction = str(s.get("deduction_code", "1"))

        # 有統編 + 三聯式/電子發票 → deduction_code 應為 "1"
        if tax_id and inv_type in ("三聯式", "電子發票") and deduction != "1":
            errors.append(
                f"[8] staging_id={s['id']} 扣抵代號不一致: "
                f"有統編+{inv_type} 但 deduction_code={deduction}（應為 1）"
            )
        # 無統編 → deduction_code 應為 "2"
        if not tax_id and deduction == "1":
            errors.append(
                f"[8] staging_id={s['id']} 扣抵代號不一致: "
                f"無統編但 deduction_code=1（應為 2）"
            )

    is_valid = len(errors) == 0
    if is_valid:
        logger.info(f"稅期 {tax_period} 驗證通過（{len(all_stagings)} 筆）")
    else:
        logger.warning(f"稅期 {tax_period} 驗證失敗: {len(errors)} 項錯誤")
        for err in errors:
            logger.warning(f"  {err}")

    return (is_valid, errors)


def _get_pending_for_period(tax_period: str) -> list[dict]:
    """取得該稅期的 pending 記錄。

    因為 state_manager 的 get_pending_stagings 是依 chat_id 篩選，
    這裡直接查詢以稅期篩選。
    """
    from state_manager import _get_conn
    conn = _get_conn()
    rows = conn.execute(
        "SELECT * FROM purchase_staging WHERE status='pending' AND tax_period=?",
        (tax_period,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# =====================================================================
# 5. 稅期工具函數
# =====================================================================

def get_tax_period_months(tax_period: str) -> tuple[int, int]:
    """解析稅期字串，回傳起迄月份。

    台灣營業稅為雙月制：1-2月、3-4月、5-6月、7-8月、9-10月、11-12月。

    Args:
        tax_period: 格式 "YYYY-MM-MM"，如 "2026-01-02"

    Returns:
        (起始月, 結束月) 的 tuple，如 (1, 2)

    Raises:
        ValueError: 格式不正確
    """
    parts = tax_period.split("-")
    if len(parts) != 3:
        raise ValueError(f"稅期格式錯誤: '{tax_period}'（應為 YYYY-MM-MM）")

    try:
        start_month = int(parts[1])
        end_month = int(parts[2])
    except ValueError:
        raise ValueError(f"稅期月份解析失敗: '{tax_period}'")

    # 驗證雙月制合理性
    if start_month < 1 or start_month > 11 or start_month % 2 != 1:
        raise ValueError(f"起始月份不合法: {start_month}（應為奇數月 1-11）")
    if end_month != start_month + 1:
        raise ValueError(f"結束月份不合法: {end_month}（應為 {start_month + 1}）")

    return (start_month, end_month)
