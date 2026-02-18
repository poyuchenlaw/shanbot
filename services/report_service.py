"""報表生成服務 — 採購報告、預算報告、月報表"""

import logging
import os
from datetime import datetime
from typing import Optional

import state_manager as sm

logger = logging.getLogger("shanbot.report")


def generate_purchase_report(year_month: str, output_dir: str = None) -> Optional[str]:
    """生成採購報告 Excel（成本 + 比價 + 建議）"""
    stagings = sm.get_stagings_by_month(year_month)
    if not stagings:
        logger.warning(f"No stagings for {year_month}")
        return None

    if not output_dir:
        output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                                   "data", "reports")
    os.makedirs(output_dir, exist_ok=True)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        logger.error("openpyxl not installed")
        return None

    wb = Workbook()

    # === Sheet 1: 採購彙總 ===
    ws1 = wb.active
    ws1.title = "採購彙總"

    # 標題
    ws1.append([f"{year_month} 採購報告"])
    ws1.merge_cells("A1:G1")
    ws1["A1"].font = Font(size=14, bold=True)
    ws1.append([])  # 空行

    # 表頭
    headers = ["日期", "供應商", "品項數", "未稅金額", "稅額", "含稅總額", "發票號碼"]
    ws1.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row=3, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", fill_type="solid")

    # 資料
    total_subtotal = 0
    total_tax = 0
    total_amount = 0

    for s in stagings:
        items = sm.get_purchase_items(s["id"])
        invoice = f"{s.get('invoice_prefix', '')}{s.get('invoice_number', '')}"
        ws1.append([
            s["purchase_date"],
            s["supplier_name"],
            len(items),
            s["subtotal"],
            s["tax_amount"],
            s["total_amount"],
            invoice,
        ])
        total_subtotal += s.get("subtotal", 0) or 0
        total_tax += s.get("tax_amount", 0) or 0
        total_amount += s.get("total_amount", 0) or 0

    # 合計列
    ws1.append([])
    ws1.append(["合計", "", len(stagings), total_subtotal, total_tax, total_amount, ""])
    last_row = ws1.max_row
    for col in range(1, len(headers) + 1):
        ws1.cell(row=last_row, column=col).font = Font(bold=True)

    # 欄寬
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 8
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 10
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 14

    # === Sheet 2: 品項明細 ===
    ws2 = wb.create_sheet("品項明細")
    ws2.append(["日期", "供應商", "品名", "數量", "單位", "單價", "金額", "分類", "會計科目"])
    for col in range(1, 10):
        ws2.cell(row=1, column=col).font = Font(bold=True)
        ws2.cell(row=1, column=col).fill = PatternFill(start_color="E0E0E0", fill_type="solid")

    for s in stagings:
        items = sm.get_purchase_items(s["id"])
        for item in items:
            ws2.append([
                s["purchase_date"],
                s["supplier_name"],
                item["item_name"],
                item["quantity"],
                item["unit"],
                item["unit_price"],
                item["amount"],
                item["category"],
                item["account_code"],
            ])

    # === Sheet 3: 分類統計 ===
    ws3 = wb.create_sheet("分類統計")
    ws3.append(["分類", "會計科目", "筆數", "金額"])
    for col in range(1, 5):
        ws3.cell(row=1, column=col).font = Font(bold=True)
        ws3.cell(row=1, column=col).fill = PatternFill(start_color="E0E0E0", fill_type="solid")

    # 統計分類
    category_stats = {}
    for s in stagings:
        items = sm.get_purchase_items(s["id"])
        for item in items:
            cat = item.get("category", "其他")
            if cat not in category_stats:
                category_stats[cat] = {"count": 0, "amount": 0, "account_code": item.get("account_code", "")}
            category_stats[cat]["count"] += 1
            category_stats[cat]["amount"] += item.get("amount", 0) or 0

    for cat, stats in sorted(category_stats.items()):
        ws3.append([cat, stats["account_code"], stats["count"], stats["amount"]])

    # 儲存
    filepath = os.path.join(output_dir, f"採購報告_{year_month}.xlsx")
    wb.save(filepath)
    logger.info(f"Purchase report saved: {filepath}")
    return filepath


def generate_monthly_report(year_month: str, output_dir: str = None) -> Optional[str]:
    """生成月報表 Excel（暫存統整 + 分類匯總 + 憑證目錄）"""
    stagings = sm.get_stagings_by_month(year_month)
    if not stagings:
        return None

    if not output_dir:
        output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                                   "data", "reports")
    os.makedirs(output_dir, exist_ok=True)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return None

    wb = Workbook()

    # Sheet 1: 月度總覽
    ws1 = wb.active
    ws1.title = "月度總覽"

    total_subtotal = sum(s.get("subtotal", 0) or 0 for s in stagings)
    total_tax = sum(s.get("tax_amount", 0) or 0 for s in stagings)
    total_amount = sum(s.get("total_amount", 0) or 0 for s in stagings)
    invoice_count = sum(1 for s in stagings if s.get("invoice_number"))
    receipt_count = sum(1 for s in stagings if not s.get("invoice_number"))

    ws1.append([f"{year_month} 月報表"])
    ws1["A1"].font = Font(size=14, bold=True)
    ws1.append([])
    ws1.append(["項目", "數值"])
    ws1.append(["採購筆數", len(stagings)])
    ws1.append(["發票張數", invoice_count])
    ws1.append(["收據/經手人證明", receipt_count])
    ws1.append(["未稅總額", total_subtotal])
    ws1.append(["進項稅額", total_tax])
    ws1.append(["含稅總額", total_amount])

    # 更新月度成本表
    sm.upsert_monthly_cost(
        year_month,
        ingredient_total=total_subtotal,
        taxable_purchase_total=total_subtotal,
        input_tax_total=total_tax,
        deductible_tax=total_tax,
        invoice_count=invoice_count,
        receipt_count=receipt_count,
    )

    # Sheet 2: 憑證目錄
    ws2 = wb.create_sheet("憑證目錄")
    ws2.append(["#", "日期", "供應商", "發票號碼", "金額", "稅額", "狀態"])
    for col in range(1, 8):
        ws2.cell(row=1, column=col).font = Font(bold=True)

    for i, s in enumerate(stagings, 1):
        invoice = f"{s.get('invoice_prefix', '')}{s.get('invoice_number', '')}"
        ws2.append([
            i, s["purchase_date"], s["supplier_name"],
            invoice or "(無發票)",
            s["total_amount"], s["tax_amount"], s["status"],
        ])

    filepath = os.path.join(output_dir, f"月報表_{year_month}.xlsx")
    wb.save(filepath)
    logger.info(f"Monthly report saved: {filepath}")
    return filepath


def generate_annual_report(year: str, output_dir: str = None) -> Optional[str]:
    """生成年度報表 Excel（12 個月彙總）"""
    if not output_dir:
        output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                                   "data", "reports")
    os.makedirs(output_dir, exist_ok=True)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        logger.error("openpyxl not installed")
        return None

    wb = Workbook()

    # Sheet 1: 月度彙總
    ws1 = wb.active
    ws1.title = "月度彙總"

    ws1.append([f"{year} 年度報表"])
    ws1.merge_cells("A1:H1")
    ws1["A1"].font = Font(size=14, bold=True)
    ws1.append([])

    headers = ["月份", "採購筆數", "發票數", "收據數", "未稅總額", "進項稅額", "含稅總額", "狀態"]
    ws1.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row=3, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", fill_type="solid")

    grand_subtotal = 0
    grand_tax = 0
    grand_total = 0
    has_data = False

    for m in range(1, 13):
        ym = f"{year}-{m:02d}"
        stats = sm.get_staging_stats(ym)
        total_count = stats.get("total", 0) or 0
        if total_count == 0:
            continue

        has_data = True
        mc = sm.get_monthly_cost(ym)
        inv_count = mc.get("invoice_count", 0) if mc else 0
        rec_count = mc.get("receipt_count", 0) if mc else 0
        subtotal = stats.get("total_amount", 0) or 0
        tax = stats.get("total_tax", 0) or 0
        total_amt = subtotal  # total_amount already includes tax
        is_locked = "已鎖定" if mc and mc.get("is_locked") else "開放"

        ws1.append([ym, total_count, inv_count, rec_count,
                     subtotal - tax, tax, subtotal, is_locked])
        grand_subtotal += (subtotal - tax)
        grand_tax += tax
        grand_total += subtotal

    if not has_data:
        return None

    ws1.append([])
    ws1.append(["合計", "", "", "", grand_subtotal, grand_tax, grand_total, ""])
    last_row = ws1.max_row
    for col in range(1, len(headers) + 1):
        ws1.cell(row=last_row, column=col).font = Font(bold=True)

    for c, w in [("A", 10), ("B", 10), ("C", 8), ("D", 8),
                  ("E", 14), ("F", 12), ("G", 14), ("H", 8)]:
        ws1.column_dimensions[c].width = w

    # Sheet 2: 分類年度統計
    ws2 = wb.create_sheet("分類統計")
    ws2.append(["分類", "會計科目", "年度筆數", "年度金額"])
    for col in range(1, 5):
        ws2.cell(row=1, column=col).font = Font(bold=True)
        ws2.cell(row=1, column=col).fill = PatternFill(start_color="E0E0E0",
                                                         fill_type="solid")

    cat_totals = {}
    for m in range(1, 13):
        ym = f"{year}-{m:02d}"
        stagings = sm.get_stagings_by_month(ym)
        for s in stagings:
            items = sm.get_purchase_items(s["id"])
            for item in items:
                cat = item.get("category", "其他")
                if cat not in cat_totals:
                    cat_totals[cat] = {"count": 0, "amount": 0,
                                       "account_code": item.get("account_code", "")}
                cat_totals[cat]["count"] += 1
                cat_totals[cat]["amount"] += item.get("amount", 0) or 0

    for cat, data in sorted(cat_totals.items(), key=lambda x: -x[1]["amount"]):
        ws2.append([cat, data["account_code"], data["count"], data["amount"]])

    filepath = os.path.join(output_dir, f"年度報表_{year}.xlsx")
    wb.save(filepath)
    logger.info(f"Annual report saved: {filepath}")
    return filepath
