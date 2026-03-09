"""薪資/人事管理服務 — 薪資表生成、契約解析、勞健保計算"""

import calendar
import json
import logging
import os
import re
from datetime import datetime
from typing import Optional

logger = logging.getLogger("shanbot.salary")

# === Taiwan Labor Law Constants (2025) ===

MINIMUM_WAGE_MONTHLY = 28590
MINIMUM_WAGE_HOURLY = 190

# 勞保費率：普通事故 11% + 就業保險 1% = 12%
# 雇主 70%, 勞工 20%, 政府 10%
LABOR_INS_RATE = 0.12
LABOR_INS_EMPLOYEE_SHARE = 0.20  # employee pays 20% of total rate

# 健保費率 5.17%, 員工負擔 30%
HEALTH_INS_RATE = 0.0517
HEALTH_INS_EMPLOYEE_SHARE = 0.30

# 勞退雇主提繳 6%（不從員工薪資扣除）
PENSION_EMPLOYER_RATE = 0.06

# 伙食津貼免稅上限
MEAL_ALLOWANCE_TAX_FREE = 3000

# 加班費倍率
OVERTIME_RATE_FIRST_2H = 1 + 1 / 3     # 前2小時 = 1.34
OVERTIME_RATE_NEXT_2H = 1 + 2 / 3      # 後2小時 = 1.67

# 勞保投保級距表 (2025)
LABOR_INS_GRADES = [
    28590, 29100, 30300, 31800, 33300, 34800, 36300, 38200,
    40100, 42000, 43900, 45800, 48200, 50600, 53000, 55400,
    57800, 60800, 63800, 66800, 69800, 72800, 76500, 80200,
    83900, 87600, 92100, 96600, 101100, 105600, 110100, 115500,
    120900, 126300, 131700, 137100, 142500, 147900, 150000,
]

# 健保投保級距表 (2025)
HEALTH_INS_GRADES = [
    28590, 29100, 30300, 31800, 33300, 34800, 36300, 38200,
    40100, 42000, 43900, 45800, 48200, 50600, 53000, 55400,
    57800, 60800, 63800, 66800, 69800, 72800, 76500, 80200,
    83900, 87600, 92100, 96600, 101100, 105600, 110100, 115500,
    120900, 126300, 131700, 137100, 142500, 147900, 150000,
    156400, 162800, 169200, 175600, 182000, 189500, 197000,
    204500, 212000, 219500,
]

# 薪資所得扣繳稅額表（簡化版 — 單身無扶養，月薪級距）
# (下限, 上限, 扣繳額)
INCOME_TAX_BRACKETS = [
    (0, 40020, 0),
    (40021, 40530, 40),
    (40531, 42500, 140),
    (42501, 45000, 300),
    (45001, 47500, 460),
    (47501, 50000, 640),
    (50001, 52500, 840),
    (52501, 55000, 1060),
    (55001, 57500, 1280),
    (57501, 60000, 1500),
    (60001, 62500, 1720),
    (62501, 65000, 1960),
    (65001, 67500, 2200),
    (67501, 70000, 2440),
    (70001, 73500, 2780),
    (73501, 76500, 3080),
    (76501, 79500, 3400),
    (79501, 83000, 3770),
    (83001, 86500, 4180),
    (86501, 90000, 4590),
    (90001, 93500, 5020),
    (93501, 97000, 5450),
    (97001, 101000, 5930),
    (101001, 105000, 6430),
    (105001, 110000, 7020),
    (110001, 115000, 7630),
    (115001, 120000, 8260),
    (120001, 125000, 8900),
    (125001, 130000, 9560),
    (130001, 140000, 10700),
    (140001, 150000, 12000),
    (150001, 160000, 13400),
    (160001, 170000, 14800),
    (170001, 190000, 17600),
    (190001, 999999999, 0),  # 超過 19 萬用 5% 計算
]


def _get_base_path() -> str:
    """取得 GDrive 基礎路徑"""
    from services.gdrive_service import GDRIVE_LOCAL
    return GDRIVE_LOCAL


def _find_nearest_grade(salary: int, grade_table: list[int]) -> int:
    """找到最接近（且 >= salary）的投保級距"""
    for grade in grade_table:
        if grade >= salary:
            return grade
    return grade_table[-1]


def mask_id_number(id_number: str) -> str:
    """身分證字號遮罩：A123***789"""
    if not id_number or len(id_number) < 10:
        return id_number or ""
    return f"{id_number[:4]}***{id_number[7:]}"


# === 1. 薪資計算 ===

def calculate_overtime_pay(base_salary: int, overtime_hours: float) -> int:
    """計算加班費

    Args:
        base_salary: 月底薪
        overtime_hours: 加班時數

    Returns:
        加班費金額
    """
    # 時薪 = 月薪 / 30 / 8
    hourly_rate = base_salary / 30 / 8
    total = 0.0

    if overtime_hours <= 2:
        total = overtime_hours * hourly_rate * OVERTIME_RATE_FIRST_2H
    else:
        total = 2 * hourly_rate * OVERTIME_RATE_FIRST_2H
        remaining = min(overtime_hours - 2, 2)
        total += remaining * hourly_rate * OVERTIME_RATE_NEXT_2H

    return round(total)


def calculate_deductions(
    base_salary: int,
    meal_allowance: int = 2400,
    overtime_pay: int = 0,
    bonus: int = 0,
    labor_grade: int = None,
    health_grade: int = None,
    pension_self_rate: float = 0.0,
    dependents: int = 0,
) -> dict:
    """計算台灣勞健保、勞退自提、所得稅扣繳

    Args:
        base_salary: 底薪
        meal_allowance: 伙食津貼
        overtime_pay: 加班費
        bonus: 獎金
        labor_grade: 勞保投保級距（None = 自動計算）
        health_grade: 健保投保級距（None = 自動計算）
        pension_self_rate: 勞退自提比例 (0-6%)
        dependents: 扶養親屬人數

    Returns:
        dict with labor_ins, health_ins, pension_self, income_tax,
             gross_salary, total_deductions, net_salary
    """
    gross_salary = base_salary + meal_allowance + overtime_pay + bonus

    # 勞保自付額
    if labor_grade is None:
        labor_grade = _find_nearest_grade(base_salary, LABOR_INS_GRADES)
    labor_ins = round(labor_grade * LABOR_INS_RATE * LABOR_INS_EMPLOYEE_SHARE)

    # 健保自付額：投保金額 × 5.17% × 30% × (1 + 眷屬人數)
    # 眷屬人數上限 3 人（超過部分由政府補助）
    if health_grade is None:
        health_grade = _find_nearest_grade(base_salary, HEALTH_INS_GRADES)
    effective_dependents = min(dependents, 3)
    health_ins = round(
        health_grade * HEALTH_INS_RATE * HEALTH_INS_EMPLOYEE_SHARE
        * (1 + effective_dependents)
    )

    # 勞退自提（從薪資扣除，但免稅）
    pension_self = 0
    if pension_self_rate > 0:
        pension_grade = _find_nearest_grade(base_salary, LABOR_INS_GRADES)
        pension_self = round(pension_grade * (pension_self_rate / 100))

    # 所得稅扣繳
    # 課稅所得 = 應發 - 伙食津貼免稅額 - 勞退自提（免稅）
    taxable_meal = min(meal_allowance, MEAL_ALLOWANCE_TAX_FREE)
    taxable_income = gross_salary - taxable_meal - pension_self
    income_tax = _calc_withholding_tax(taxable_income, dependents)

    total_deductions = labor_ins + health_ins + pension_self + income_tax
    net_salary = gross_salary - total_deductions

    return {
        "gross_salary": gross_salary,
        "labor_ins": labor_ins,
        "labor_grade": labor_grade,
        "health_ins": health_ins,
        "health_grade": health_grade,
        "pension_self": pension_self,
        "income_tax": income_tax,
        "total_deductions": total_deductions,
        "net_salary": net_salary,
    }


def _calc_withholding_tax(taxable_income: int, dependents: int = 0) -> int:
    """簡易扣繳稅額計算"""
    # 每多一個扶養人，免稅額增加約 1540/月（2025 年）
    adjusted = taxable_income - (dependents * 1540)
    if adjusted <= 0:
        return 0

    for low, high, tax in INCOME_TAX_BRACKETS:
        if low <= adjusted <= high:
            return tax

    # 超過最高級距 → 5% 概算
    if adjusted > 190000:
        return round(adjusted * 0.05)

    return 0


# === 2. Excel 模板生成 ===

def generate_salary_template(year_month: str) -> str:
    """生成薪資表 Excel 模板

    Args:
        year_month: 格式 "YYYY-MM"

    Returns:
        生成的檔案完整路徑
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # === Sheet 1: 員工薪資表 ===
    ws = wb.active
    ws.title = "員工薪資表"

    # Style
    header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(name="微軟正黑體", size=14, bold=True)
    normal_font = Font(name="微軟正黑體", size=10)
    money_fmt = '#,##0'
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:M1")
    ws["A1"] = f"員工薪資表 — {year_month}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:M2")
    ws["A2"] = f"製表日期：{datetime.now().strftime('%Y-%m-%d')}"
    ws["A2"].font = Font(name="微軟正黑體", size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "員工姓名", "底薪", "伙食津貼", "加班時數", "加班費",
        "獎金", "應發合計", "勞保自付", "健保自付",
        "勞退自提", "所得稅", "其他扣款", "實發金額",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Pre-fill 10 empty rows with formulas
    for row in range(4, 14):
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = normal_font
            if col >= 2:
                cell.number_format = money_fmt

        # Formula: 應發合計 (G) = 底薪(B) + 伙食津貼(C) + 加班費(E) + 獎金(F)
        ws.cell(row=row, column=7).value = f"=B{row}+C{row}+E{row}+F{row}"
        # Formula: 實發金額 (M) = 應發合計(G) - 勞保(H) - 健保(I) - 勞退(J) - 所得稅(K) - 其他(L)
        ws.cell(row=row, column=13).value = f"=G{row}-H{row}-I{row}-J{row}-K{row}-L{row}"

    # Totals row
    total_row = 14
    ws.cell(row=total_row, column=1, value="合計").font = Font(name="微軟正黑體", size=11, bold=True)
    for col in range(2, 14):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col)
        if col != 4:  # skip 加班時數 sum
            cell.value = f"=SUM({col_letter}4:{col_letter}13)"
        cell.font = Font(name="微軟正黑體", size=10, bold=True)
        cell.number_format = money_fmt
        cell.border = thin_border

    # Column widths
    widths = [12, 10, 10, 8, 10, 10, 12, 10, 10, 10, 10, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = "A4"

    # === Sheet 2: 說明 ===
    ws2 = wb.create_sheet("說明")
    ws2["A1"] = "台灣勞健保費率參考表（2025年度）"
    ws2["A1"].font = Font(name="微軟正黑體", size=13, bold=True)

    info_data = [
        ("", ""),
        ("【基本工資】", ""),
        ("月薪基本工資", f"{MINIMUM_WAGE_MONTHLY:,} 元"),
        ("時薪基本工資", f"{MINIMUM_WAGE_HOURLY} 元"),
        ("", ""),
        ("【勞工保險】", ""),
        ("費率", "12%（普通事故 11% + 就業保險 1%）"),
        ("雇主負擔", "70%"),
        ("勞工負擔", "20%"),
        ("政府負擔", "10%"),
        ("員工實付", "投保級距 × 12% × 20% = 2.4%"),
        ("", ""),
        ("【全民健保】", ""),
        ("費率", "5.17%"),
        ("員工負擔比例", "30%"),
        ("計算公式", "投保金額 × 5.17% × 30% × (1+眷屬人數)"),
        ("眷屬上限", "3人（超過由政府補助）"),
        ("", ""),
        ("【勞工退休金】", ""),
        ("雇主提繳", "6%（強制，不從員工扣）"),
        ("員工自提", "0~6%（自願，從薪資扣，免稅）"),
        ("", ""),
        ("【伙食津貼】", ""),
        ("免稅上限", f"{MEAL_ALLOWANCE_TAX_FREE:,} 元/月"),
        ("", ""),
        ("【加班費】", ""),
        ("前 2 小時", "時薪 × 1.34"),
        ("第 3~4 小時", "時薪 × 1.67"),
        ("時薪計算", "月薪 ÷ 30 ÷ 8"),
    ]
    for i, (label, value) in enumerate(info_data, 2):
        ws2.cell(row=i, column=1, value=label).font = Font(
            name="微軟正黑體", size=10,
            bold=label.startswith("【"),
        )
        ws2.cell(row=i, column=2, value=value).font = Font(name="微軟正黑體", size=10)

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 50

    # Save
    parts = year_month.split("-")
    year = parts[0]
    month = f"{int(parts[1]):02d}月"
    base = _get_base_path()
    save_dir = os.path.join(base, year, month, "薪資表")
    os.makedirs(save_dir, exist_ok=True)
    filepath = os.path.join(save_dir, f"薪資表_{year_month}.xlsx")
    wb.save(filepath)
    logger.info(f"Salary template generated: {filepath}")
    return filepath


def generate_employee_template() -> str:
    """生成員工資料表 Excel 模板

    Returns:
        生成的檔案完整路徑
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "員工資料表"

    # Style
    header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(name="微軟正黑體", size=14, bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = "員工資料表"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "姓名", "身分證字號", "職稱", "部門", "到職日",
        "底薪", "伙食津貼", "勞保級距", "健保級距",
        "勞退自提%", "扶養人數", "薪轉帳號",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Pre-fill 20 empty rows
    for row in range(3, 23):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

    # Column widths
    widths = [10, 14, 10, 10, 12, 10, 10, 10, 10, 10, 8, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    # Save
    base = _get_base_path()
    save_dir = os.path.join(base, "員工資料")
    os.makedirs(save_dir, exist_ok=True)
    filepath = os.path.join(save_dir, "員工資料表.xlsx")
    wb.save(filepath)
    logger.info(f"Employee template generated: {filepath}")
    return filepath


# === 3. Excel 解析 ===

def parse_salary_excel(file_path: str) -> list[dict]:
    """解析上傳的薪資表 Excel

    Args:
        file_path: Excel 檔案路徑

    Returns:
        list of payroll record dicts
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    records = []
    # 找到表頭列
    header_row = None
    header_map = {}
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            val = str(cell.value or "").strip()
            if val in ("員工姓名", "姓名"):
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        logger.warning(f"Cannot find header row in {file_path}")
        return records

    # Map columns
    column_aliases = {
        "員工姓名": "name", "姓名": "name",
        "底薪": "base_salary",
        "伙食津貼": "meal_allowance",
        "加班時數": "overtime_hours",
        "加班費": "overtime_pay",
        "獎金": "bonus",
        "應發合計": "gross_salary",
        "勞保自付": "labor_insurance",
        "健保自付": "health_insurance",
        "勞退自提": "pension_self",
        "所得稅": "income_tax",
        "其他扣款": "other_deductions",
        "實發金額": "net_salary",
    }

    for cell in ws[header_row]:
        val = str(cell.value or "").strip()
        if val in column_aliases:
            header_map[cell.column] = column_aliases[val]

    # Read data rows
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        record = {}
        for cell in row:
            field = header_map.get(cell.column)
            if field:
                record[field] = cell.value

        # Skip empty rows / total rows
        name = record.get("name")
        if not name or str(name).strip() in ("", "合計", "總計"):
            continue

        # Convert numeric fields
        for f in ["base_salary", "meal_allowance", "overtime_hours", "overtime_pay",
                   "bonus", "gross_salary", "labor_insurance", "health_insurance",
                   "pension_self", "income_tax", "other_deductions", "net_salary"]:
            val = record.get(f)
            if val is not None:
                try:
                    record[f] = float(val) if f == "overtime_hours" else int(float(val))
                except (ValueError, TypeError):
                    record[f] = 0
            else:
                record[f] = 0

        records.append(record)

    logger.info(f"Parsed {len(records)} salary records from {file_path}")
    return records


def parse_employee_excel(file_path: str) -> list[dict]:
    """解析員工資料表 Excel

    Returns:
        list of employee record dicts
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    records = []
    header_row = None
    header_map = {}

    column_aliases = {
        "姓名": "name",
        "身分證字號": "id_number",
        "職稱": "position",
        "部門": "department",
        "到職日": "hire_date",
        "底薪": "base_salary",
        "伙食津貼": "meal_allowance",
        "勞保級距": "labor_insurance_grade",
        "健保級距": "health_insurance_grade",
        "勞退自提%": "pension_self_rate",
        "扶養人數": "tax_dependents",
        "薪轉帳號": "bank_account",
    }

    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            val = str(cell.value or "").strip()
            if val == "姓名":
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        return records

    for cell in ws[header_row]:
        val = str(cell.value or "").strip()
        if val in column_aliases:
            header_map[cell.column] = column_aliases[val]

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        record = {}
        for cell in row:
            field = header_map.get(cell.column)
            if field:
                record[field] = cell.value

        name = record.get("name")
        if not name or str(name).strip() == "":
            continue

        # Type convert
        for f in ["base_salary", "meal_allowance", "labor_insurance_grade",
                   "health_insurance_grade", "tax_dependents"]:
            val = record.get(f)
            if val is not None:
                try:
                    record[f] = int(float(val))
                except (ValueError, TypeError):
                    record[f] = 0

        val = record.get("pension_self_rate")
        if val is not None:
            try:
                record["pension_self_rate"] = float(val)
            except (ValueError, TypeError):
                record["pension_self_rate"] = 0.0

        # hire_date normalization
        hd = record.get("hire_date")
        if hd and isinstance(hd, datetime):
            record["hire_date"] = hd.strftime("%Y-%m-%d")
        elif hd:
            record["hire_date"] = str(hd).strip()

        records.append(record)

    logger.info(f"Parsed {len(records)} employee records from {file_path}")
    return records


# === 4. 契約解析（Gemini VLM）===

def parse_contract_image(image_path: str) -> Optional[dict]:
    """用 Gemini VLM 解析勞動契約圖片/PDF

    Args:
        image_path: 圖片或 PDF 本地路徑

    Returns:
        dict with: name, id_number, position, base_salary, hire_date, work_hours
        or None on failure
    """
    import base64
    import requests

    GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
    GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")

    if not GEMINI_API_KEY:
        logger.warning("No GEMINI_API_KEY for contract parsing")
        return None

    try:
        with open(image_path, "rb") as f:
            image_bytes = f.read()
        b64_image = base64.b64encode(image_bytes).decode("utf-8")

        ext = os.path.splitext(image_path)[1].lower()
        mime_map = {
            ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
            ".png": "image/png", ".webp": "image/webp",
            ".pdf": "application/pdf",
        }
        mime_type = mime_map.get(ext, "image/jpeg")

        contract_schema = {
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "員工姓名"},
                "id_number": {"type": "string", "description": "身分證字號"},
                "position": {"type": "string", "description": "職稱/職位"},
                "department": {"type": "string", "description": "部門"},
                "base_salary": {"type": "number", "description": "底薪/月薪（數字）"},
                "hire_date": {"type": "string", "description": "到職日（YYYY-MM-DD 格式，民國年+1911）"},
                "work_hours": {"type": "string", "description": "工作時間描述"},
                "contract_type": {"type": "string", "description": "契約類型（定期/不定期）"},
            },
            "required": ["name"],
        }

        url = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"
        payload = {
            "contents": [{
                "parts": [
                    {"inline_data": {"mime_type": mime_type, "data": b64_image}},
                    {"text": (
                        "請辨識這份勞動契約/聘僱合約的內容。\n"
                        "提取：員工姓名、身分證字號、職稱、部門、底薪（月薪金額）、到職日、工作時間、契約類型。\n"
                        "到職日請轉換為 YYYY-MM-DD 格式（民國年+1911=西元年）。\n"
                        "底薪請只填數字，不要包含逗號或「元」。\n"
                        "如果欄位看不清楚或找不到，請填空字串。"
                    )},
                ],
            }],
            "generationConfig": {
                "responseMimeType": "application/json",
                "responseJsonSchema": contract_schema,
                "temperature": 0.1,
                "maxOutputTokens": 2048,
            },
        }

        resp = requests.post(url, params={"key": GEMINI_API_KEY}, json=payload, timeout=30)

        if resp.status_code == 200:
            data = resp.json()
            text = data["candidates"][0]["content"]["parts"][0]["text"]
            result = json.loads(text)
            logger.info(f"Contract parsed: name={result.get('name')}")
            return result
        else:
            logger.error(f"Gemini contract parse error: {resp.status_code} {resp.text[:200]}")
            return None

    except Exception as e:
        logger.error(f"Contract parse error: {e}")
        return None


# === 5. 薪資匯入 ===

def import_salary_from_sheet(file_path: str, year_month: str) -> dict:
    """讀取已填寫完成的薪資表，匯入 DB

    Args:
        file_path: Excel 路徑
        year_month: 歸屬月份 "YYYY-MM"

    Returns:
        dict with: count, total_gross, total_net, errors
    """
    import state_manager as sm

    records = parse_salary_excel(file_path)
    if not records:
        return {"count": 0, "total_gross": 0, "total_net": 0, "errors": ["找不到有效的薪資資料"]}

    count = 0
    total_gross = 0
    total_net = 0
    errors = []

    for rec in records:
        name = str(rec.get("name", "")).strip()
        if not name:
            continue

        # Find employee in DB
        employees = sm.list_employees(status="active")
        emp = None
        for e in employees:
            if e["name"] == name:
                emp = e
                break

        if not emp:
            errors.append(f"找不到員工「{name}」，請先建立員工資料")
            continue

        try:
            sm.add_payroll(
                employee_id=emp["id"],
                year_month=year_month,
                base_salary=rec.get("base_salary", 0),
                meal_allowance=rec.get("meal_allowance", 0),
                overtime_hours=rec.get("overtime_hours", 0),
                overtime_pay=rec.get("overtime_pay", 0),
                bonus=rec.get("bonus", 0),
                gross_salary=rec.get("gross_salary", 0),
                labor_insurance=rec.get("labor_insurance", 0),
                health_insurance=rec.get("health_insurance", 0),
                pension_self=rec.get("pension_self", 0),
                income_tax=rec.get("income_tax", 0),
                other_deductions=rec.get("other_deductions", 0),
                net_salary=rec.get("net_salary", 0),
                status="draft",
            )
            count += 1
            total_gross += rec.get("gross_salary", 0)
            total_net += rec.get("net_salary", 0)
        except Exception as e:
            errors.append(f"匯入「{name}」失敗：{str(e)}")

    logger.info(f"Imported {count} payroll records for {year_month}")
    return {
        "count": count,
        "total_gross": total_gross,
        "total_net": total_net,
        "errors": errors,
    }


# === 6. GDrive 資料夾管理 ===

def create_employee_folder(employee_name: str) -> str:
    """在 GDrive 建立員工資料夾結構

    Args:
        employee_name: 員工姓名

    Returns:
        資料夾完整路徑
    """
    base = _get_base_path()
    folder = os.path.join(base, "員工資料", employee_name)
    subfolders = ["契約", "薪資", "其他"]
    for sub in subfolders:
        os.makedirs(os.path.join(folder, sub), exist_ok=True)
    logger.info(f"Employee folder created: {folder}")
    return folder


def archive_contract(image_path: str, employee_name: str) -> str:
    """將契約檔案歸檔到員工資料夾

    Args:
        image_path: 原始檔案路徑
        employee_name: 員工姓名

    Returns:
        歸檔後的路徑
    """
    import shutil

    folder = create_employee_folder(employee_name)
    ext = os.path.splitext(image_path)[1]
    date_str = datetime.now().strftime("%y%m%d")
    dest_filename = f"{date_str}_勞動契約{ext}"
    dest_path = os.path.join(folder, "契約", dest_filename)
    shutil.copy2(image_path, dest_path)
    logger.info(f"Contract archived: {dest_path}")
    return dest_path


# === 7. 菜單模板 ===

def generate_menu_template(year_month: str) -> str:
    """生成菜單排程 Excel 模板

    Args:
        year_month: 格式 "YYYY-MM"

    Returns:
        生成的檔案完整路徑
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "菜單排程"

    # Styles
    header_font = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    title_font = Font(name="微軟正黑體", size=14, bold=True)
    date_font = Font(name="微軟正黑體", size=10, bold=True)
    normal_font = Font(name="微軟正黑體", size=10)
    weekend_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = f"菜單排程表 — {year_month}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "日期", "午餐主菜", "午餐副菜1", "午餐副菜2", "午餐湯品",
        "晚餐主菜", "晚餐副菜1", "晚餐副菜2", "晚餐湯品",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Pre-fill dates for the month
    parts = year_month.split("-")
    year = int(parts[0])
    month = int(parts[1])
    _, days_in_month = calendar.monthrange(year, month)

    weekday_names = ["一", "二", "三", "四", "五", "六", "日"]

    for day in range(1, days_in_month + 1):
        row = day + 2
        dt = datetime(year, month, day)
        weekday = dt.weekday()
        day_name = weekday_names[weekday]
        date_str = f"{month}/{day}（{day_name}）"

        cell = ws.cell(row=row, column=1, value=date_str)
        cell.font = date_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

        # Style weekend rows
        is_weekend = weekday >= 5
        for col in range(1, 10):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            c.font = normal_font
            if is_weekend:
                c.fill = weekend_fill

    # Column widths
    ws.column_dimensions["A"].width = 14
    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.freeze_panes = "A3"

    # Save
    month_str = f"{month:02d}月"
    base = _get_base_path()
    save_dir = os.path.join(base, str(year), month_str, "菜單企劃")
    os.makedirs(save_dir, exist_ok=True)
    filepath = os.path.join(save_dir, f"菜單_{year_month}.xlsx")
    wb.save(filepath)
    logger.info(f"Menu template generated: {filepath}")
    return filepath


def parse_menu_excel(file_path: str) -> list[dict]:
    """解析已填寫的菜單 Excel

    Returns:
        list of dicts: {date, meal_type, slot, dish_name}
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    slot_map = {
        2: ("lunch", "主菜"),
        3: ("lunch", "副菜1"),
        4: ("lunch", "副菜2"),
        5: ("lunch", "湯品"),
        6: ("dinner", "主菜"),
        7: ("dinner", "副菜1"),
        8: ("dinner", "副菜2"),
        9: ("dinner", "湯品"),
    }

    records = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        date_cell = row[0].value
        if not date_cell:
            continue

        # Parse date like "3/1（一）" → extract month/day
        date_str = str(date_cell).strip()
        m = re.match(r"(\d+)/(\d+)", date_str)
        if not m:
            continue

        month = int(m.group(1))
        day = int(m.group(2))

        # Infer year from file context (current year)
        year = datetime.now().year
        try:
            schedule_date = f"{year}-{month:02d}-{day:02d}"
        except ValueError:
            continue

        for col_idx, (meal_type, slot) in slot_map.items():
            if col_idx < len(row) + 1:
                cell = row[col_idx - 1]  # 0-indexed
                dish_name = str(cell.value or "").strip()
                if dish_name:
                    records.append({
                        "date": schedule_date,
                        "meal_type": meal_type,
                        "slot": slot,
                        "dish_name": dish_name,
                    })

    logger.info(f"Parsed {len(records)} menu items from {file_path}")
    return records
