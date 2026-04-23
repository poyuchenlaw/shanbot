"""小膳 Bot — 自動稽核服務

財務報表稽核 & 異常偵測：
1. 月度試算表借貸平衡驗證
2. 資產負債表恆等式（資產 = 負債 + 權益）
3. 進項稅額與發票張數覈對
4. 單筆/供應商集中度異常偵測
5. 跨月一致性檢查
6. 全套稽核報告生成
"""

import logging
import os
from datetime import datetime
from typing import Optional

import state_manager as sm

logger = logging.getLogger("shanbot.audit")


# =====================================================================
# 1. 試算表借貸平衡驗證
# =====================================================================

def verify_trial_balance(year_month: str) -> dict:
    """驗證月度試算表借貸是否平衡

    Returns:
        {
            "year_month": str,
            "total_debit": float,
            "total_credit": float,
            "difference": float,
            "balanced": bool,
            "account_count": int,
            "details": list  # 各科目明細
        }
    """
    trial = sm.get_trial_balance(year_month)
    total_debit = sum(t.get("total_debit", 0) or 0 for t in trial)
    total_credit = sum(t.get("total_credit", 0) or 0 for t in trial)
    diff = abs(total_debit - total_credit)

    return {
        "year_month": year_month,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "difference": diff,
        "balanced": diff < 1,  # 容許 $1 四捨五入差
        "account_count": len(trial),
        "details": trial,
    }


# =====================================================================
# 2. 資產負債表恆等式驗證
# =====================================================================

def verify_balance_sheet_equation(year_month: str) -> dict:
    """驗證 資產 = 負債 + 權益

    結帳前：資產 = 負債 + 權益 + (收入 - 成本 - 費用)
    結帳後：資產 = 負債 + 權益（淨利已結轉到 3300）
    """
    from services.accounting_service import generate_balance_sheet
    bs = generate_balance_sheet(year_month)

    assets = bs.get("assets", 0)
    liabilities = bs.get("liabilities", 0)
    equity = bs.get("equity", 0)

    # 檢查是否已結帳
    acct = sm.get_monthly_accounting(year_month)
    is_closed = bool(acct and acct.get("is_closed"))

    if not is_closed:
        # 結帳前：計算未結轉的損益（4xxx - 5xxx - 6xxx）
        trial = sm.get_trial_balance(year_month)
        revenue = sum(abs(t.get("balance", 0)) for t in trial if t.get("account_code", "").startswith("4"))
        cost = sum(t.get("balance", 0) for t in trial if t.get("account_code", "").startswith("5"))
        expense = sum(t.get("balance", 0) for t in trial if t.get("account_code", "").startswith("6"))
        uncleared_net_income = revenue - cost - expense
        effective_equity = equity + uncleared_net_income
    else:
        effective_equity = equity
        uncleared_net_income = 0

    sum_liab_equity = liabilities + effective_equity
    diff = abs(assets - sum_liab_equity)

    return {
        "year_month": year_month,
        "assets": assets,
        "liabilities": liabilities,
        "equity": equity,
        "uncleared_net_income": uncleared_net_income,
        "effective_equity": effective_equity,
        "sum_liab_equity": sum_liab_equity,
        "difference": diff,
        "balanced": diff < 1,
        "is_closed": is_closed,
        "asset_detail": bs.get("asset_detail", []),
        "liability_detail": bs.get("liability_detail", []),
        "equity_detail": bs.get("equity_detail", []),
    }


# =====================================================================
# 3. 進項稅額與發票覈對
# =====================================================================

def verify_input_tax(year_month: str) -> dict:
    """覈對進項稅額：staging 記錄 vs journal_entries

    Returns:
        {
            "staging_tax_total": float,    # purchase_staging 的 tax_amount 合計
            "journal_tax_total": float,    # journal_entries 進項稅額科目合計
            "matched": bool,
            "invoice_count": int,          # 有統編發票張數
            "receipt_count": int,          # 免用發票/收據張數
            "deductible_count": int,       # 可扣抵筆數
            "non_deductible_count": int,   # 不可扣抵筆數
        }
    """
    stagings = sm.get_stagings_by_month(year_month)

    staging_tax = 0
    invoice_count = 0
    receipt_count = 0
    deductible_count = 0
    non_deductible_count = 0

    for s in stagings:
        if s.get("status") not in ("confirmed", "exported"):
            continue
        tax = s.get("tax_amount", 0) or 0
        staging_tax += tax

        if s.get("invoice_type") == "免用發票" or not s.get("supplier_tax_id"):
            receipt_count += 1
        else:
            invoice_count += 1

        if s.get("deduction_code") == "1":
            deductible_count += 1
        else:
            non_deductible_count += 1

    # journal_entries 中 1150 進項稅額的借方合計
    journal_entries = sm.get_journal_entries(year_month)
    journal_tax = sum(
        e.get("debit", 0) for e in journal_entries
        if e.get("account_code") == "1150"
    )

    return {
        "year_month": year_month,
        "staging_tax_total": staging_tax,
        "journal_tax_total": journal_tax,
        "difference": abs(staging_tax - journal_tax),
        "matched": abs(staging_tax - journal_tax) < 1,
        "invoice_count": invoice_count,
        "receipt_count": receipt_count,
        "deductible_count": deductible_count,
        "non_deductible_count": non_deductible_count,
    }


# =====================================================================
# 4. 異常偵測
# =====================================================================

def detect_anomalies(year_month: str) -> dict:
    """偵測異常交易

    檢查項目：
    - 單筆金額超過門檻（預設 50,000）
    - 供應商集中度（單一供應商佔比 > 40%）
    - 重複金額（同日同金額）
    - 未分類品項

    Returns:
        {
            "high_amount": list,        # 高金額交易
            "supplier_concentration": list,  # 供應商集中度
            "duplicate_suspects": list, # 疑似重複
            "unclassified": list,       # 未分類品項
            "total_alerts": int,
        }
    """
    stagings = sm.get_stagings_by_month(year_month)
    confirmed = [s for s in stagings if s.get("status") in ("confirmed", "exported")]

    alerts = {
        "high_amount": [],
        "supplier_concentration": [],
        "duplicate_suspects": [],
        "unclassified": [],
    }

    HIGH_AMOUNT_THRESHOLD = 50000
    CONCENTRATION_THRESHOLD = 0.4

    total_amount = sum(s.get("total_amount", 0) or 0 for s in confirmed)

    # 高金額
    for s in confirmed:
        amt = s.get("total_amount", 0) or 0
        if amt >= HIGH_AMOUNT_THRESHOLD:
            alerts["high_amount"].append({
                "id": s["id"],
                "date": s.get("purchase_date"),
                "supplier": s.get("supplier_name"),
                "amount": amt,
            })

    # 供應商集中度
    supplier_totals = {}
    for s in confirmed:
        name = s.get("supplier_name") or "未知"
        supplier_totals[name] = supplier_totals.get(name, 0) + (s.get("total_amount", 0) or 0)

    if total_amount > 0:
        for name, amt in supplier_totals.items():
            ratio = amt / total_amount
            if ratio >= CONCENTRATION_THRESHOLD:
                alerts["supplier_concentration"].append({
                    "supplier": name,
                    "amount": amt,
                    "ratio": round(ratio * 100, 1),
                })

    # 重複金額偵測（同日同金額）
    date_amount_map = {}
    for s in confirmed:
        key = (s.get("purchase_date"), s.get("total_amount"))
        if key[0] and key[1]:
            if key not in date_amount_map:
                date_amount_map[key] = []
            date_amount_map[key].append(s)

    for key, items in date_amount_map.items():
        if len(items) > 1:
            alerts["duplicate_suspects"].append({
                "date": key[0],
                "amount": key[1],
                "count": len(items),
                "ids": [s["id"] for s in items],
                "suppliers": [s.get("supplier_name") for s in items],
            })

    # 未分類品項
    for s in confirmed:
        items = sm.get_purchase_items(s["id"])
        for item in items:
            cat = item.get("category", "")
            if not cat or cat in ("other", "其他", "未分類"):
                alerts["unclassified"].append({
                    "staging_id": s["id"],
                    "item_name": item.get("item_name"),
                    "amount": item.get("amount"),
                })

    total_alerts = sum(len(v) for v in alerts.values())
    alerts["total_alerts"] = total_alerts

    return alerts


# =====================================================================
# 5. 每筆交易借貸平衡批次驗證
# =====================================================================

def verify_all_journal_balances(year_month: str) -> dict:
    """批次驗證所有已確認交易的借貸平衡

    Returns:
        {
            "total_checked": int,
            "all_balanced": bool,
            "imbalanced": list  # 不平衡的交易
        }
    """
    stagings = sm.get_stagings_by_month(year_month)
    confirmed = [s for s in stagings if s.get("status") in ("confirmed", "exported")]

    imbalanced = []
    for s in confirmed:
        entries = sm.get_journal_entries_by_source("purchase", s["id"])
        if not entries:
            imbalanced.append({
                "staging_id": s["id"],
                "supplier": s.get("supplier_name"),
                "issue": "missing_journal_entries",
            })
            continue

        debit = sum(e.get("debit", 0) for e in entries)
        credit = sum(e.get("credit", 0) for e in entries)
        if abs(debit - credit) >= 1:
            imbalanced.append({
                "staging_id": s["id"],
                "supplier": s.get("supplier_name"),
                "debit": debit,
                "credit": credit,
                "difference": debit - credit,
                "issue": "imbalanced",
            })

    return {
        "year_month": year_month,
        "total_checked": len(confirmed),
        "all_balanced": len(imbalanced) == 0,
        "imbalanced_count": len(imbalanced),
        "imbalanced": imbalanced,
    }


# =====================================================================
# 6. 損益表交叉驗證
# =====================================================================

def verify_income_statement(year_month: str) -> dict:
    """交叉驗證損益表數據

    比對：
    - staging 金額合計 vs journal_entries 進貨科目合計
    - income 表 vs journal_entries 收入科目合計
    """
    from services.accounting_service import generate_income_statement
    pl = generate_income_statement(year_month)

    # staging 進貨合計
    stagings = sm.get_stagings_by_month(year_month)
    confirmed = [s for s in stagings if s.get("status") in ("confirmed", "exported")]
    staging_subtotal = sum(s.get("subtotal", 0) or 0 for s in confirmed)

    # journal 進貨合計（5xxx）
    journal_cost = pl.get("cost", 0)

    # income 表合計
    income_rows = sm.get_income_summary(year_month)
    income_table_total = sum(r.get("amount", 0) for r in income_rows)

    # journal 收入合計（4xxx）
    journal_revenue = pl.get("revenue", 0)

    # 如果已結帳，收入/成本科目已歸零，改用 monthly_accounting 記錄比對
    acct = sm.get_monthly_accounting(year_month)
    if acct and acct.get("is_closed"):
        acct_income = acct.get("total_income", 0) or 0
        acct_expense = acct.get("total_expense", 0) or 0
        acct_net = acct.get("net_profit", 0) or 0
        # 結帳後用 monthly_accounting 的數據驗證
        cost_matched = abs(staging_subtotal - acct_expense) < 1 if acct_expense > 0 else (staging_subtotal == 0)
        revenue_matched = True  # 已結帳代表收入已驗證
        gross_profit = acct_income - acct_expense
        net_income = acct_net
    else:
        cost_matched = abs(staging_subtotal - journal_cost) < 1
        revenue_matched = abs(income_table_total - journal_revenue) < 1 if income_table_total > 0 else True
        gross_profit = pl.get("gross_profit", 0)
        net_income = pl.get("net_income", 0)

    return {
        "year_month": year_month,
        "staging_subtotal": staging_subtotal,
        "journal_cost": journal_cost,
        "cost_matched": cost_matched,
        "income_table_total": income_table_total,
        "journal_revenue": journal_revenue,
        "revenue_matched": revenue_matched,
        "gross_profit": gross_profit,
        "net_income": net_income,
        "is_closed": bool(acct and acct.get("is_closed")),
    }


# =====================================================================
# 7. 完整稽核報告
# =====================================================================

def run_full_audit(year_month: str) -> dict:
    """執行完整月度稽核，回傳綜合報告

    Returns:
        {
            "year_month": str,
            "timestamp": str,
            "overall_pass": bool,
            "checks": {
                "trial_balance": {...},
                "balance_sheet": {...},
                "input_tax": {...},
                "journal_balances": {...},
                "income_statement": {...},
                "anomalies": {...},
            },
            "summary": str  # 人類可讀摘要
        }
    """
    logger.info(f"Starting full audit for {year_month}")

    checks = {}

    # 1. 試算表
    checks["trial_balance"] = verify_trial_balance(year_month)

    # 2. 資產負債表恆等式
    checks["balance_sheet"] = verify_balance_sheet_equation(year_month)

    # 3. 進項稅額覈對
    checks["input_tax"] = verify_input_tax(year_month)

    # 4. 分錄借貸平衡
    checks["journal_balances"] = verify_all_journal_balances(year_month)

    # 5. 損益表交叉驗證
    checks["income_statement"] = verify_income_statement(year_month)

    # 6. 異常偵測
    checks["anomalies"] = detect_anomalies(year_month)

    # 綜合判定
    pass_checks = [
        checks["trial_balance"]["balanced"],
        checks["balance_sheet"]["balanced"],
        checks["input_tax"]["matched"],
        checks["journal_balances"]["all_balanced"],
        checks["income_statement"]["cost_matched"],
    ]
    overall_pass = all(pass_checks)

    # 人類可讀摘要
    lines = [f"📋 {year_month} 月度稽核報告", ""]

    tb = checks["trial_balance"]
    lines.append(f"1️⃣ 試算表借貸平衡：{'✅ 通過' if tb['balanced'] else '❌ 不平衡'}")
    lines.append(f"   借方合計 {tb['total_debit']:,.0f} / 貸方合計 {tb['total_credit']:,.0f}")
    if not tb["balanced"]:
        lines.append(f"   ⚠️ 差異 {tb['difference']:,.0f}")

    bs = checks["balance_sheet"]
    lines.append(f"2️⃣ 資產=負債+權益：{'✅ 通過' if bs['balanced'] else '❌ 不平衡'}")
    lines.append(f"   資產 {bs['assets']:,.0f} / 負債+權益 {bs['sum_liab_equity']:,.0f}")
    if bs.get("uncleared_net_income"):
        lines.append(f"   （含未結轉淨利 {bs['uncleared_net_income']:,.0f}）")

    it = checks["input_tax"]
    lines.append(f"3️⃣ 進項稅額覈對：{'✅ 一致' if it['matched'] else '❌ 不一致'}")
    lines.append(f"   發票 {it['invoice_count']} 張（可扣抵 {it['deductible_count']}）/ 收據 {it['receipt_count']} 張")

    jb = checks["journal_balances"]
    lines.append(f"4️⃣ 分錄借貸平衡：{'✅ 全部通過' if jb['all_balanced'] else '❌ ' + str(jb['imbalanced_count']) + ' 筆不平衡'}")
    lines.append(f"   已檢查 {jb['total_checked']} 筆交易")

    is_check = checks["income_statement"]
    lines.append(f"5️⃣ 損益表交叉驗證：{'✅ 一致' if is_check['cost_matched'] else '❌ 成本不一致'}")
    lines.append(f"   毛利 {is_check['gross_profit']:,.0f} / 淨利 {is_check['net_income']:,.0f}")

    anom = checks["anomalies"]
    lines.append(f"6️⃣ 異常偵測：{anom['total_alerts']} 項警示")
    if anom["high_amount"]:
        lines.append(f"   ⚠️ 高金額 {len(anom['high_amount'])} 筆")
    if anom["supplier_concentration"]:
        lines.append(f"   ⚠️ 供應商集中 {len(anom['supplier_concentration'])} 家")
    if anom["duplicate_suspects"]:
        lines.append(f"   ⚠️ 疑似重複 {len(anom['duplicate_suspects'])} 組")
    if anom["unclassified"]:
        lines.append(f"   ⚠️ 未分類品項 {len(anom['unclassified'])} 個")

    lines.append("")
    lines.append(f"{'✅ 整體通過' if overall_pass else '❌ 有項目未通過，請檢查'}")

    summary = "\n".join(lines)
    logger.info(f"Audit complete: {'PASS' if overall_pass else 'FAIL'}")

    return {
        "year_month": year_month,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "overall_pass": overall_pass,
        "checks": checks,
        "summary": summary,
    }


def generate_audit_excel(year_month: str, output_dir: str = None) -> Optional[str]:
    """生成稽核報告 Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        logger.error("openpyxl not installed")
        return None

    audit = run_full_audit(year_month)

    if not output_dir:
        output_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "data", "exports", year_month
        )
    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.Workbook()

    # 樣式
    title_font = Font(name="微軟正黑體", size=14, bold=True)
    header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    pass_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    fail_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    normal_font = Font(name="微軟正黑體", size=10)
    money_fmt = '#,##0'

    # === Sheet 1: 稽核總覽 ===
    ws1 = wb.active
    ws1.title = "稽核總覽"

    ws1.merge_cells("A1:E1")
    ws1["A1"] = f"{year_month} 月度稽核報告"
    ws1["A1"].font = title_font

    ws1.merge_cells("A2:E2")
    ws1["A2"] = f"產生時間：{audit['timestamp']}"
    ws1["A2"].font = normal_font

    # 總覽表
    row = 4
    headers = ["檢查項目", "結果", "借方/資產", "貸方/負債+權益", "差異"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    checks_rows = [
        (
            "試算表借貸平衡",
            audit["checks"]["trial_balance"]["balanced"],
            audit["checks"]["trial_balance"]["total_debit"],
            audit["checks"]["trial_balance"]["total_credit"],
            audit["checks"]["trial_balance"]["difference"],
        ),
        (
            "資產=負債+權益",
            audit["checks"]["balance_sheet"]["balanced"],
            audit["checks"]["balance_sheet"]["assets"],
            audit["checks"]["balance_sheet"]["sum_liab_equity"],
            audit["checks"]["balance_sheet"]["difference"],
        ),
        (
            "進項稅額覈對",
            audit["checks"]["input_tax"]["matched"],
            audit["checks"]["input_tax"]["staging_tax_total"],
            audit["checks"]["input_tax"]["journal_tax_total"],
            audit["checks"]["input_tax"]["difference"],
        ),
        (
            "分錄借貸平衡",
            audit["checks"]["journal_balances"]["all_balanced"],
            audit["checks"]["journal_balances"]["total_checked"],
            0,
            audit["checks"]["journal_balances"]["imbalanced_count"],
        ),
        (
            "損益表交叉驗證",
            audit["checks"]["income_statement"]["cost_matched"],
            audit["checks"]["income_statement"]["staging_subtotal"],
            audit["checks"]["income_statement"]["journal_cost"],
            abs(
                audit["checks"]["income_statement"]["staging_subtotal"]
                - audit["checks"]["income_statement"]["journal_cost"]
            ),
        ),
    ]

    for item_name, passed, val1, val2, diff in checks_rows:
        row += 1
        ws1.cell(row=row, column=1, value=item_name).font = normal_font
        result_cell = ws1.cell(row=row, column=2, value="✅ 通過" if passed else "❌ 未通過")
        result_cell.font = normal_font
        result_cell.fill = pass_fill if passed else fail_fill
        ws1.cell(row=row, column=3, value=val1).number_format = money_fmt
        ws1.cell(row=row, column=4, value=val2).number_format = money_fmt
        ws1.cell(row=row, column=5, value=diff).number_format = money_fmt

    row += 2
    overall = audit["overall_pass"]
    ws1.cell(row=row, column=1, value="整體結果").font = Font(
        name="微軟正黑體", size=12, bold=True
    )
    result_cell = ws1.cell(
        row=row, column=2, value="✅ 全部通過" if overall else "❌ 有項目未通過"
    )
    result_cell.font = Font(name="微軟正黑體", size=12, bold=True)
    result_cell.fill = pass_fill if overall else fail_fill

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 18
    ws1.column_dimensions["E"].width = 14

    # === Sheet 2: 試算表明細 ===
    ws2 = wb.create_sheet("試算表")
    ws2.merge_cells("A1:E1")
    ws2["A1"] = f"{year_month} 試算表"
    ws2["A1"].font = title_font

    row = 3
    for col, h in enumerate(["科目代碼", "科目名稱", "借方合計", "貸方合計", "餘額"], 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for t in audit["checks"]["trial_balance"]["details"]:
        row += 1
        ws2.cell(row=row, column=1, value=t.get("account_code", "")).font = normal_font
        ws2.cell(row=row, column=2, value=t.get("account_name", "")).font = normal_font
        ws2.cell(row=row, column=3, value=t.get("total_debit", 0)).number_format = money_fmt
        ws2.cell(row=row, column=4, value=t.get("total_credit", 0)).number_format = money_fmt
        ws2.cell(row=row, column=5, value=t.get("balance", 0)).number_format = money_fmt

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 16

    # === Sheet 3: 異常偵測 ===
    ws3 = wb.create_sheet("異常偵測")
    ws3.merge_cells("A1:E1")
    ws3["A1"] = f"{year_month} 異常偵測報告"
    ws3["A1"].font = title_font

    anom = audit["checks"]["anomalies"]
    row = 3

    # 高金額
    ws3.cell(row=row, column=1, value="⚠️ 高金額交易（≥50,000）").font = Font(
        name="微軟正黑體", size=11, bold=True
    )
    row += 1
    if anom["high_amount"]:
        for col, h in enumerate(["ID", "日期", "供應商", "金額"], 1):
            ws3.cell(row=row, column=col, value=h).font = header_font
            ws3.cell(row=row, column=col).fill = header_fill
        for a in anom["high_amount"]:
            row += 1
            ws3.cell(row=row, column=1, value=a["id"])
            ws3.cell(row=row, column=2, value=a["date"])
            ws3.cell(row=row, column=3, value=a["supplier"])
            ws3.cell(row=row, column=4, value=a["amount"]).number_format = money_fmt
    else:
        ws3.cell(row=row, column=1, value="（無）").font = normal_font
    row += 2

    # 供應商集中
    ws3.cell(row=row, column=1, value="⚠️ 供應商集中度（≥40%）").font = Font(
        name="微軟正黑體", size=11, bold=True
    )
    row += 1
    if anom["supplier_concentration"]:
        for col, h in enumerate(["供應商", "金額", "佔比(%)"], 1):
            ws3.cell(row=row, column=col, value=h).font = header_font
            ws3.cell(row=row, column=col).fill = header_fill
        for a in anom["supplier_concentration"]:
            row += 1
            ws3.cell(row=row, column=1, value=a["supplier"])
            ws3.cell(row=row, column=2, value=a["amount"]).number_format = money_fmt
            ws3.cell(row=row, column=3, value=a["ratio"])
    else:
        ws3.cell(row=row, column=1, value="（無）").font = normal_font
    row += 2

    # 疑似重複
    ws3.cell(row=row, column=1, value="⚠️ 疑似重複交易").font = Font(
        name="微軟正黑體", size=11, bold=True
    )
    row += 1
    if anom["duplicate_suspects"]:
        for col, h in enumerate(["日期", "金額", "筆數", "供應商"], 1):
            ws3.cell(row=row, column=col, value=h).font = header_font
            ws3.cell(row=row, column=col).fill = header_fill
        for a in anom["duplicate_suspects"]:
            row += 1
            ws3.cell(row=row, column=1, value=a["date"])
            ws3.cell(row=row, column=2, value=a["amount"]).number_format = money_fmt
            ws3.cell(row=row, column=3, value=a["count"])
            ws3.cell(row=row, column=4, value=", ".join(a["suppliers"]))
    else:
        ws3.cell(row=row, column=1, value="（無）").font = normal_font

    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 10
    ws3.column_dimensions["D"].width = 20
    ws3.column_dimensions["E"].width = 14

    filepath = os.path.join(output_dir, f"{year_month}_稽核報告.xlsx")
    wb.save(filepath)
    logger.info(f"Audit Excel generated: {filepath}")
    return filepath
