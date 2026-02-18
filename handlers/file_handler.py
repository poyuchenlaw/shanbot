"""檔案上傳處理 — Excel/PDF 自動分類歸檔（v2.2 內容檢視版）"""

import logging
import os
import re
from datetime import datetime

import state_manager as sm

logger = logging.getLogger("shanbot.file")

FILES_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "files")

# 檔名關鍵字 → doc_category 映射（順序有意義：前面優先）
_KEYWORD_CATEGORY = [
    (["薪資", "薪水", "salary", "工資", "勞健保"], "payroll"),
    (["租約", "租賃", "lease", "rent", "房租"], "general"),
    (["借款", "貸款", "loan", "利息"], "financing"),
    (["折舊", "資產", "asset", "設備"], "fixed_asset"),
    (["保險", "insurance"], "general"),
    (["投資", "investment"], "investment"),
    (["合約", "contract", "契約"], "revenue"),
    (["採購", "purchase", "進貨"], "expenditure"),
    (["菜單", "成本", "menu"], "production"),
]

# 八大循環中文名稱
CATEGORY_LABELS = {
    "revenue": "收入循環",
    "expenditure": "支出循環",
    "payroll": "人力資源循環",
    "production": "生產循環",
    "financing": "融資循環",
    "investment": "投資循環",
    "fixed_asset": "固定資產循環",
    "general": "一般循環",
}

# 分類 → 建議歸檔短名
_CATEGORY_SHORT = {
    "revenue": "收入",
    "expenditure": "採購",
    "payroll": "薪資表",
    "production": "生產成本",
    "financing": "融資",
    "investment": "投資",
    "fixed_asset": "固定資產",
    "general": "一般文件",
}


def classify_by_filename(filename: str) -> str:
    """根據檔名關鍵字自動分類，回傳 doc_category"""
    lower = filename.lower()
    for keywords, category in _KEYWORD_CATEGORY:
        for kw in keywords:
            if kw in lower:
                return category
    return "general"


def _classify_by_keywords(text: str) -> str | None:
    """給定文本字串，用關鍵字映射嘗試分類"""
    lower = text.lower()
    for keywords, category in _KEYWORD_CATEGORY:
        for kw in keywords:
            if kw in lower:
                return category
    return None


def detect_file_type(filename: str) -> str:
    """根據副檔名判斷檔案類型"""
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xls", ".csv")):
        return "excel"
    elif lower.endswith(".pdf"):
        return "pdf"
    elif lower.endswith((".jpg", ".jpeg", ".png", ".gif")):
        return "image"
    return "other"


def inspect_excel_content(local_path: str) -> dict:
    """讀取 Excel 檔案內容，回傳 metadata + 內容分類建議

    回傳:
        {
            "sheet_names": [...],
            "headers": {"Sheet1": [...], ...},
            "content_keywords": [...],
            "suggested_category": "payroll" | None,
            "content_summary": "...",
        }
    """
    result = {
        "sheet_names": [],
        "headers": {},
        "content_keywords": [],
        "suggested_category": None,
        "content_summary": "",
    }
    try:
        import openpyxl
        wb = openpyxl.load_workbook(local_path, read_only=True, data_only=True)
        result["sheet_names"] = wb.sheetnames

        all_text_parts = []
        for sheet_name in wb.sheetnames[:3]:  # 最多讀 3 個 sheet
            ws = wb[sheet_name]
            headers = []
            row_count = 0
            for row in ws.iter_rows(max_row=10, values_only=True):
                row_count += 1
                cells = [str(c) for c in row if c is not None]
                if row_count == 1:
                    headers = cells
                all_text_parts.extend(cells)

            result["headers"][sheet_name] = headers

        wb.close()

        # 將所有文字合併做關鍵字分析
        combined = " ".join(all_text_parts[:200])  # 限制長度
        result["content_keywords"] = _extract_keywords(combined)
        result["suggested_category"] = _classify_by_keywords(combined)

        # 摘要（前 100 字）
        summary_parts = [p for p in all_text_parts[:20] if p.strip() and p != "None"]
        result["content_summary"] = " | ".join(summary_parts)[:200]

    except Exception as e:
        logger.warning(f"Excel content inspection failed: {e}")

    return result


def inspect_pdf_content(local_path: str) -> dict:
    """讀取 PDF 文字內容，回傳 metadata + 內容分類建議

    回傳:
        {
            "page_count": N,
            "content_keywords": [...],
            "suggested_category": "payroll" | None,
            "content_summary": "...",
        }
    """
    result = {
        "page_count": 0,
        "content_keywords": [],
        "suggested_category": None,
        "content_summary": "",
    }
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(local_path)
        result["page_count"] = len(doc)

        all_text = []
        for page_idx in range(min(3, len(doc))):  # 最多讀 3 頁
            page = doc[page_idx]
            text = page.get_text("text")
            if text:
                all_text.append(text)
        doc.close()

        combined = " ".join(all_text)[:2000]  # 限制長度
        result["content_keywords"] = _extract_keywords(combined)
        result["suggested_category"] = _classify_by_keywords(combined)

        # 摘要
        clean = re.sub(r"\s+", " ", combined).strip()
        result["content_summary"] = clean[:200]

    except Exception as e:
        logger.warning(f"PDF content inspection failed: {e}")

    return result


def _extract_keywords(text: str) -> list[str]:
    """從文字中提取會計相關關鍵字"""
    found = []
    all_keywords = {
        "薪資": "payroll", "薪水": "payroll", "勞保": "payroll",
        "健保": "payroll", "工資": "payroll", "扣繳": "payroll",
        "租約": "general", "租賃": "general", "房租": "general",
        "借款": "financing", "貸款": "financing", "利息": "financing",
        "折舊": "fixed_asset", "資產": "fixed_asset", "設備": "fixed_asset",
        "保險": "general", "投資": "investment",
        "合約": "revenue", "契約": "revenue", "收入": "revenue",
        "採購": "expenditure", "進貨": "expenditure", "發票": "expenditure",
        "菜單": "production", "食材": "production", "成本": "production",
        "損益": "revenue", "資產負債": "fixed_asset",
        "現金流量": "financing", "權益": "investment",
    }
    lower = text.lower()
    for kw in all_keywords:
        if kw in lower and kw not in found:
            found.append(kw)
    return found[:10]


def build_smart_filename(
    original_filename: str,
    doc_category: str,
    content_summary: str,
    year_month: str,
) -> str:
    """根據分類和內容建立有意義的歸檔檔名

    格式：YYYYMM_分類短名_原始檔名
    例如：202602_薪資表_1月份薪資.xlsx
    """
    ext = os.path.splitext(original_filename)[1]  # .xlsx, .pdf
    category_label = _CATEGORY_SHORT.get(doc_category, "文件")
    ym_short = year_month.replace("-", "")  # 202602

    # 清理原始檔名（去掉副檔名、去掉路徑字元）
    base = os.path.splitext(original_filename)[0]
    base = base.replace("/", "_").replace("\\", "_")

    # 如果原始檔名已經包含分類關鍵字，不重複加
    category_already = False
    for kw in (_CATEGORY_SHORT.get(doc_category, ""),):
        if kw and kw in base:
            category_already = True
            break

    if category_already:
        smart_name = f"{ym_short}_{base}{ext}"
    else:
        smart_name = f"{ym_short}_{category_label}_{base}{ext}"

    return smart_name


async def handle_file_received(
    line_service, message_id: str, filename: str,
    group_id: str, user_id: str, reply_token: str,
) -> str | None:
    """完整檔案處理流程：下載 → 內容檢視 → 分類 → 重命名 → 儲存 → GDrive → Flex"""

    # 1. 下載檔案
    file_bytes = line_service.get_content(message_id)
    if not file_bytes:
        return "❌ 檔案下載失敗，請重新上傳"

    # 2. 判斷類型
    file_type = detect_file_type(filename)
    if file_type not in ("excel", "pdf"):
        return (
            f"收到檔案：{filename}\n"
            "目前支援 Excel (.xlsx/.xls) 和 PDF 格式。\n"
            "請轉換格式後重新上傳。"
        )

    # 3. 先存到暫存位置（後面可能重命名）
    os.makedirs(FILES_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = filename.replace("/", "_").replace("\\", "_")
    temp_filename = f"{timestamp}_{safe_name}"
    local_path = os.path.join(FILES_DIR, temp_filename)

    with open(local_path, "wb") as f:
        f.write(file_bytes)
    logger.info(f"File saved: {local_path} ({len(file_bytes)} bytes)")

    # 4. 內容檢視 + 智能分類
    year_month = datetime.now().strftime("%Y-%m")
    content_info = {}
    content_category = None

    if file_type == "excel":
        content_info = inspect_excel_content(local_path)
        content_category = content_info.get("suggested_category")
    elif file_type == "pdf":
        content_info = inspect_pdf_content(local_path)
        content_category = content_info.get("suggested_category")

    # 分類優先級：內容檢視 > 檔名關鍵字 > 預設
    filename_category = classify_by_filename(filename)
    if content_category:
        doc_category = content_category
        classification_method = "content"
    elif filename_category != "general":
        doc_category = filename_category
        classification_method = "filename"
    else:
        doc_category = "general"
        classification_method = "default"

    logger.info(
        f"Classification: {doc_category} (method={classification_method}, "
        f"content={content_category}, filename={filename_category})"
    )

    # 5. 智能重命名
    content_summary = content_info.get("content_summary", "")
    smart_name = build_smart_filename(filename, doc_category, content_summary, year_month)

    # 重命名本地檔案
    smart_path = os.path.join(FILES_DIR, smart_name)
    if smart_path != local_path:
        try:
            os.rename(local_path, smart_path)
            local_path = smart_path
            logger.info(f"File renamed: {temp_filename} → {smart_name}")
        except Exception as e:
            logger.warning(f"Rename failed, keeping original: {e}")
            smart_name = temp_filename

    # 6. 收集完整 metadata
    metadata = {
        "original_filename": filename,
        "smart_filename": smart_name,
        "original_size": len(file_bytes),
        "classification_method": classification_method,
    }
    if content_info.get("sheet_names"):
        metadata["sheet_names"] = content_info["sheet_names"]
    if content_info.get("headers"):
        metadata["headers"] = content_info["headers"]
    if content_info.get("page_count"):
        metadata["page_count"] = content_info["page_count"]
    if content_info.get("content_keywords"):
        metadata["content_keywords"] = content_info["content_keywords"]
    if content_summary:
        metadata["content_summary"] = content_summary[:200]

    # 7. 記錄到 DB
    doc_id = sm.add_financial_document(
        chat_id=group_id,
        user_id=user_id,
        filename=smart_name,
        file_type=file_type,
        doc_category=doc_category,
        local_path=local_path,
        year_month=year_month,
        metadata=metadata,
    )

    # 8. 上傳到 GDrive（用智能檔名）
    gdrive_path = None
    try:
        from services.gdrive_service import upload_financial_doc
        gdrive_path = await upload_financial_doc(
            local_path, year_month, doc_category, smart_name
        )
        if gdrive_path:
            sm.update_financial_document(doc_id, gdrive_path=gdrive_path)
            logger.info(f"Financial doc #{doc_id} uploaded to GDrive: {gdrive_path}")
    except Exception as e:
        logger.warning(f"GDrive upload skipped: {e}")

    # 9. 回覆 Flex
    try:
        from services.flex_builder import build_file_upload_result_flex
        doc_info = {
            "id": doc_id,
            "filename": smart_name,
            "original_filename": filename,
            "file_type": file_type,
            "doc_category": doc_category,
            "category_label": CATEGORY_LABELS.get(doc_category, "一般循環"),
            "classification_method": classification_method,
            "content_keywords": content_info.get("content_keywords", []),
            "gdrive_path": gdrive_path or "",
            "year_month": year_month,
        }
        flex = build_file_upload_result_flex(doc_info)
        line_service.reply_flex(
            reply_token,
            f"📁 檔案已歸檔：{smart_name}",
            flex,
        )
        return None  # 已用 Flex 回覆
    except Exception as e:
        logger.warning(f"Flex reply failed, fallback to text: {e}")

    # Fallback 文字回覆
    category_label = CATEGORY_LABELS.get(doc_category, "一般循環")
    method_label = {"content": "內容分析", "filename": "檔名分析", "default": "預設"}
    gdrive_note = f"\n☁️ 已同步：{gdrive_path}" if gdrive_path else ""
    keywords_note = ""
    if content_info.get("content_keywords"):
        keywords_note = f"\n🔑 偵測到：{', '.join(content_info['content_keywords'][:5])}"
    return (
        f"📁 檔案已歸檔（#{doc_id}）\n"
        f"原檔名：{filename}\n"
        f"歸檔名：{smart_name}\n"
        f"分類：{category_label}（{method_label.get(classification_method, '')}）\n"
        f"月份：{year_month}{keywords_note}{gdrive_note}\n\n"
        f"如需修改分類，請輸入：\n"
        f"「修改分類 #{doc_id} 分類名」"
    )
