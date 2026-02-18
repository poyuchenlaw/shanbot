"""report_service.py generate_annual_report() 單元測試 — 年度報表生成驗證"""

import os
import sys
import tempfile
import unittest
from unittest.mock import patch, MagicMock

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

# Pre-import so state_manager is in sys.modules for local `import state_manager as sm`
import state_manager  # noqa: F401


def _make_staging(sid, date, supplier, subtotal, tax, total,
                  invoice_prefix="", invoice_number="", status="confirmed"):
    return {
        "id": sid,
        "purchase_date": date,
        "supplier_name": supplier,
        "subtotal": subtotal,
        "tax_amount": tax,
        "total_amount": total,
        "invoice_prefix": invoice_prefix,
        "invoice_number": invoice_number,
        "status": status,
    }


def _make_item(name, qty, unit, price, amount, category="蔬菜", account="5110"):
    return {
        "item_name": name,
        "quantity": qty,
        "unit": unit,
        "unit_price": price,
        "amount": amount,
        "category": category,
        "account_code": account,
    }


# ---------------------------------------------------------------------------
# TestAnnualReport
# ---------------------------------------------------------------------------
class TestAnnualReport(unittest.TestCase):

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        for f in os.listdir(self.tmpdir):
            try:
                os.unlink(os.path.join(self.tmpdir, f))
            except OSError:
                pass
        # Also clean nested dirs
        for root, dirs, files in os.walk(self.tmpdir, topdown=False):
            for f in files:
                try:
                    os.unlink(os.path.join(root, f))
                except OSError:
                    pass
            for d in dirs:
                try:
                    os.rmdir(os.path.join(root, d))
                except OSError:
                    pass
        try:
            os.rmdir(self.tmpdir)
        except OSError:
            pass

    @patch("state_manager.get_staging_stats")
    def test_no_data_returns_none(self, mock_stats):
        """所有月份都沒資料 => 回傳 None"""
        mock_stats.return_value = {"total": 0}

        from services.report_service import generate_annual_report
        result = generate_annual_report("2025", output_dir=self.tmpdir)
        self.assertIsNone(result)

    @patch("state_manager.get_purchase_items", return_value=[])
    @patch("state_manager.get_stagings_by_month", return_value=[])
    @patch("state_manager.get_monthly_cost")
    @patch("state_manager.get_staging_stats")
    def test_generates_xlsx_file(self, mock_stats, mock_mc, mock_stg, mock_items):
        """有資料 => 產出 .xlsx 檔案"""
        mock_stats.side_effect = lambda ym: (
            {"total": 5, "total_amount": 50000, "total_tax": 2500}
            if ym == "2025-03" else {"total": 0}
        )
        mock_mc.return_value = {
            "invoice_count": 3, "receipt_count": 2, "is_locked": False,
        }

        from services.report_service import generate_annual_report
        result = generate_annual_report("2025", output_dir=self.tmpdir)
        self.assertIsNotNone(result)
        self.assertTrue(result.endswith(".xlsx"))
        self.assertTrue(os.path.exists(result))

    @patch("state_manager.get_purchase_items", return_value=[])
    @patch("state_manager.get_stagings_by_month", return_value=[])
    @patch("state_manager.get_monthly_cost")
    @patch("state_manager.get_staging_stats")
    def test_correct_filename(self, mock_stats, mock_mc, mock_stg, mock_items):
        """檔名格式：年度報表_YYYY.xlsx"""
        mock_stats.side_effect = lambda ym: (
            {"total": 1, "total_amount": 1000, "total_tax": 50}
            if ym == "2025-01" else {"total": 0}
        )
        mock_mc.return_value = {
            "invoice_count": 1, "receipt_count": 0, "is_locked": False,
        }

        from services.report_service import generate_annual_report
        result = generate_annual_report("2025", output_dir=self.tmpdir)
        self.assertTrue(result.endswith("年度報表_2025.xlsx"))

    @patch("state_manager.get_purchase_items", return_value=[])
    @patch("state_manager.get_stagings_by_month", return_value=[])
    @patch("state_manager.get_monthly_cost")
    @patch("state_manager.get_staging_stats")
    def test_sheet_names(self, mock_stats, mock_mc, mock_stg, mock_items):
        """工作表名稱：月度彙總、分類統計"""
        mock_stats.side_effect = lambda ym: (
            {"total": 2, "total_amount": 2000, "total_tax": 100}
            if ym == "2025-06" else {"total": 0}
        )
        mock_mc.return_value = {
            "invoice_count": 2, "receipt_count": 0, "is_locked": False,
        }

        from services.report_service import generate_annual_report
        filepath = generate_annual_report("2025", output_dir=self.tmpdir)

        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        self.assertEqual(wb.sheetnames, ["月度彙總", "分類統計"])
        wb.close()

    @patch("state_manager.get_purchase_items", return_value=[])
    @patch("state_manager.get_stagings_by_month", return_value=[])
    @patch("state_manager.get_monthly_cost")
    @patch("state_manager.get_staging_stats")
    def test_multiple_months(self, mock_stats, mock_mc, mock_stg, mock_items):
        """多個月份 => 每月一列"""
        mock_stats.side_effect = lambda ym: (
            {"total": 3, "total_amount": 10000, "total_tax": 500}
            if ym in ("2025-01", "2025-02", "2025-03") else {"total": 0}
        )
        mock_mc.return_value = {
            "invoice_count": 2, "receipt_count": 1, "is_locked": False,
        }

        from services.report_service import generate_annual_report
        filepath = generate_annual_report("2025", output_dir=self.tmpdir)

        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb["月度彙總"]
        data_rows = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row[0] and str(row[0]).startswith("2025-"):
                data_rows.append(row)
        self.assertEqual(len(data_rows), 3)
        wb.close()

    @patch("state_manager.get_purchase_items")
    @patch("state_manager.get_stagings_by_month")
    @patch("state_manager.get_monthly_cost")
    @patch("state_manager.get_staging_stats")
    def test_category_aggregation(self, mock_stats, mock_mc, mock_stg, mock_items):
        """分類統計 Sheet — 跨月聚合"""
        mock_stats.side_effect = lambda ym: (
            {"total": 1, "total_amount": 5000, "total_tax": 250}
            if ym in ("2025-01", "2025-02") else {"total": 0}
        )
        mock_mc.return_value = {
            "invoice_count": 1, "receipt_count": 0, "is_locked": False,
        }

        staging_jan = _make_staging(1, "2025-01-10", "菜商A", 4750, 250, 5000)
        staging_feb = _make_staging(2, "2025-02-15", "肉商B", 4750, 250, 5000)

        mock_stg.side_effect = lambda ym: (
            [staging_jan] if ym == "2025-01"
            else [staging_feb] if ym == "2025-02"
            else []
        )

        items_jan = [_make_item("高麗菜", 10, "斤", 35, 350, "蔬菜", "5110")]
        items_feb = [_make_item("豬肉", 5, "斤", 120, 600, "肉類", "5120")]

        mock_items.side_effect = lambda sid: (
            items_jan if sid == 1 else items_feb if sid == 2 else []
        )

        from services.report_service import generate_annual_report
        filepath = generate_annual_report("2025", output_dir=self.tmpdir)

        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws2 = wb["分類統計"]
        categories = []
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0]:
                categories.append(row[0])
        self.assertIn("蔬菜", categories)
        self.assertIn("肉類", categories)
        wb.close()

    @patch("state_manager.get_purchase_items", return_value=[])
    @patch("state_manager.get_stagings_by_month", return_value=[])
    @patch("state_manager.get_monthly_cost")
    @patch("state_manager.get_staging_stats")
    def test_grand_totals_sum(self, mock_stats, mock_mc, mock_stg, mock_items):
        """合計列數值正確"""
        mock_stats.side_effect = lambda ym: (
            {"total": 2, "total_amount": 10000, "total_tax": 500}
            if ym == "2025-01"
            else {"total": 3, "total_amount": 20000, "total_tax": 1000}
            if ym == "2025-02"
            else {"total": 0}
        )
        mock_mc.return_value = {
            "invoice_count": 1, "receipt_count": 0, "is_locked": False,
        }

        from services.report_service import generate_annual_report
        filepath = generate_annual_report("2025", output_dir=self.tmpdir)

        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb["月度彙總"]
        totals_row = None
        for row in ws.iter_rows(values_only=True):
            if row[0] == "合計":
                totals_row = row
                break
        self.assertIsNotNone(totals_row)
        # grand_total (col G / index 6) = 10000 + 20000 = 30000
        self.assertEqual(totals_row[6], 30000)
        wb.close()

    @patch("state_manager.get_purchase_items", return_value=[])
    @patch("state_manager.get_stagings_by_month", return_value=[])
    @patch("state_manager.get_monthly_cost")
    @patch("state_manager.get_staging_stats")
    def test_locked_status(self, mock_stats, mock_mc, mock_stg, mock_items):
        """is_locked => '已鎖定'"""
        mock_stats.side_effect = lambda ym: (
            {"total": 1, "total_amount": 5000, "total_tax": 250}
            if ym == "2025-04" else {"total": 0}
        )
        mock_mc.return_value = {
            "invoice_count": 1, "receipt_count": 0, "is_locked": True,
        }

        from services.report_service import generate_annual_report
        filepath = generate_annual_report("2025", output_dir=self.tmpdir)

        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb["月度彙總"]
        status = ws.cell(row=4, column=8).value
        self.assertEqual(status, "已鎖定")
        wb.close()

    @patch("state_manager.get_purchase_items", return_value=[])
    @patch("state_manager.get_stagings_by_month", return_value=[])
    @patch("state_manager.get_monthly_cost")
    @patch("state_manager.get_staging_stats")
    def test_unlocked_status(self, mock_stats, mock_mc, mock_stg, mock_items):
        """is_locked=False => '開放'"""
        mock_stats.side_effect = lambda ym: (
            {"total": 1, "total_amount": 5000, "total_tax": 250}
            if ym == "2025-07" else {"total": 0}
        )
        mock_mc.return_value = {
            "invoice_count": 1, "receipt_count": 0, "is_locked": False,
        }

        from services.report_service import generate_annual_report
        filepath = generate_annual_report("2025", output_dir=self.tmpdir)

        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb["月度彙總"]
        status = ws.cell(row=4, column=8).value
        self.assertEqual(status, "開放")
        wb.close()

    @patch("state_manager.get_purchase_items", return_value=[])
    @patch("state_manager.get_stagings_by_month", return_value=[])
    @patch("state_manager.get_monthly_cost")
    @patch("state_manager.get_staging_stats")
    def test_no_monthly_cost_record(self, mock_stats, mock_mc, mock_stg, mock_items):
        """get_monthly_cost 回傳 None => invoice/receipt count = 0"""
        mock_stats.side_effect = lambda ym: (
            {"total": 1, "total_amount": 3000, "total_tax": 150}
            if ym == "2025-09" else {"total": 0}
        )
        mock_mc.return_value = None

        from services.report_service import generate_annual_report
        filepath = generate_annual_report("2025", output_dir=self.tmpdir)

        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb["月度彙總"]
        inv_count = ws.cell(row=4, column=3).value
        rec_count = ws.cell(row=4, column=4).value
        self.assertEqual(inv_count, 0)
        self.assertEqual(rec_count, 0)
        wb.close()

    @patch("state_manager.get_purchase_items")
    @patch("state_manager.get_stagings_by_month")
    @patch("state_manager.get_monthly_cost")
    @patch("state_manager.get_staging_stats")
    def test_category_sorted_by_amount_desc(self, mock_stats, mock_mc, mock_stg, mock_items):
        """分類統計按金額降序排列"""
        mock_stats.side_effect = lambda ym: (
            {"total": 1, "total_amount": 10000, "total_tax": 500}
            if ym == "2025-05" else {"total": 0}
        )
        mock_mc.return_value = {
            "invoice_count": 1, "receipt_count": 0, "is_locked": False,
        }

        staging = _make_staging(1, "2025-05-01", "綜合商", 9500, 500, 10000)
        mock_stg.side_effect = lambda ym: (
            [staging] if ym == "2025-05" else []
        )

        mock_items.return_value = [
            _make_item("雞肉", 10, "斤", 80, 800, "肉類", "5120"),
            _make_item("高麗菜", 20, "斤", 35, 700, "蔬菜", "5110"),
            _make_item("醬油", 5, "瓶", 40, 200, "調味料", "5130"),
        ]

        from services.report_service import generate_annual_report
        filepath = generate_annual_report("2025", output_dir=self.tmpdir)

        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws2 = wb["分類統計"]
        amounts = []
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0] and row[3] is not None:
                amounts.append(row[3])
        self.assertEqual(amounts, sorted(amounts, reverse=True))
        wb.close()

    @patch("state_manager.get_purchase_items", return_value=[])
    @patch("state_manager.get_stagings_by_month", return_value=[])
    @patch("state_manager.get_monthly_cost")
    @patch("state_manager.get_staging_stats")
    def test_output_dir_created(self, mock_stats, mock_mc, mock_stg, mock_items):
        """output_dir 不存在 => 自動建立"""
        nested_dir = os.path.join(self.tmpdir, "sub", "reports")
        mock_stats.side_effect = lambda ym: (
            {"total": 1, "total_amount": 1000, "total_tax": 50}
            if ym == "2025-12" else {"total": 0}
        )
        mock_mc.return_value = {
            "invoice_count": 1, "receipt_count": 0, "is_locked": False,
        }

        from services.report_service import generate_annual_report
        result = generate_annual_report("2025", output_dir=nested_dir)
        self.assertIsNotNone(result)
        self.assertTrue(os.path.isdir(nested_dir))


if __name__ == "__main__":
    unittest.main()
