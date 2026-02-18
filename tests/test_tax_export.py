"""小膳 Bot — 稅務匯出服務單元測試

測試範圍：
- MOF TXT 固定寬度格式（81 bytes）
- 文中資訊 Excel 匯出
- 經手人證明表 PDF
- 匯出前 7 點驗證
- 稅期工具函數
"""

import os
import sys
import tempfile
from unittest.mock import MagicMock, patch

import pytest

# 確保能 import 專案模組
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))


# =====================================================================
# Mock 資料
# =====================================================================

def _make_staging(
    staging_id=1,
    supplier_name="好鮮水產行",
    supplier_tax_id="12345678",
    invoice_prefix="AB",
    invoice_number="12345678",
    invoice_type="三聯式",
    invoice_format_code="21",
    purchase_date="2026-01-15",
    subtotal=10000,
    tax_type="1",
    tax_rate=0.05,
    tax_amount=500,
    total_amount=10500,
    deduction_code="1",
    handler_name="",
    handler_note="",
    status="confirmed",
    year_month="2026-01",
    tax_period="2026-01-02",
    notes="",
    supplier_id=1,
    **overrides,
):
    """建立測試用 staging 記錄。"""
    data = {
        "id": staging_id,
        "user_id": "U001",
        "chat_id": "C001",
        "image_message_id": "",
        "local_image_path": "",
        "gdrive_path": "",
        "supplier_id": supplier_id,
        "supplier_name": supplier_name,
        "supplier_tax_id": supplier_tax_id,
        "invoice_prefix": invoice_prefix,
        "invoice_number": invoice_number,
        "invoice_type": invoice_type,
        "invoice_format_code": invoice_format_code,
        "purchase_date": purchase_date,
        "subtotal": subtotal,
        "tax_type": tax_type,
        "tax_rate": tax_rate,
        "tax_amount": tax_amount,
        "total_amount": total_amount,
        "deduction_code": deduction_code,
        "raw_ocr_text": "",
        "ocr_confidence": 0.95,
        "handler_name": handler_name,
        "handler_note": handler_note,
        "status": status,
        "year_month": year_month,
        "tax_period": tax_period,
        "notes": notes,
        "created_at": "2026-01-15 10:00:00",
        "confirmed_at": "2026-01-15 11:00:00",
        "exported_at": None,
    }
    data.update(overrides)
    return data


def _make_market_staging(staging_id=10, **overrides):
    """建立測試用市場採購（免用發票）記錄。"""
    defaults = {
        "staging_id": staging_id,
        "supplier_name": "果菜市場攤販",
        "supplier_tax_id": "",
        "invoice_prefix": "",
        "invoice_number": "",
        "invoice_type": "免用發票",
        "invoice_format_code": "",
        "subtotal": 3000,
        "tax_amount": 0,
        "total_amount": 3000,
        "handler_name": "王小明",
        "deduction_code": "",
        "tax_type": "3",
    }
    defaults.update(overrides)
    return _make_staging(**defaults)


def _make_items(staging_id=1):
    """建立測試用品項明細。"""
    return [
        {
            "id": 1,
            "staging_id": staging_id,
            "ingredient_id": 1,
            "item_name": "高麗菜",
            "brand": "",
            "quantity": 20,
            "unit": "kg",
            "unit_price": 25,
            "amount": 500,
            "tax_amount": 25,
            "category": "蔬菜",
            "account_code": "5110",
            "is_handwritten": 0,
            "original_unit_price": None,
            "confidence": 0.95,
            "created_at": "2026-01-15 10:00:00",
        },
        {
            "id": 2,
            "staging_id": staging_id,
            "ingredient_id": 2,
            "item_name": "豬五花",
            "brand": "",
            "quantity": 10,
            "unit": "kg",
            "unit_price": 180,
            "amount": 1800,
            "tax_amount": 90,
            "category": "肉類",
            "account_code": "5110",
            "is_handwritten": 0,
            "original_unit_price": None,
            "confidence": 0.92,
            "created_at": "2026-01-15 10:00:00",
        },
    ]


# =====================================================================
# Fixtures
# =====================================================================

@pytest.fixture
def output_dir():
    """建立臨時輸出目錄。"""
    with tempfile.TemporaryDirectory(prefix="shanbot_test_") as tmpdir:
        yield tmpdir


@pytest.fixture
def env_vars():
    """設定測試用環境變數。"""
    original = {}
    test_vars = {
        "COMPANY_TAX_ID": "89012345",
        "COMPANY_TAX_REG_NO": "123456789",
        "COMPANY_NAME": "測試團膳有限公司",
    }
    for key, val in test_vars.items():
        original[key] = os.environ.get(key)
        os.environ[key] = val
    yield test_vars
    # 還原
    for key, orig_val in original.items():
        if orig_val is None:
            os.environ.pop(key, None)
        else:
            os.environ[key] = orig_val


@pytest.fixture
def mock_sm():
    """Mock state_manager 的所有相關函數。"""
    stagings = [
        _make_staging(staging_id=1, purchase_date="2026-01-15"),
        _make_staging(
            staging_id=2,
            supplier_name="大豐肉品",
            supplier_tax_id="87654321",
            invoice_prefix="CD",
            invoice_number="00987654",
            purchase_date="2026-02-10",
            subtotal=20000,
            tax_amount=1000,
            total_amount=21000,
        ),
    ]
    items_map = {
        1: _make_items(staging_id=1),
        2: [
            {
                "id": 3,
                "staging_id": 2,
                "ingredient_id": 3,
                "item_name": "里肌肉",
                "brand": "大豐",
                "quantity": 15,
                "unit": "kg",
                "unit_price": 200,
                "amount": 3000,
                "tax_amount": 150,
                "category": "肉類",
                "account_code": "5110",
                "is_handwritten": 0,
                "original_unit_price": None,
                "confidence": 0.9,
                "created_at": "2026-02-10 10:00:00",
            }
        ],
    }

    with patch("services.tax_export_service.sm") as mock:
        mock.get_confirmed_stagings.return_value = stagings
        mock.get_purchase_items.side_effect = lambda sid: items_map.get(sid, [])
        mock.get_supplier.return_value = {"id": 1, "name": "好鮮水產行", "has_uniform_invoice": 1}
        mock.add_tax_export.return_value = 1
        mock.get_account_mapping.return_value = {"account_code": "5110", "account_name": "進貨—蔬菜類"}
        yield mock, stagings, items_map


# =====================================================================
# 1. MOF TXT 格式測試
# =====================================================================

class TestMofTxt:
    """MOF 進項 TXT 匯出測試。"""

    def test_line_length_exactly_81_bytes(self, output_dir, env_vars, mock_sm):
        """每行長度必須恰好 81 bytes。"""
        # 重新載入模組以取得新的環境變數
        import services.tax_export_service as svc
        svc.COMPANY_TAX_ID = env_vars["COMPANY_TAX_ID"]
        svc.COMPANY_TAX_REG_NO = env_vars["COMPANY_TAX_REG_NO"]

        filepath = svc.export_mof_txt("2026-01-02", output_dir)
        assert os.path.exists(filepath)

        with open(filepath, "rb") as f:
            content = f.read()

        # 以 CRLF 分割（二進位模式確保 \r\n 不被轉換）
        raw_lines = content.split(b"\r\n")
        # 移除最後空行
        if raw_lines and raw_lines[-1] == b"":
            raw_lines = raw_lines[:-1]

        assert len(raw_lines) == 2, f"應有 2 行，實際 {len(raw_lines)} 行"

        for i, line in enumerate(raw_lines):
            assert len(line) == 81, (
                f"第 {i+1} 行長度 {len(line)} bytes，應為 81 bytes\n"
                f"內容: '{line.decode('ascii')}'"
            )

    def test_field_positions(self, output_dir, env_vars, mock_sm):
        """驗證各欄位位置正確（0-indexed）。"""
        import services.tax_export_service as svc
        svc.COMPANY_TAX_ID = env_vars["COMPANY_TAX_ID"]
        svc.COMPANY_TAX_REG_NO = env_vars["COMPANY_TAX_REG_NO"]

        filepath = svc.export_mof_txt("2026-01-02", output_dir)

        with open(filepath, "rb") as f:
            lines = [l.decode("ascii") for l in f.read().split(b"\r\n") if l]
        line = lines[0]

        # 格式代號（位置 0-1，2 碼）
        assert line[0:2] == "21", f"格式代號: '{line[0:2]}'"

        # 稅籍編號（位置 2-10，9 碼）
        assert line[2:11] == "123456789", f"稅籍編號: '{line[2:11]}'"

        # 流水號（位置 11-17，7 碼，第一筆 0000001）
        assert line[11:18] == "0000001", f"流水號: '{line[11:18]}'"

        # 所屬年度（位置 18-20，3 碼，民國 115 年）
        assert line[18:21] == "115", f"所屬年度: '{line[18:21]}'"

        # 所屬月份（位置 21-22，2 碼）
        assert line[21:23] == "02", f"所屬月份: '{line[21:23]}'"

        # 買受人統編（位置 23-30，8 碼）
        assert line[23:31] == "89012345", f"買受人統編: '{line[23:31]}'"

        # 銷售人統編（位置 31-38，8 碼）
        assert line[31:39] == "12345678", f"銷售人統編: '{line[31:39]}'"

        # 發票字軌（位置 39-40，2 碼）
        assert line[39:41] == "AB", f"發票字軌: '{line[39:41]}'"

        # 發票號碼（位置 41-48，8 碼）
        assert line[41:49] == "12345678", f"發票號碼: '{line[41:49]}'"

        # 銷售額（位置 49-60，12 碼）
        assert line[49:61] == "000000010000", f"銷售額: '{line[49:61]}'"

        # 課稅別（位置 61，1 碼）
        assert line[61] == "1", f"課稅別: '{line[61]}'"

        # 稅額（位置 62-71，10 碼）
        assert line[62:72] == "0000000500", f"稅額: '{line[62:72]}'"

        # 扣抵代號（位置 72，1 碼）
        assert line[72] == "1", f"扣抵代號: '{line[72]}'"

        # 保留（位置 73-77，5 碼空白）
        assert line[73:78] == "     ", f"保留欄: '{line[73:78]}'"

        # 特種稅率（位置 78，1 碼空白）
        assert line[78] == " ", f"特種稅率: '{line[78]}'"

        # 彙加註記（位置 79，1 碼空白）
        assert line[79] == " ", f"彙加註記: '{line[79]}'"

        # 通關註記（位置 80，1 碼空白）
        assert line[80] == " ", f"通關註記: '{line[80]}'"

    def test_sequential_numbering(self, output_dir, env_vars, mock_sm):
        """流水號應連續遞增。"""
        import services.tax_export_service as svc
        svc.COMPANY_TAX_ID = env_vars["COMPANY_TAX_ID"]
        svc.COMPANY_TAX_REG_NO = env_vars["COMPANY_TAX_REG_NO"]

        filepath = svc.export_mof_txt("2026-01-02", output_dir)

        with open(filepath, "rb") as f:
            lines = [l.decode("ascii") for l in f.read().split(b"\r\n") if l]

        for i, line in enumerate(lines):
            expected_seq = str(i + 1).zfill(7)
            actual_seq = line[11:18]
            assert actual_seq == expected_seq, (
                f"第 {i+1} 行流水號: '{actual_seq}'，應為 '{expected_seq}'"
            )

    def test_no_data_raises_error(self, output_dir, env_vars):
        """無資料時應拋出 ValueError。"""
        import services.tax_export_service as svc
        svc.COMPANY_TAX_ID = env_vars["COMPANY_TAX_ID"]
        svc.COMPANY_TAX_REG_NO = env_vars["COMPANY_TAX_REG_NO"]

        with patch("services.tax_export_service.sm") as mock:
            mock.get_confirmed_stagings.return_value = []
            with pytest.raises(ValueError, match="無已確認"):
                svc.export_mof_txt("2026-01-02", output_dir)

    def test_missing_env_raises_error(self, output_dir):
        """環境變數未設定時應拋出 ValueError。"""
        import services.tax_export_service as svc
        svc.COMPANY_TAX_ID = ""
        svc.COMPANY_TAX_REG_NO = ""

        with pytest.raises(ValueError, match="COMPANY_TAX_ID"):
            svc.export_mof_txt("2026-01-02", output_dir)

    def test_market_purchases_excluded(self, output_dir, env_vars):
        """免用發票的市場採購不應出現在 MOF TXT。"""
        import services.tax_export_service as svc
        svc.COMPANY_TAX_ID = env_vars["COMPANY_TAX_ID"]
        svc.COMPANY_TAX_REG_NO = env_vars["COMPANY_TAX_REG_NO"]

        stagings = [
            _make_staging(staging_id=1),
            _make_market_staging(staging_id=2),  # 免用發票
        ]

        with patch("services.tax_export_service.sm") as mock:
            mock.get_confirmed_stagings.return_value = stagings
            mock.add_tax_export.return_value = 1
            filepath = svc.export_mof_txt("2026-01-02", output_dir)

        with open(filepath, "rb") as f:
            lines = [l.decode("ascii") for l in f.read().split(b"\r\n") if l]

        # 只應有 1 行（市場採購被排除）
        assert len(lines) == 1

    def test_roc_year_conversion(self, output_dir, env_vars, mock_sm):
        """西元年應正確轉換為民國年。"""
        import services.tax_export_service as svc
        svc.COMPANY_TAX_ID = env_vars["COMPANY_TAX_ID"]
        svc.COMPANY_TAX_REG_NO = env_vars["COMPANY_TAX_REG_NO"]

        filepath = svc.export_mof_txt("2026-01-02", output_dir)

        with open(filepath, "rb") as f:
            line = f.read().split(b"\r\n")[0].decode("ascii")

        # 2026 - 1911 = 115
        assert line[18:21] == "115"

    def test_export_record_created(self, output_dir, env_vars, mock_sm):
        """匯出後應在 tax_exports 記錄。"""
        import services.tax_export_service as svc
        svc.COMPANY_TAX_ID = env_vars["COMPANY_TAX_ID"]
        svc.COMPANY_TAX_REG_NO = env_vars["COMPANY_TAX_REG_NO"]

        mock, _, _ = mock_sm
        svc.export_mof_txt("2026-01-02", output_dir)

        mock.add_tax_export.assert_called_once()
        call_kwargs = mock.add_tax_export.call_args
        assert call_kwargs[1]["export_type"] == "MOF_TXT"
        assert call_kwargs[1]["record_count"] == 2

    def test_all_lines_ascii_only(self, output_dir, env_vars, mock_sm):
        """所有行只能包含 ASCII 字元。"""
        import services.tax_export_service as svc
        svc.COMPANY_TAX_ID = env_vars["COMPANY_TAX_ID"]
        svc.COMPANY_TAX_REG_NO = env_vars["COMPANY_TAX_REG_NO"]

        filepath = svc.export_mof_txt("2026-01-02", output_dir)

        with open(filepath, "rb") as f:
            content = f.read()

        # 除了 \r\n 之外，所有字元應在 0x20-0x7E
        for byte in content:
            assert byte in (0x0D, 0x0A) or (0x20 <= byte <= 0x7E), (
                f"非 ASCII 字元: 0x{byte:02X}"
            )


# =====================================================================
# 2. 文中資訊 Excel 匯出測試
# =====================================================================

class TestWintonExcel:
    """文中資訊 Excel 匯出測試。"""

    def test_excel_file_created(self, output_dir, env_vars, mock_sm):
        """Excel 檔應成功建立。"""
        import services.tax_export_service as svc
        filepath = svc.export_winton_excel("2026-01-02", output_dir)
        assert os.path.exists(filepath)
        assert filepath.endswith(".xlsx")

    def test_headers_correct(self, output_dir, env_vars, mock_sm):
        """標題列應正確且完整。"""
        import services.tax_export_service as svc
        import openpyxl

        filepath = svc.export_winton_excel("2026-01-02", output_dir)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        expected_headers = [
            "日期", "傳票號碼", "摘要", "借方科目", "借方金額",
            "貸方科目", "貸方金額", "進項稅額", "發票號碼", "統一編號", "備註", "扣抵類別",
        ]

        for col_idx, expected in enumerate(expected_headers, start=1):
            actual = ws.cell(row=1, column=col_idx).value
            assert actual == expected, f"欄 {col_idx}: '{actual}' != '{expected}'"

    def test_headers_bold(self, output_dir, env_vars, mock_sm):
        """標題列應為粗體。"""
        import services.tax_export_service as svc
        import openpyxl

        filepath = svc.export_winton_excel("2026-01-02", output_dir)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        for col_idx in range(1, 13):
            cell = ws.cell(row=1, column=col_idx)
            assert cell.font.bold, f"欄 {col_idx} 標題未設粗體"

    def test_data_rows_count(self, output_dir, env_vars, mock_sm):
        """資料列數應與 staging 記錄數一致。"""
        import services.tax_export_service as svc
        import openpyxl

        filepath = svc.export_winton_excel("2026-01-02", output_dir)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # 第 1 列為標題，第 2-3 列為資料
        assert ws.max_row == 3, f"列數: {ws.max_row}（應為 3：1 標題 + 2 資料）"

    def test_column_a_date_format(self, output_dir, env_vars, mock_sm):
        """A 欄日期格式應為 YYYY/MM/DD。"""
        import services.tax_export_service as svc
        import openpyxl

        filepath = svc.export_winton_excel("2026-01-02", output_dir)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        date_val = ws.cell(row=2, column=1).value
        assert date_val == "2026/01/15", f"日期: '{date_val}'"

    def test_column_b_voucher_sequential(self, output_dir, env_vars, mock_sm):
        """B 欄傳票號碼應遞增。"""
        import services.tax_export_service as svc
        import openpyxl

        filepath = svc.export_winton_excel("2026-01-02", output_dir)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        assert ws.cell(row=2, column=2).value == "V0001"
        assert ws.cell(row=3, column=2).value == "V0002"

    def test_column_f_payable_account(self, output_dir, env_vars, mock_sm):
        """F 欄應固定為 2101（應付帳款）。"""
        import services.tax_export_service as svc
        import openpyxl

        filepath = svc.export_winton_excel("2026-01-02", output_dir)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=6).value
            assert val == "2101", f"第 {row} 列貸方科目: '{val}'"

    def test_column_i_invoice_number(self, output_dir, env_vars, mock_sm):
        """I 欄發票號碼 = 字軌 + 號碼。"""
        import services.tax_export_service as svc
        import openpyxl

        filepath = svc.export_winton_excel("2026-01-02", output_dir)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        val = ws.cell(row=2, column=9).value
        assert val == "AB12345678", f"發票號碼: '{val}'"

    def test_amounts_numeric(self, output_dir, env_vars, mock_sm):
        """金額欄位應為數值型態。"""
        import services.tax_export_service as svc
        import openpyxl

        filepath = svc.export_winton_excel("2026-01-02", output_dir)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # E 欄（借方金額）
        assert isinstance(ws.cell(row=2, column=5).value, (int, float))
        # G 欄（貸方金額）
        assert isinstance(ws.cell(row=2, column=7).value, (int, float))
        # H 欄（進項稅額）
        assert isinstance(ws.cell(row=2, column=8).value, (int, float))

    def test_no_data_raises_error(self, output_dir, env_vars):
        """無資料時應拋出 ValueError。"""
        import services.tax_export_service as svc

        with patch("services.tax_export_service.sm") as mock:
            mock.get_confirmed_stagings.return_value = []
            with pytest.raises(ValueError, match="無已確認"):
                svc.export_winton_excel("2026-01-02", output_dir)

    def test_export_record_created(self, output_dir, env_vars, mock_sm):
        """匯出後應記錄至 tax_exports。"""
        import services.tax_export_service as svc
        mock, _, _ = mock_sm
        svc.export_winton_excel("2026-01-02", output_dir)

        mock.add_tax_export.assert_called_once()
        call_kwargs = mock.add_tax_export.call_args
        assert call_kwargs[1]["export_type"] == "WINTON_EXCEL"

    def test_deduction_column_present(self, output_dir, env_vars, mock_sm):
        """L 欄扣抵類別應存在且正確。"""
        import services.tax_export_service as svc
        import openpyxl

        filepath = svc.export_winton_excel("2026-01-02", output_dir)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # 標題
        assert ws.cell(row=1, column=12).value == "扣抵類別"
        # 資料（staging deduction_code 預設為 "1" → 可扣抵）
        val = ws.cell(row=2, column=12).value
        assert val == "可扣抵", f"扣抵類別: '{val}'"

    def test_deduction_column_non_deductible(self, output_dir, env_vars):
        """不可扣抵的記錄應顯示「不可扣抵」。"""
        import services.tax_export_service as svc
        import openpyxl

        stagings = [_make_staging(staging_id=1, deduction_code="2")]
        with patch("services.tax_export_service.sm") as mock:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_purchase_items.return_value = _make_items()
            mock.get_account_mapping.return_value = {"account_code": "5110"}
            mock.add_tax_export.return_value = 1

            filepath = svc.export_winton_excel("2026-01-02", output_dir)

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        val = ws.cell(row=2, column=12).value
        assert val == "不可扣抵", f"扣抵類別: '{val}'"


# =====================================================================
# 3. 經手人證明表 PDF 測試
# =====================================================================

class TestHandlerCert:
    """經手人證明表 PDF 匯出測試。"""

    def test_pdf_file_created(self, output_dir, env_vars):
        """PDF 檔應成功建立。"""
        import services.tax_export_service as svc
        svc.COMPANY_NAME = env_vars["COMPANY_NAME"]

        market_stagings = [
            _make_market_staging(staging_id=10, purchase_date="2026-01-15"),
        ]

        with patch("services.tax_export_service.sm") as mock:
            mock.get_confirmed_stagings.return_value = market_stagings
            mock.get_purchase_items.return_value = _make_items(staging_id=10)
            mock.get_supplier.return_value = {
                "id": 99, "name": "果菜市場攤販", "has_uniform_invoice": 0,
            }
            mock.add_tax_export.return_value = 1
            filepath = svc.export_handler_cert("2026-01-02", output_dir)

        assert os.path.exists(filepath)
        assert filepath.endswith(".pdf")
        # PDF 檔案不應為空
        assert os.path.getsize(filepath) > 0

    def test_no_market_data_raises_error(self, output_dir, env_vars):
        """無市場採購時應拋出 ValueError。"""
        import services.tax_export_service as svc

        # 只有有統編的發票記錄
        normal_stagings = [_make_staging(staging_id=1)]

        with patch("services.tax_export_service.sm") as mock:
            mock.get_confirmed_stagings.return_value = normal_stagings
            mock.get_supplier.return_value = {
                "id": 1, "name": "好鮮水產行", "has_uniform_invoice": 1,
            }
            with pytest.raises(ValueError, match="無免用發票"):
                svc.export_handler_cert("2026-01-02", output_dir)

    def test_multiple_dates_separate_pages(self, output_dir, env_vars):
        """不同日期應分頁。"""
        import services.tax_export_service as svc
        svc.COMPANY_NAME = env_vars["COMPANY_NAME"]

        # 兩天的市場採購
        market_stagings = [
            _make_market_staging(staging_id=10, purchase_date="2026-01-15"),
            _make_market_staging(staging_id=11, purchase_date="2026-01-16"),
        ]

        with patch("services.tax_export_service.sm") as mock:
            mock.get_confirmed_stagings.return_value = market_stagings
            mock.get_purchase_items.return_value = _make_items()
            mock.get_supplier.return_value = {
                "id": 99, "name": "果菜市場攤販", "has_uniform_invoice": 0,
            }
            mock.add_tax_export.return_value = 1
            filepath = svc.export_handler_cert("2026-01-02", output_dir)

        assert os.path.exists(filepath)
        # PDF 存在且大小合理（兩頁應比一頁大）
        assert os.path.getsize(filepath) > 1000

    def test_export_record_created(self, output_dir, env_vars):
        """匯出後應記錄至 tax_exports。"""
        import services.tax_export_service as svc
        svc.COMPANY_NAME = env_vars["COMPANY_NAME"]

        market_stagings = [
            _make_market_staging(staging_id=10, purchase_date="2026-01-15"),
        ]

        with patch("services.tax_export_service.sm") as mock:
            mock.get_confirmed_stagings.return_value = market_stagings
            mock.get_purchase_items.return_value = _make_items(staging_id=10)
            mock.get_supplier.return_value = {
                "id": 99, "name": "果菜市場攤販", "has_uniform_invoice": 0,
            }
            mock.add_tax_export.return_value = 1
            svc.export_handler_cert("2026-01-02", output_dir)

        mock.add_tax_export.assert_called_once()
        call_kwargs = mock.add_tax_export.call_args
        assert call_kwargs[1]["export_type"] == "HANDLER_CERT"


# =====================================================================
# 4. 匯出前驗證測試
# =====================================================================

class TestValidation:
    """匯出前 7 點驗證測試。"""

    def test_all_pass(self):
        """全部通過場景。"""
        import services.tax_export_service as svc

        stagings = [
            _make_staging(staging_id=1),
            _make_staging(staging_id=2, supplier_tax_id="87654321",
                          invoice_prefix="CD", invoice_number="00987654",
                          subtotal=20000, tax_amount=1000, total_amount=21000),
        ]

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 1, "has_uniform_invoice": 1}
            mock_pending.return_value = []  # 無 pending

            is_valid, errors = svc.validate_before_export("2026-01-02")

        assert is_valid is True
        assert errors == []

    def test_check1_pending_records(self):
        """檢核 1: 有未確認記錄應失敗。"""
        import services.tax_export_service as svc

        stagings = [_make_staging(staging_id=1)]
        pending = [_make_staging(staging_id=99, status="pending")]

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 1, "has_uniform_invoice": 1}
            mock_pending.return_value = pending

            is_valid, errors = svc.validate_before_export("2026-01-02")

        assert is_valid is False
        assert any("[1]" in e for e in errors)

    def test_check2_missing_tax_id(self):
        """檢核 2: 發票缺統編應失敗。"""
        import services.tax_export_service as svc

        stagings = [_make_staging(staging_id=1, supplier_tax_id="")]

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 1, "has_uniform_invoice": 1}
            mock_pending.return_value = []

            is_valid, errors = svc.validate_before_export("2026-01-02")

        assert is_valid is False
        assert any("[2]" in e for e in errors)

    def test_check2_market_purchase_exempt(self):
        """檢核 2: 市場採購免統編檢查。"""
        import services.tax_export_service as svc

        # 市場採購無統編但不應報錯
        stagings = [
            _make_staging(staging_id=1),  # 正常發票
            _make_market_staging(staging_id=2),  # 市場採購（無統編）
        ]

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 99, "has_uniform_invoice": 0}
            mock_pending.return_value = []

            is_valid, errors = svc.validate_before_export("2026-01-02")

        # 不應因市場採購缺統編而報錯（但市場採購缺經手人可能報別的錯）
        check2_errors = [e for e in errors if "[2]" in e]
        assert len(check2_errors) == 0

    def test_check3_invalid_invoice_prefix(self):
        """檢核 3: 發票字軌格式錯誤。"""
        import services.tax_export_service as svc

        stagings = [_make_staging(staging_id=1, invoice_prefix="A1")]  # 應為 2 英文

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 1, "has_uniform_invoice": 1}
            mock_pending.return_value = []

            is_valid, errors = svc.validate_before_export("2026-01-02")

        assert is_valid is False
        assert any("[3]" in e for e in errors)

    def test_check3_invalid_invoice_number(self):
        """檢核 3: 發票號碼格式錯誤。"""
        import services.tax_export_service as svc

        stagings = [_make_staging(staging_id=1, invoice_number="ABCD")]  # 非數字

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 1, "has_uniform_invoice": 1}
            mock_pending.return_value = []

            is_valid, errors = svc.validate_before_export("2026-01-02")

        assert is_valid is False
        assert any("[3]" in e for e in errors)

    def test_check4_tax_amount_mismatch(self):
        """檢核 4: 稅額與未稅金額不符。"""
        import services.tax_export_service as svc

        # subtotal=10000 → 預期稅額=500，但填了 800
        stagings = [_make_staging(staging_id=1, subtotal=10000, tax_amount=800,
                                  total_amount=10800)]

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 1, "has_uniform_invoice": 1}
            mock_pending.return_value = []

            is_valid, errors = svc.validate_before_export("2026-01-02")

        assert is_valid is False
        assert any("[4]" in e for e in errors)

    def test_check4_tax_tolerance_within_1(self):
        """檢核 4: 稅額在容差 1 元內應通過。"""
        import services.tax_export_service as svc

        # subtotal=10001 → 預期稅額=500.05，四捨五入=500，差 0.05 < 1
        stagings = [_make_staging(staging_id=1, subtotal=10001, tax_amount=500,
                                  total_amount=10501)]

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 1, "has_uniform_invoice": 1}
            mock_pending.return_value = []

            is_valid, errors = svc.validate_before_export("2026-01-02")

        check4_errors = [e for e in errors if "[4]" in e]
        assert len(check4_errors) == 0

    def test_check5_debit_credit_imbalance(self):
        """檢核 5: 借貸不平衡。"""
        import services.tax_export_service as svc

        # subtotal(10000) + tax(500) = 10500 (借方)
        # total_amount = 9000 (貸方) → 不平衡
        stagings = [_make_staging(staging_id=1, subtotal=10000, tax_amount=500,
                                  total_amount=9000)]

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 1, "has_uniform_invoice": 1}
            mock_pending.return_value = []

            is_valid, errors = svc.validate_before_export("2026-01-02")

        assert is_valid is False
        assert any("[5]" in e for e in errors)

    def test_check6_period_mismatch(self):
        """檢核 6: 稅期不一致。"""
        import services.tax_export_service as svc

        stagings = [_make_staging(staging_id=1, tax_period="2026-03-04")]  # 不同稅期

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 1, "has_uniform_invoice": 1}
            mock_pending.return_value = []

            is_valid, errors = svc.validate_before_export("2026-01-02")

        assert is_valid is False
        assert any("[6]" in e for e in errors)

    def test_check7_market_missing_handler(self):
        """檢核 7: 市場採購缺經手人。"""
        import services.tax_export_service as svc

        stagings = [_make_market_staging(staging_id=1, handler_name="")]

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 99, "has_uniform_invoice": 0}
            mock_pending.return_value = []

            is_valid, errors = svc.validate_before_export("2026-01-02")

        assert is_valid is False
        assert any("[7]" in e for e in errors)

    def test_check7_market_with_handler_passes(self):
        """檢核 7: 市場採購有經手人應通過。"""
        import services.tax_export_service as svc

        stagings = [_make_market_staging(staging_id=1, handler_name="王小明")]

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 99, "has_uniform_invoice": 0}
            mock_pending.return_value = []

            is_valid, errors = svc.validate_before_export("2026-01-02")

        check7_errors = [e for e in errors if "[7]" in e]
        assert len(check7_errors) == 0

    def test_check8_deduction_inconsistency_triplicate_wrong_code(self):
        """檢核 8: 有統編+三聯式但 deduction_code!=1 應失敗。"""
        import services.tax_export_service as svc

        stagings = [_make_staging(
            staging_id=1,
            supplier_tax_id="12345678",
            invoice_type="三聯式",
            deduction_code="2",  # 錯誤：三聯式應為 1
        )]

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 1, "has_uniform_invoice": 1}
            mock_pending.return_value = []

            is_valid, errors = svc.validate_before_export("2026-01-02")

        assert is_valid is False
        assert any("[8]" in e for e in errors)

    def test_check8_no_tax_id_with_deduction1_should_fail(self):
        """檢核 8: 無統編但 deduction_code=1 應失敗。"""
        import services.tax_export_service as svc

        stagings = [_make_staging(
            staging_id=1,
            supplier_tax_id="",
            invoice_type="收據",
            deduction_code="1",  # 錯誤：無統編不應為 1
        )]

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 1, "has_uniform_invoice": 1}
            mock_pending.return_value = []

            is_valid, errors = svc.validate_before_export("2026-01-02")

        assert any("[8]" in e for e in errors)

    def test_check8_consistent_deduction_passes(self):
        """檢核 8: 正確的扣抵分類應通過。"""
        import services.tax_export_service as svc

        stagings = [_make_staging(
            staging_id=1,
            supplier_tax_id="12345678",
            invoice_type="三聯式",
            deduction_code="1",  # 正確
        )]

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 1, "has_uniform_invoice": 1}
            mock_pending.return_value = []

            is_valid, errors = svc.validate_before_export("2026-01-02")

        check8_errors = [e for e in errors if "[8]" in e]
        assert len(check8_errors) == 0

    def test_multiple_errors_collected(self):
        """多項錯誤應全部收集（不因第一項失敗就中斷）。"""
        import services.tax_export_service as svc

        stagings = [
            _make_staging(staging_id=1, supplier_tax_id="", invoice_prefix="X",
                          subtotal=10000, tax_amount=999, total_amount=8000),
        ]

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = stagings
            mock.get_supplier.return_value = {"id": 1, "has_uniform_invoice": 1}
            mock_pending.return_value = [_make_staging(staging_id=99, status="pending")]

            is_valid, errors = svc.validate_before_export("2026-01-02")

        assert is_valid is False
        # 應有多項錯誤
        assert len(errors) >= 3

    def test_no_records_returns_error(self):
        """無任何記錄應回傳錯誤。"""
        import services.tax_export_service as svc

        with patch("services.tax_export_service.sm") as mock, \
             patch("services.tax_export_service._get_pending_for_period") as mock_pending:
            mock.get_confirmed_stagings.return_value = []
            mock_pending.return_value = []

            is_valid, errors = svc.validate_before_export("2026-01-02")

        assert is_valid is False
        assert any("[0]" in e for e in errors)


# =====================================================================
# 5. 稅期工具函數測試
# =====================================================================

class TestTaxPeriod:
    """稅期工具函數測試。"""

    def test_jan_feb(self):
        """1-2 月稅期。"""
        from services.tax_export_service import get_tax_period_months
        assert get_tax_period_months("2026-01-02") == (1, 2)

    def test_mar_apr(self):
        """3-4 月稅期。"""
        from services.tax_export_service import get_tax_period_months
        assert get_tax_period_months("2026-03-04") == (3, 4)

    def test_may_jun(self):
        """5-6 月稅期。"""
        from services.tax_export_service import get_tax_period_months
        assert get_tax_period_months("2026-05-06") == (5, 6)

    def test_jul_aug(self):
        """7-8 月稅期。"""
        from services.tax_export_service import get_tax_period_months
        assert get_tax_period_months("2026-07-08") == (7, 8)

    def test_sep_oct(self):
        """9-10 月稅期。"""
        from services.tax_export_service import get_tax_period_months
        assert get_tax_period_months("2026-09-10") == (9, 10)

    def test_nov_dec(self):
        """11-12 月稅期。"""
        from services.tax_export_service import get_tax_period_months
        assert get_tax_period_months("2026-11-12") == (11, 12)

    def test_invalid_format_no_dash(self):
        """格式錯誤（缺分隔符）。"""
        from services.tax_export_service import get_tax_period_months
        with pytest.raises(ValueError, match="格式錯誤"):
            get_tax_period_months("202601")

    def test_invalid_even_start_month(self):
        """起始月為偶數應錯誤。"""
        from services.tax_export_service import get_tax_period_months
        with pytest.raises(ValueError, match="起始月份不合法"):
            get_tax_period_months("2026-02-03")

    def test_invalid_non_consecutive(self):
        """結束月非起始月+1 應錯誤。"""
        from services.tax_export_service import get_tax_period_months
        with pytest.raises(ValueError, match="結束月份不合法"):
            get_tax_period_months("2026-01-04")

    def test_different_years(self):
        """不同年份應正常解析。"""
        from services.tax_export_service import get_tax_period_months
        assert get_tax_period_months("2025-11-12") == (11, 12)
        assert get_tax_period_months("2030-01-02") == (1, 2)


# =====================================================================
# 輔助函數測試
# =====================================================================

class TestHelpers:
    """輔助函數測試。"""

    def test_is_market_purchase_by_invoice_type(self):
        """invoice_type='免用發票' 應判定為市場採購。"""
        from services.tax_export_service import _is_market_purchase
        staging = _make_market_staging()
        with patch("services.tax_export_service.sm"):
            assert _is_market_purchase(staging) is True

    def test_is_market_purchase_by_supplier_flag(self):
        """供應商 has_uniform_invoice=0 應判定為市場採購。"""
        from services.tax_export_service import _is_market_purchase
        staging = _make_staging(
            invoice_type="",  # 未明確標記
            supplier_id=99,
        )
        with patch("services.tax_export_service.sm") as mock:
            mock.get_supplier.return_value = {
                "id": 99, "name": "菜市場", "has_uniform_invoice": 0,
            }
            assert _is_market_purchase(staging) is True

    def test_is_not_market_purchase(self):
        """一般發票不應判定為市場採購。"""
        from services.tax_export_service import _is_market_purchase
        staging = _make_staging()
        with patch("services.tax_export_service.sm") as mock:
            mock.get_supplier.return_value = {
                "id": 1, "name": "好鮮水產行", "has_uniform_invoice": 1,
            }
            assert _is_market_purchase(staging) is False

    def test_format_number_integer(self):
        """整數格式化。"""
        from services.tax_export_service import _format_number
        assert _format_number(1000) == "1,000"
        assert _format_number(0) == "0"

    def test_format_number_float(self):
        """浮點數格式化。"""
        from services.tax_export_service import _format_number
        assert _format_number(10.5) == "10.50"
        assert _format_number(10.0) == "10"  # 整數不帶小數

    def test_format_number_none(self):
        """None 應回傳 "0"。"""
        from services.tax_export_service import _format_number
        assert _format_number(None) == "0"
