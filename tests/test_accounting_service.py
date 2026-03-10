"""會計自動化服務測試 — 完整會計循環驗證"""

import asyncio
import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))


def _run(coro):
    return asyncio.get_event_loop().run_until_complete(coro)


_original_db_path = None


def _setup_db():
    global _original_db_path
    import state_manager as sm
    _original_db_path = sm.DB_PATH
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    sm.DB_PATH = path
    sm.init_db()
    return path


def _teardown_db(path):
    import state_manager as sm
    sm.DB_PATH = _original_db_path
    try:
        os.unlink(path)
    except OSError:
        pass


def _create_confirmed_staging(sm, supplier="好鮮水產行", total=5000,
                               subtotal=4762, tax=238):
    """建立已確認的 staging 記錄"""
    sid = sm.add_purchase_staging("U001", "C001", purchase_date="2026-03-15")
    sm.update_purchase_staging(sid, supplier_name=supplier, total_amount=total,
                               subtotal=subtotal, tax_amount=tax,
                               year_month="2026-03", status="confirmed",
                               deduction_code="1")
    return sid


def _add_items(sm, staging_id):
    """加入品項明細"""
    sm.add_purchase_item(staging_id, item_name="高麗菜", quantity=10, unit="斤",
                         unit_price=35, amount=350, category="蔬菜")
    sm.add_purchase_item(staging_id, item_name="豬肉", quantity=5, unit="斤",
                         unit_price=180, amount=900, category="肉類")
    sm.add_purchase_item(staging_id, item_name="雞蛋", quantity=2, unit="箱",
                         unit_price=600, amount=1200, category="蛋豆")


class TestJournalEntryGeneration(unittest.TestCase):
    """Test 1: 分錄生成 — 借方＝貸方"""

    def setUp(self):
        self.db_path = _setup_db()

    def tearDown(self):
        _teardown_db(self.db_path)

    def test_basic_journal_entries(self):
        """基本分錄生成 — 有明細、有稅"""
        import state_manager as sm
        from services.accounting_service import generate_journal_entries, verify_balance

        sid = _create_confirmed_staging(sm)
        _add_items(sm, sid)

        entries = generate_journal_entries(sid)
        self.assertGreater(len(entries), 0, "應生成分錄")

        # 借方包含進貨科目
        debit_entries = [e for e in entries if e["side"] == "debit"]
        self.assertGreater(len(debit_entries), 0, "應有借方分錄")

        # 貸方包含現金
        credit_entries = [e for e in entries if e["side"] == "credit"]
        self.assertGreater(len(credit_entries), 0, "應有貸方分錄")

        # 借貸平衡
        balance = verify_balance(sid)
        self.assertTrue(balance["balanced"], f"借貸不平衡！差額: {balance['difference']}")

    def test_no_items_journal(self):
        """無明細 — 整筆作為進貨"""
        import state_manager as sm
        from services.accounting_service import generate_journal_entries, verify_balance

        sid = _create_confirmed_staging(sm, total=3000, subtotal=2857, tax=143)
        # 不加品項
        entries = generate_journal_entries(sid)
        self.assertGreater(len(entries), 0)

        balance = verify_balance(sid)
        self.assertTrue(balance["balanced"])

    def test_no_tax_deduction(self):
        """不可扣抵 — 無進項稅額分錄"""
        import state_manager as sm
        from services.accounting_service import generate_journal_entries

        sid = sm.add_purchase_staging("U001", "C001", purchase_date="2026-03-15")
        sm.update_purchase_staging(sid, supplier_name="菜市場阿嬤",
                                   total_amount=800, subtotal=800, tax_amount=0,
                                   year_month="2026-03", status="confirmed",
                                   deduction_code="3")
        sm.add_purchase_item(sid, item_name="空心菜", quantity=5, unit="把",
                             unit_price=30, amount=150, category="蔬菜")

        entries = generate_journal_entries(sid)
        tax_entries = [e for e in entries if e["account"] == "進項稅額"]
        self.assertEqual(len(tax_entries), 0, "不可扣抵不應有進項稅額")

    def test_multiple_categories(self):
        """多分類 — 借方按分類拆分"""
        import state_manager as sm
        from services.accounting_service import generate_journal_entries, verify_balance

        sid = _create_confirmed_staging(sm)
        _add_items(sm, sid)  # 蔬菜+肉類+蛋豆

        entries = generate_journal_entries(sid)
        debit_entries = [e for e in entries if e["side"] == "debit"]

        # 應有 3 個品類的借方分錄 + 1 個進項稅額
        account_names = [e["account"] for e in debit_entries]
        self.assertIn("進貨—蔬菜類", account_names)
        self.assertIn("進貨—肉類", account_names)
        self.assertIn("進貨—蛋豆類", account_names)

        balance = verify_balance(sid)
        self.assertTrue(balance["balanced"])

    def test_idempotent_generation(self):
        """冪等 — 重複生成不會重複分錄"""
        import state_manager as sm
        from services.accounting_service import generate_journal_entries

        sid = _create_confirmed_staging(sm)
        _add_items(sm, sid)

        entries1 = generate_journal_entries(sid)
        entries2 = generate_journal_entries(sid)  # 再生成一次

        # DB 中只有最新一批
        db_entries = sm.get_journal_entries_by_source("purchase", sid)
        self.assertEqual(len(db_entries), len(entries2))


class TestTrialBalance(unittest.TestCase):
    """Test 2: 試算表 — 全月借方合計＝貸方合計"""

    def setUp(self):
        self.db_path = _setup_db()

    def tearDown(self):
        _teardown_db(self.db_path)

    def test_monthly_trial_balance(self):
        import state_manager as sm
        from services.accounting_service import generate_journal_entries

        # 建立多筆交易
        for supplier, total in [("好鮮水產行", 5000), ("阿明豆腐店", 1200)]:
            sid = _create_confirmed_staging(sm, supplier=supplier, total=total,
                                            subtotal=round(total/1.05),
                                            tax=total - round(total/1.05))
            sm.add_purchase_item(sid, item_name="食材", quantity=1, unit="批",
                                 unit_price=total, amount=round(total/1.05),
                                 category="蔬菜")
            generate_journal_entries(sid)

        trial = sm.get_trial_balance("2026-03")
        total_debit = sum(t["total_debit"] for t in trial)
        total_credit = sum(t["total_credit"] for t in trial)
        self.assertAlmostEqual(total_debit, total_credit, delta=1,
                               msg="全月試算表借貸不平衡")

    def test_journal_summary(self):
        import state_manager as sm
        from services.accounting_service import generate_journal_entries

        sid = _create_confirmed_staging(sm)
        _add_items(sm, sid)
        generate_journal_entries(sid)

        summary = sm.get_journal_summary("2026-03")
        self.assertGreater(summary["count"], 0)
        self.assertTrue(summary["balanced"])


class TestAccountingExcel(unittest.TestCase):
    """Test 3: Excel 帳冊生成（8 頁）"""

    def setUp(self):
        self.db_path = _setup_db()
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        _teardown_db(self.db_path)
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_excel_generation(self):
        import state_manager as sm
        from services import accounting_service as acct

        # 覆蓋輸出目錄
        acct.ACCOUNTING_DIR = self.tmpdir

        sid = _create_confirmed_staging(sm)
        _add_items(sm, sid)
        acct.generate_journal_entries(sid)

        path = acct.generate_accounting_excel("2026-03")
        self.assertIsNotNone(path)
        self.assertTrue(os.path.exists(path))

        # 驗證 Excel 內容 — 8 個工作表
        import openpyxl
        wb = openpyxl.load_workbook(path)
        self.assertIn("進貨日記帳", wb.sheetnames)
        self.assertIn("月度費用彙總", wb.sheetnames)
        self.assertIn("試算表", wb.sheetnames)
        self.assertIn("分錄明細", wb.sheetnames)
        self.assertIn("損益表", wb.sheetnames)
        self.assertIn("資產負債表", wb.sheetnames)
        self.assertIn("總分類帳", wb.sheetnames)

    def test_empty_month(self):
        import state_manager as sm
        from services import accounting_service as acct

        acct.ACCOUNTING_DIR = self.tmpdir

        path = acct.generate_accounting_excel("2026-04")
        self.assertIsNotNone(path)
        self.assertTrue(os.path.exists(path))


class TestProcessAfterArchive(unittest.TestCase):
    """Test 4: 歸檔後自動會計"""

    def setUp(self):
        self.db_path = _setup_db()
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        _teardown_db(self.db_path)
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_process_after_archive(self):
        import state_manager as sm
        from services import accounting_service as acct

        acct.ACCOUNTING_DIR = self.tmpdir

        sid = _create_confirmed_staging(sm)
        _add_items(sm, sid)

        result = acct.process_after_archive(sid)
        self.assertIn("分錄", result)
        self.assertIn("借貸平衡", result)

    def test_process_missing_staging(self):
        from services import accounting_service as acct
        acct.ACCOUNTING_DIR = self.tmpdir

        result = acct.process_after_archive(99999)
        self.assertEqual(result, "")


class TestMonthlyAccounting(unittest.TestCase):
    """Test 5: 月度會計總表"""

    def setUp(self):
        self.db_path = _setup_db()

    def tearDown(self):
        _teardown_db(self.db_path)

    def test_upsert_monthly_accounting(self):
        import state_manager as sm

        mid = sm.upsert_monthly_accounting("2026-03",
                                           total_income=100000,
                                           total_expense=60000,
                                           net_profit=40000)
        self.assertGreater(mid, 0)

        record = sm.get_monthly_accounting("2026-03")
        self.assertIsNotNone(record)
        self.assertEqual(record["total_income"], 100000)
        self.assertEqual(record["total_expense"], 60000)

    def test_update_existing(self):
        import state_manager as sm

        sm.upsert_monthly_accounting("2026-03", total_expense=50000)
        sm.upsert_monthly_accounting("2026-03", total_expense=60000)

        record = sm.get_monthly_accounting("2026-03")
        self.assertEqual(record["total_expense"], 60000)

    def test_monthly_report_text(self):
        import state_manager as sm
        from services.accounting_service import get_monthly_report_text

        sid = _create_confirmed_staging(sm)
        _add_items(sm, sid)
        from services.accounting_service import generate_journal_entries
        generate_journal_entries(sid)

        text = get_monthly_report_text("2026-03")
        self.assertIn("2026-03", text)
        self.assertIn("進貨支出", text)


class TestJournalEntryCRUD(unittest.TestCase):
    """Test 6: 分錄 CRUD 函數"""

    def setUp(self):
        self.db_path = _setup_db()

    def tearDown(self):
        _teardown_db(self.db_path)

    def test_add_and_get(self):
        import state_manager as sm

        eid = sm.add_journal_entry(
            entry_date="2026-03-15", year_month="2026-03",
            source_type="purchase", source_id=1,
            description="進貨", account_code="5110",
            account_name="進貨—蔬菜類", debit=1000, credit=0,
        )
        self.assertGreater(eid, 0)

        entries = sm.get_journal_entries("2026-03")
        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0]["debit"], 1000)

    def test_get_by_source(self):
        import state_manager as sm

        sm.add_journal_entry("2026-03-15", "2026-03", "purchase", 1,
                             "進貨A", "5110", "進貨", debit=500)
        sm.add_journal_entry("2026-03-15", "2026-03", "purchase", 1,
                             "付款A", "1100", "現金", credit=500)
        sm.add_journal_entry("2026-03-16", "2026-03", "purchase", 2,
                             "進貨B", "5110", "進貨", debit=300)

        entries_1 = sm.get_journal_entries_by_source("purchase", 1)
        self.assertEqual(len(entries_1), 2)

        entries_2 = sm.get_journal_entries_by_source("purchase", 2)
        self.assertEqual(len(entries_2), 1)

    def test_delete_by_source(self):
        import state_manager as sm

        sm.add_journal_entry("2026-03-15", "2026-03", "purchase", 1,
                             "進貨", "5110", "進貨", debit=500)
        sm.add_journal_entry("2026-03-15", "2026-03", "purchase", 1,
                             "付款", "1100", "現金", credit=500)

        sm.delete_journal_entries_by_source("purchase", 1)
        entries = sm.get_journal_entries_by_source("purchase", 1)
        self.assertEqual(len(entries), 0)


class TestIncomeJournal(unittest.TestCase):
    """Test 7: 收入分錄 — 含銷項稅額"""

    def setUp(self):
        self.db_path = _setup_db()

    def tearDown(self):
        _teardown_db(self.db_path)

    def test_income_journal_with_tax(self):
        """收入 → 分錄含銷項稅額"""
        import state_manager as sm
        from services.accounting_service import generate_income_journal_entries

        income_id = sm.add_income("2026-03", 105000, "團膳收入", "現金", "2026-03-31")
        entries = generate_income_journal_entries(income_id)
        self.assertGreater(len(entries), 0)

        # 應有銷項稅額
        tax_entries = [e for e in entries if e["account"] == "銷項稅額"]
        self.assertEqual(len(tax_entries), 1)
        self.assertEqual(tax_entries[0]["amount"], 5000)

        # 營業收入未稅
        revenue = [e for e in entries if e["account"] == "營業收入"]
        self.assertEqual(len(revenue), 1)
        self.assertEqual(revenue[0]["amount"], 100000)

        # 借貸平衡
        total_d = sum(e["amount"] for e in entries if e["side"] == "debit")
        total_c = sum(e["amount"] for e in entries if e["side"] == "credit")
        self.assertEqual(total_d, total_c)


class TestPayrollJournal(unittest.TestCase):
    """Test 8: 薪資分錄"""

    def setUp(self):
        self.db_path = _setup_db()

    def tearDown(self):
        _teardown_db(self.db_path)

    def test_payroll_journal(self):
        """薪資 → 分錄（含勞健保/稅）"""
        import state_manager as sm
        from services.accounting_service import generate_payroll_journal_entries

        # 建立員工和薪資
        emp_id = sm.add_employee(name="王小明", base_salary=30000,
                                 meal_allowance=2400)
        sm.add_payroll(
            employee_id=emp_id, year_month="2026-03",
            base_salary=30000, meal_allowance=2400,
            overtime_pay=0, bonus=0, gross_salary=32400,
            labor_insurance=686, health_insurance=442,
            pension_self=0, income_tax=0,
            net_salary=31272, status="confirmed",
        )

        entries = generate_payroll_journal_entries("2026-03")
        self.assertGreater(len(entries), 0)

        # 應有薪資費用借方
        salary_d = [e for e in entries if e["account"] == "薪資費用" and e["side"] == "debit"]
        self.assertGreater(len(salary_d), 0)

        # 應有現金貸方
        cash_c = [e for e in entries if e["account"] == "現金" and e["side"] == "credit"]
        self.assertGreater(len(cash_c), 0)

        # 借貸平衡
        total_d = sum(e["amount"] for e in entries if e["side"] == "debit")
        total_c = sum(e["amount"] for e in entries if e["side"] == "credit")
        self.assertAlmostEqual(total_d, total_c, delta=1)


class TestPeriodEndClosing(unittest.TestCase):
    """Test 9: 期末結帳"""

    def setUp(self):
        self.db_path = _setup_db()

    def tearDown(self):
        _teardown_db(self.db_path)

    def test_closing(self):
        """結帳 → 收入費用結轉到本期損益"""
        import state_manager as sm
        from services.accounting_service import (
            generate_journal_entries, generate_income_journal_entries,
            perform_period_end_closing,
        )

        # 進貨
        sid = _create_confirmed_staging(sm)
        _add_items(sm, sid)
        generate_journal_entries(sid)

        # 收入
        inc_id = sm.add_income("2026-03", 50000, "團膳收入", "現金", "2026-03-31")
        generate_income_journal_entries(inc_id)

        # 結帳
        result = perform_period_end_closing("2026-03")
        self.assertEqual(result["status"], "closed")
        self.assertGreater(result["total_revenue"], 0)

        # 月度表標記已結帳
        acct = sm.get_monthly_accounting("2026-03")
        self.assertIsNotNone(acct)
        self.assertEqual(acct["is_closed"], 1)

    def test_double_closing(self):
        """重複結帳 → 返回 already_closed"""
        import state_manager as sm
        from services.accounting_service import perform_period_end_closing

        sm.upsert_monthly_accounting("2026-03", is_closed=1)
        result = perform_period_end_closing("2026-03")
        self.assertEqual(result["status"], "already_closed")


class TestFinancialStatements(unittest.TestCase):
    """Test 10: 財務報表"""

    def setUp(self):
        self.db_path = _setup_db()

    def tearDown(self):
        _teardown_db(self.db_path)

    def test_income_statement(self):
        """損益表 — 收入 - 成本 - 費用 = 淨利"""
        import state_manager as sm
        from services.accounting_service import (
            generate_journal_entries, generate_income_journal_entries,
            generate_income_statement,
        )

        sid = _create_confirmed_staging(sm)
        _add_items(sm, sid)
        generate_journal_entries(sid)

        inc_id = sm.add_income("2026-03", 100000, "團膳收入", "現金", "2026-03-31")
        generate_income_journal_entries(inc_id)

        pl = generate_income_statement("2026-03")
        self.assertGreater(pl["revenue"], 0)
        self.assertGreater(pl["cost"], 0)
        self.assertEqual(pl["net_income"], pl["revenue"] - pl["cost"] - pl["expense"])

    def test_balance_sheet(self):
        """資產負債表 — 有資料就能跑"""
        import state_manager as sm
        from services.accounting_service import (
            generate_journal_entries, generate_balance_sheet,
        )

        sid = _create_confirmed_staging(sm)
        _add_items(sm, sid)
        generate_journal_entries(sid)

        bs = generate_balance_sheet("2026-03")
        # 只有進貨分錄時，資產（現金為負）和成本分錄都在
        self.assertIsNotNone(bs["year_month"])


class TestVATSummary(unittest.TestCase):
    """Test 11: 營業稅摘要"""

    def setUp(self):
        self.db_path = _setup_db()

    def tearDown(self):
        _teardown_db(self.db_path)

    def test_vat_summary(self):
        import state_manager as sm
        from services.accounting_service import (
            generate_journal_entries, generate_income_journal_entries,
            get_vat_summary,
        )

        # 進項
        sid = _create_confirmed_staging(sm, total=10500, subtotal=10000, tax=500)
        _add_items(sm, sid)
        generate_journal_entries(sid)

        # 銷項
        inc_id = sm.add_income("2026-03", 21000, "收入", "現金", "2026-03-31")
        generate_income_journal_entries(inc_id)

        vat = get_vat_summary("2026-03")
        self.assertGreater(vat["input_tax"], 0)
        self.assertGreater(vat["output_tax"], 0)
        self.assertIn(vat["status"], ["應繳", "留抵"])


class TestChartOfAccounts(unittest.TestCase):
    """Test 12: 會計科目表"""

    def setUp(self):
        self.db_path = _setup_db()

    def tearDown(self):
        _teardown_db(self.db_path)

    def test_full_chart(self):
        import state_manager as sm

        coa = sm.get_chart_of_accounts()
        self.assertGreater(len(coa), 40, "應有 40+ 科目")

        # 各類別都有
        categories = set(a["category"] for a in coa)
        self.assertIn("asset", categories)
        self.assertIn("liability", categories)
        self.assertIn("equity", categories)
        self.assertIn("revenue", categories)
        self.assertIn("cost", categories)
        self.assertIn("expense", categories)

    def test_filter_by_category(self):
        import state_manager as sm

        assets = sm.get_chart_of_accounts("asset")
        self.assertGreater(len(assets), 0)
        for a in assets:
            self.assertEqual(a["category"], "asset")


class TestFixedAssets(unittest.TestCase):
    """Test 13: 固定資產"""

    def setUp(self):
        self.db_path = _setup_db()

    def tearDown(self):
        _teardown_db(self.db_path)

    def test_add_and_depreciate(self):
        import state_manager as sm
        from services.accounting_service import generate_depreciation_entries

        aid = sm.add_fixed_asset(
            name="冷藏設備", category="機器設備",
            purchase_date="2026-01-01", cost=120000,
            useful_life_months=60, salvage_value=0,
        )
        self.assertGreater(aid, 0)

        entries = generate_depreciation_entries("2026-03")
        self.assertEqual(len(entries), 2)  # 借折舊+貸累計

        # 月折舊 = 120000 / 60 = 2000
        dep_entry = [e for e in entries if e["account"] == "折舊費用"][0]
        self.assertEqual(dep_entry["amount"], 2000)

        # 累計折舊已更新
        asset = sm.get_fixed_assets()[0]
        self.assertEqual(asset["accumulated_depreciation"], 2000)


class TestGeneralLedger(unittest.TestCase):
    """Test 14: 總分類帳"""

    def setUp(self):
        self.db_path = _setup_db()

    def tearDown(self):
        _teardown_db(self.db_path)

    def test_general_ledger(self):
        import state_manager as sm
        from services.accounting_service import generate_journal_entries

        sid = _create_confirmed_staging(sm)
        _add_items(sm, sid)
        generate_journal_entries(sid)

        # 全科目
        ledger = sm.get_general_ledger("2026-03")
        self.assertGreater(len(ledger), 0)

        # 指定科目
        cash_ledger = sm.get_general_ledger("2026-03", "1100")
        self.assertGreater(len(cash_ledger), 0)
        for e in cash_ledger:
            self.assertEqual(e["account_code"], "1100")


class TestTrainingDocument(unittest.TestCase):
    """Test 15: 教育訓練文件生成"""

    def setUp(self):
        self.db_path = _setup_db()
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        _teardown_db(self.db_path)
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_generate_training_doc(self):
        from services import accounting_service as acct

        acct.ACCOUNTING_DIR = self.tmpdir
        path = acct.generate_training_document(self.tmpdir)
        self.assertTrue(os.path.exists(path))
        self.assertTrue(path.endswith(".docx"))


if __name__ == "__main__":
    unittest.main()
